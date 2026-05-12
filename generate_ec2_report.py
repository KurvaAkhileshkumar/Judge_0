#!/usr/bin/env python3
"""
generate_ec2_report.py
──────────────────────
Generates a rich, colour-coded Excel workbook from one or more
load_test_report.json files produced by the Go load tester (main.go).

Usage:
    # Single report
    python generate_ec2_report.py [report.json] [--metrics metrics.jsonl] [--out report.xlsx]

    # Multi-run scenario comparison (auto-detect all load_test_report*.json)
    python generate_ec2_report.py report.json --compare

    # Or pass explicit comparison files
    python generate_ec2_report.py report.json --compare file2.json file3.json ...

Arguments:
    report.json       — Primary JSON report (default: load_test_report.json)
    --metrics FILE    — JSONL file from collect_ec2_metrics.py (optional)
    --out FILE        — Output .xlsx path (default: primary report stem + .xlsx)
    --workers N       — Judge0 worker containers (default: 3)
    --runners N       — MAX_RUNNERS per worker (default: 2)
    --compare [FILE …]— Add Sheet 4 multi-run comparison; if no FILEs given,
                        auto-detects all load_test_report*.json in the same directory

Sheets produced:
  Sheet 1 — System Info
      EC2 instance spec, Docker service layout (with direct-mode variant),
      Judge0 conf, Flask/gunicorn conf, grading worker, PostgreSQL, Redis,
      load-test run config incl. concurrency mode and result delivery method.

  Sheet 2 — Load Test Metrics
      Concurrency mode analysis (flat-concurrent/ramp/pipelined-batch),
      KPI overview, latency percentiles (P50–P99.9) with severity colouring,
      throughput, score distribution (0%/partial/100%), submission outcomes,
      idempotency cache analysis (warm-cache scenario), status distribution,
      per-problem pass rates, solution-type breakdown, batch analysis,
      and (if --metrics provided) live system metrics time-series with
      conditional colouring on CPU/RAM/queue-depth cells.

  Sheet 3 — User Test Case Results
      Per-user aggregated marks summary table at the top, then a detailed
      per-submission table: User ID, Problem, Solution, Harness Status,
      Score, Total TCs, Marks %, Partial flag, Latency, Error notes, and
      per-TC columns (Status | Got | Expected) up to the widest submission.

  Sheet 4 — Scenario Comparison (only with --compare)
      Side-by-side comparison of multiple load test runs: users, concurrency
      mode, ramp, throughput, latency percentiles, pass/fail/error rates,
      rate-limited count, duplicate count, peak in-flight, and score
      distribution — one column per report file.
"""

import argparse
import datetime
import json
import math
import sys
from pathlib import Path
from datetime import timezone as _tz

_UTC = _tz.utc

def _now_utc() -> str:
    return datetime.datetime.now(_UTC).strftime("%Y-%m-%d %H:%M:%S")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed.  Run: pip install -r requirements-report.txt")
    sys.exit(1)


# ── Colour palette ────────────────────────────────────────────────────────────

_C = {
    # Header tiers
    "hdr_dark":      "1F4E79",   # dark navy — sheet/section titles
    "hdr_mid":       "2E75B6",   # medium blue — sub-section titles
    "hdr_light":     "BDD7EE",   # light blue — column headers
    # Section dividers
    "section":       "D6E4F0",
    # Outcome colours
    "pass":          "C6EFCE",   # green
    "pass_fg":       "276221",
    "fail":          "FFCCCC",   # pink-red
    "fail_fg":       "9C0006",
    "tle":           "FFEB9C",   # yellow
    "tle_fg":        "9C6500",
    "error":         "FFD966",   # amber
    "error_fg":      "7D5000",
    "blocked":       "D9D9D9",   # grey
    "blocked_fg":    "404040",
    "rate_limited":  "F4B942",   # orange
    "rate_fg":       "6D3B00",
    "sys_error":     "FF6666",   # red
    "sys_fg":        "660000",
    "judge0_err":    "FF6666",
    "judge0_fg":     "660000",
    "unknown":       "F2F2F2",
    # Misc
    "white":         "FFFFFF",
    "alt_row":       "EBF3FB",
    "light_grey":    "F2F2F2",
    "partial":       "FFD966",   # amber for partial marks
    "partial_fg":    "7D5000",
    # CPU / RAM severity
    "cpu_ok":        "C6EFCE",
    "cpu_warn":      "FFEB9C",
    "cpu_crit":      "FFCCCC",
}

# Map harness status → (bg, fg)
_STATUS_STYLE = {
    "PASS":         (_C["pass"],        _C["pass_fg"]),
    "FAIL":         (_C["fail"],        _C["fail_fg"]),
    "TLE":          (_C["tle"],         _C["tle_fg"]),
    "ERROR":        (_C["error"],       _C["error_fg"]),
    "BLOCKED":      (_C["blocked"],     _C["blocked_fg"]),
    "RATE_LIMITED": (_C["rate_limited"],_C["rate_fg"]),
    "SYSTEM_ERROR": (_C["sys_error"],   _C["sys_fg"]),
    "JUDGE0_ERROR": (_C["judge0_err"],  _C["judge0_fg"]),
    "UNKNOWN":      (_C["unknown"],     "000000"),
    "MISSING":      (_C["unknown"],     "000000"),
}


# ── Low-level cell helpers ────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = "000000", size: int = 11) -> Font:
    return Font(bold=bold, color=color, size=size)


def _thin_border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center_align(wrap: bool = True) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left_align(wrap: bool = True) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def _style(cell, bg: str = _C["white"], fg: str = "000000",
           bold: bool = False, size: int = 11,
           halign: str = "left", wrap: bool = True):
    cell.fill = _fill(bg)
    cell.font = _font(bold=bold, color=fg, size=size)
    cell.border = _thin_border()
    cell.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)


# ── Mid-level row helpers ─────────────────────────────────────────────────────

def _write_title(ws, row: int, text: str, ncols: int, fontsize: int = 13):
    """Dark-navy full-width title row."""
    cell = ws.cell(row=row, column=1, value=text)
    _style(cell, bg=_C["hdr_dark"], fg="FFFFFF", bold=True, size=fontsize, halign="center")
    ws.row_dimensions[row].height = max(22, fontsize + 10)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=ncols)


def _write_section(ws, row: int, text: str, ncols: int):
    """Blue section-divider row spanning all columns."""
    cell = ws.cell(row=row, column=1, value=f"  {text}")
    _style(cell, bg=_C["section"], fg=_C["hdr_dark"], bold=True)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=ncols)


def _write_col_header(ws, row: int, col: int, text: str, bg: str = _C["hdr_mid"],
                      fg: str = "FFFFFF", colspan: int = 1):
    """Single column-header cell (optionally merged)."""
    cell = ws.cell(row=row, column=col, value=text)
    _style(cell, bg=bg, fg=fg, bold=True, halign="center")
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col + colspan - 1)
    return cell


def _write_kv(ws, row: int, key: str, value, alt: bool = False):
    """Two-column key: value row."""
    bg = _C["alt_row"] if alt else _C["white"]
    c1 = ws.cell(row=row, column=1, value=key)
    _style(c1, bg=bg, bold=True)
    c2 = ws.cell(row=row, column=2, value=value)
    _style(c2, bg=bg)


def _write_data_row(ws, row: int, values: list, alt: bool = False,
                    col_start: int = 1):
    """Plain data row, alternating background."""
    bg = _C["alt_row"] if alt else _C["white"]
    for ci, val in enumerate(values, col_start):
        cell = ws.cell(row=row, column=ci, value=val)
        _style(cell, bg=bg)


def _set_col_widths(ws, widths: list, col_start: int = 1):
    for i, w in enumerate(widths, col_start):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Derived helpers ───────────────────────────────────────────────────────────

def _throughput(summary: dict) -> float:
    return summary.get("throughput_per_sec", 0.0) or 0.0


def _duration(summary: dict) -> float:
    completed = summary.get("completed", 0) or 1
    tput = _throughput(summary)
    return (completed / tput) if tput > 0 else 0.0


def _pct(numerator, denominator):
    if not denominator:
        return "–"
    return f"{numerator / denominator * 100:.1f}%"


# ── Scenario / concurrency helpers ────────────────────────────────────────────

def _concurrency_label(cfg: dict) -> str:
    """Human-readable concurrency mode derived from config."""
    if cfg.get("_cumulative"):
        return (f"Multiple modes — {cfg.get('_n_runs', 'N')} scenarios aggregated "
                f"(flat / ramp / pipelined-batch — see Sheet 4 for per-run breakdown)")
    batch = cfg.get("batch_size", 0) or 0
    ramp  = cfg.get("ramp_up_sec", 0) or 0
    users = cfg.get("users", 0) or 0
    if batch > 0:
        n_batches = math.ceil(users / batch)
        return (f"Pipelined Batch  ({n_batches} batches × {batch} users/batch — "
                f"all batches fire simultaneously)")
    elif ramp > 0:
        cadence_ms = ramp * 1000 / max(users - 1, 1)
        return (f"Flat Concurrent + Ramp-up  ({ramp}s ramp → "
                f"1 user every ~{cadence_ms:.0f} ms)")
    else:
        return "Flat Concurrent  — all users fire simultaneously (no ramp-up)"


def _ramp_cadence_ms(cfg: dict) -> float:
    ramp  = cfg.get("ramp_up_sec", 0) or 0
    users = cfg.get("users", 1) or 1
    return (ramp * 1000 / max(users - 1, 1)) if ramp > 0 and users > 1 else 0.0


def _result_delivery_label(cfg: dict) -> str:
    mode = cfg.get("mode", "")
    if mode == "flask_stack":
        return "SSE  /results/stream/<ticket_id>  (Flask stack)"
    cb = cfg.get("callback_port", 0) or 0
    if cb:
        return f"Callback  (Judge0 → HTTP POST → port {cb})"
    return f"Polling  (interval {cfg.get('poll_interval', '400ms')})"


_PASCAL_TO_SNAKE = {
    "UserID":        "user_id",
    "ProblemID":     "problem_id",
    "SolutionID":    "solution_id",
    "SolutionType":  "solution_type",
    "HarnessStatus": "harness_status",
    "SourceCode":    "source_code",
    "TCDetails":     "tc_details",
    "LatencyMs":     "latency_ms",
    "Judge0Status":  "judge0_status",
    "Error":         "error",
    "GlobalTLE":     "global_tle",
    "Score":         "score",
    "TotalTCs":      "total_tcs",
}

def _normalise_result(r: dict) -> dict:
    """Return a copy of a result dict with PascalCase keys mapped to snake_case.
    Old JSON files (pre-tag) used Go's default PascalCase; new files use proper
    json tags and already have snake_case.  This makes the report handle both."""
    out = {}
    for k, v in r.items():
        out[_PASCAL_TO_SNAKE.get(k, k)] = v
    return out


def _score_distribution(all_subs: list) -> tuple:
    """Returns (full_marks_count, partial_count, zero_count)."""
    full = partial = zero = 0
    for s in all_subs:
        score = s.get("score", 0)
        total = s.get("total_tcs", 0) or 1
        if score >= total:
            full += 1
        elif score > 0:
            partial += 1
        else:
            zero += 1
    return full, partial, zero


def _scenario_label(cfg: dict) -> str:
    """Short one-line label for a report file used in Sheet 4."""
    mode  = cfg.get("mode", "direct")
    users = cfg.get("users", 0)
    ramp  = cfg.get("ramp_up_sec", 0) or 0
    batch = cfg.get("batch_size", 0) or 0
    parts = [f"{users}u"]
    if batch > 0:
        parts.append(f"batch{batch}")
    elif ramp > 0:
        parts.append(f"ramp{ramp}s")
    else:
        parts.append("concurrent")
    parts.append("flask" if mode == "flask_stack" else "direct")
    return "_".join(parts)


# ─────────────────────────────────────────────────────────────────────────────
#  Sheet 1 — System Info
# ─────────────────────────────────────────────────────────────────────────────

def build_sheet1(wb: openpyxl.Workbook, report: dict, workers: int, runners: int):
    ws = wb.active
    ws.title = "System Info"
    ws.sheet_view.showGridLines = False

    cfg     = report.get("config", {})
    summary = report.get("summary", {})
    NCOLS   = 2

    # ── Title ─────────────────────────────────────────────────────────────────
    _write_title(ws, 1, "EC2 LOAD TEST — SYSTEM CONFIGURATION", NCOLS, fontsize=14)
    ws.cell(row=2, column=1, value=(
        f"Generated: {_now_utc()} UTC"
    ))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    c2 = ws.cell(row=2, column=1)
    _style(c2, bg=_C["light_grey"], halign="center", bold=False, size=10)

    row  = 3
    alt  = False

    def section(title: str):
        nonlocal row, alt
        row += 1
        _write_section(ws, row, title, NCOLS)
        row += 1
        alt = False

    def kv(key: str, value):
        nonlocal row, alt
        _write_kv(ws, row, key, value, alt)
        row += 1
        alt = not alt

    # ── EC2 Instance ──────────────────────────────────────────────────────────
    section("EC2 Instance")
    kv("Instance Type",               "c5.xlarge")
    kv("vCPUs",                       4)
    kv("RAM",                         "8 GB")
    kv("Storage",                     "EBS (gp2 / gp3)")
    kv("OS",                          "Ubuntu 22.04.5 LTS")
    kv("Kernel",                      "Linux 5.15+ x86_64")
    kv("Region",                      "ap-south-1 (Mumbai)")
    kv("Network Performance",         "Up to 10 Gbps")

    # ── Docker Compose Services ───────────────────────────────────────────────
    section("Docker Services Layout")
    kv("Judge0 API Server (server)",  "judge0/judge0:1.13.1  —  port 2358")
    kv("Judge0 Worker containers",    f"{workers} containers  (--scale workers={workers})")
    kv("MAX_RUNNERS per worker",      runners)
    kv("Total concurrent sandboxes",  f"{workers} × {runners} = {workers * runners}")
    kv("Flask API (api)",             "gunicorn+gevent  —  port 5001→5000")
    kv("Grading Worker (grading_worker)", "worker_async.py  —  WORKER_CONCURRENCY=24")
    kv("Reconciler",                  "reconciler.py  —  scan every 60s")
    kv("PostgreSQL",                  "postgres:13  —  mem_limit=400m  reservation=128m")
    kv("Redis",                       "Latest  —  mem_limit=512m  maxmemory=512mb")
    kv("Memory budget (workers=3)",
       "OS≈1 GB + PG≤0.4 GB + Redis≤0.5 GB + server≈0.4 GB + 3×workers≤3 GB + Flask≈0.3 GB ≈ 5.6 GB")

    # ── Judge0 Configuration ──────────────────────────────────────────────────
    section("Judge0 Configuration (judge0.ec2.conf)")
    kv("MAX_RUNNERS (isolate sandboxes/worker)", runners)
    kv("MAX_QUEUE_SIZE (Resque)",     200)
    kv("CPU_TIME_LIMIT (default)",    "5 s")
    kv("WALL_TIME_LIMIT (default)",   "10 s")
    kv("MEMORY_LIMIT (default)",      "256 000 KB  (256 MB)")
    kv("MAX_MEMORY_LIMIT",            "4 194 304 KB  (4 GB)")
    kv("SOURCE_CODE_SIZE_LIMIT",      "524 288 B  (512 KB)")
    kv("MAX_PROCESSES_AND_OR_THREADS", 220)
    kv("MAX_MAX_PROCESSES_AND_OR_THREADS", 500)
    kv("OPEN_FILES_LIMIT",            65536)
    kv("INTERVAL (Resque idle poll)", "0.1 s")
    kv("RAILS_ENV / RACK_ENV",        "production")
    kv("JUDGE0_TELEMETRY_ENABLE",     False)

    # ── Judge0 API Server (Puma) ──────────────────────────────────────────────
    section("Judge0 API Server (Puma / Rails)")
    kv("WEB_CONCURRENCY (Puma processes)",   2)
    kv("RAILS_MAX_THREADS (threads/process)", 16)
    kv("HTTP slots",                          "2 × 16 = 32 concurrent HTTP requests")
    kv("net.core.somaxconn",                 4096)
    kv("net.ipv4.tcp_max_syn_backlog",       4096)

    # ── Flask API ─────────────────────────────────────────────────────────────
    section("Flask Grading API (gunicorn + gevent)")
    kv("Gunicorn worker class",  "gevent")
    kv("Gunicorn workers",       2)
    kv("Worker connections",     500)
    kv("Bind",                   "0.0.0.0:5000  (host port 5001)")
    kv("Timeout",                "1900 s")
    kv("Result delivery",        "SSE  /results/stream/<ticket_id>")
    kv("MAX_QUEUE_DEPTH",        5000)
    kv("SSE_TIMEOUT_S",          1800)
    kv("Idempotency TTL",        "2 h  (SHA-256 key stored in Redis)")
    kv("Max request body",       "1 MB  (rejects oversized payloads before parsing)")

    # ── Grading Worker ────────────────────────────────────────────────────────
    section("Grading Worker (worker_async.py)")
    kv("WORKER_CONCURRENCY",     24)
    kv("MAX_RETRY_COUNT",        3)
    kv("Retry strategy",         "Natural queue drain  —  no sleep between retries")
    kv("Result TTL",             "2 h in Redis")
    kv("Callback mode",          "Judge0 POSTs back to worker's callback server")

    # ── PostgreSQL ────────────────────────────────────────────────────────────
    section("PostgreSQL 13")
    kv("max_connections",        200)
    kv("shared_buffers",         "128 MB")
    kv("effective_cache_size",   "1 GB  (planner hint, no allocation)")
    kv("wal_buffers",            "16 MB")
    kv("work_mem",               "4 MB")
    kv("checkpoint_completion_target", 0.9)
    kv("mem_limit",              "400 m")
    kv("mem_reservation",        "128 m")

    # ── Redis ─────────────────────────────────────────────────────────────────
    section("Redis")
    kv("maxmemory",              "512 mb")
    kv("mem_limit",              "512 m")
    kv("Queue keys",             "judge0:jobs:normal  |  judge0:jobs:retry")
    kv("Result key prefix",      "judge0:result:<ticket_id>")
    kv("Pub/Sub prefix",         "judge0:notify:<ticket_id>")
    kv("Idempotency prefix",     "judge0:idem:<sha256>")

    # ── Load Test Configuration ───────────────────────────────────────────────
    section("Load Test Configuration")
    kv("Mode",                   cfg.get("mode", "N/A"))
    kv("Target URL",             cfg.get("target", "N/A"))
    kv("Virtual Users",          cfg.get("users", "N/A"))
    kv("Ramp-up Duration",       f"{cfg.get('ramp_up_sec', 0)} s")
    kv("Batch Size",             cfg.get("batch_size") or "0  (flat concurrent — no batching)")
    kv("Concurrency Mode",       _concurrency_label(cfg))
    if _ramp_cadence_ms(cfg) > 0:
        kv("Ramp-up Cadence",    f"1 virtual user launched every ~{_ramp_cadence_ms(cfg):.0f} ms")
    kv("Result Delivery",        _result_delivery_label(cfg))
    kv("Dry Run",                cfg.get("dry_run", False))
    kv("Test Duration",          f"{_duration(summary):.1f} s")
    kv("Go Load Tester",         "main.go  —  Judge0 Harness Load Tester")

    is_flask = cfg.get("mode") == "flask_stack"
    if not is_flask:
        section("Direct Judge0 Mode Notes")
        kv("Result delivery",    "Polling (default) or Callback (-callback-port N)")
        kv("No admission control", "No rate limiting / idempotency in direct mode")
        kv("No Redis queue",     "Harnesses submitted directly to Judge0 /submissions")
        kv("Security check",     "Go-side quick AST pre-check (syntax errors blocked locally)")

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 55
    ws.freeze_panes = ws.cell(row=3, column=1)


# ─────────────────────────────────────────────────────────────────────────────
#  Sheet 2 — Load Test Metrics
# ─────────────────────────────────────────────────────────────────────────────

def build_sheet2(wb: openpyxl.Workbook, report: dict, metrics_data: list):
    ws = wb.create_sheet("Load Test Metrics")
    ws.sheet_view.showGridLines = False

    cfg           = report.get("config", {})
    summary       = report.get("summary", {})
    status_counts = report.get("status_counts", {})
    problem_counts= report.get("problem_counts", {})
    type_counts   = report.get("type_counts", {})
    batches       = report.get("batches", [])
    per_problem   = report.get("per_problem_report", [])
    all_results   = [_normalise_result(r) for r in report.get("results", [])]

    NCOLS = 6
    row   = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    _write_title(ws, row, "EC2 LOAD TEST — PERFORMANCE METRICS", NCOLS, fontsize=14)
    row += 1
    ts_cell = ws.cell(row=row, column=1,
                      value=f"Generated: {_now_utc()} UTC"
                            f"   |   Report: {cfg.get('target','N/A')}"
                            f"   |   Users: {cfg.get('users','N/A')}"
                            f"   |   Mode: {cfg.get('mode','N/A')}")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    _style(ts_cell, bg=_C["light_grey"], halign="left", size=10)
    row += 2

    # ── Compute derived stats ─────────────────────────────────────────────────
    lats = sorted(r.get("latency_ms", 0) for r in all_results if "latency_ms" in r)

    def _pctile(p: float) -> int:
        if not lats:
            return summary.get(f"latency_p{int(p)}_ms", 0)
        # Use same nearest-rank formula as Go: idx = int((n-1) * p / 100)
        idx = int((len(lats) - 1) * p / 100.0)
        return lats[idx]

    avg_lat = (sum(lats) / len(lats)) if lats else 0.0
    min_lat = lats[0]  if lats else 0
    max_lat = lats[-1] if lats else 0
    # Compute all percentiles consistently via _pctile (Go-compatible formula)
    p50     = _pctile(50)
    p95     = _pctile(95)
    p99     = _pctile(99)

    completed   = summary.get("completed", 0)
    total_subs  = summary.get("total_submissions", completed) or 1
    throughput  = _throughput(summary)
    duration    = _duration(summary)

    # ── Section: KPI Overview (2 columns of key-value pairs side by side) ─────
    _write_section(ws, row, "KPI OVERVIEW", NCOLS)
    row += 1

    # Header row
    for ci, h in enumerate(["Metric", "Value", "", "Metric", "Value", ""], 1):
        bg = _C["hdr_light"] if h else _C["white"]
        c = ws.cell(row=row, column=ci, value=h)
        _style(c, bg=bg, bold=True, halign="center", fg=_C["hdr_dark"])
    row += 1

    kpis_left = [
        ("Mode",                    cfg.get("mode", "N/A")),
        ("Target URL",              cfg.get("target", "N/A")),
        ("Virtual Users",           cfg.get("users", "N/A")),
        ("Ramp-up Duration",        f"{cfg.get('ramp_up_sec', 0)} s"),
        ("Batch Size",              cfg.get("batch_size") or "flat concurrent"),
        ("Test Duration",           f"{duration:.1f} s"),
        ("Total Finalized",         completed),
        ("PASS Count",              status_counts.get("PASS", 0)),
        ("Throughput",              f"{throughput:.2f} req/s"),
        ("Peak In-Flight Jobs",     summary.get("peak_in_flight_jobs", 0)),
    ]
    kpis_right = [
        ("Errors (network/Judge0)", summary.get("errors", 0)),
        ("Blocked (security check)",summary.get("blocked_by_security", 0)),
        ("Rate Limited (HTTP 429)", summary.get("rate_limited_429", 0)),
        ("Duplicate (idempotent)",  summary.get("duplicate_idempotent", 0)),
        ("System Errors (infra)",   summary.get("system_errors", 0)),
        ("Harness Jobs (actual)",   summary.get("harness_jobs", 0)),
        ("Batch-Equivalent Jobs",   summary.get("batch_equivalent", 0)),
        ("Queue Reduction Factor",  summary.get("queue_reduction", "N/A")),
        ("Max Queue Depth (Redis)", summary.get("max_queue_depth", "N/A")),
        ("Result Delivery",         "SSE  /results/stream"),
    ]

    for i, ((lk, lv), (rk, rv)) in enumerate(zip(kpis_left, kpis_right)):
        bg = _C["alt_row"] if i % 2 else _C["white"]
        for ci, val in [(1, lk), (2, lv), (3, ""), (4, rk), (5, rv), (6, "")]:
            c = ws.cell(row=row, column=ci, value=val)
            _style(c, bg=bg, bold=(ci in (1, 4)))
        row += 1
    row += 1

    # ── Section: Latency Analysis ─────────────────────────────────────────────
    _write_section(ws, row, "LATENCY ANALYSIS", NCOLS)
    row += 1

    for ci, h in enumerate(["Percentile", "Latency (ms)", "", "Metric", "Value", ""], 1):
        bg = _C["hdr_light"] if h else _C["white"]
        c = ws.cell(row=row, column=ci, value=h)
        _style(c, bg=bg, bold=True, halign="center", fg=_C["hdr_dark"])
    row += 1

    lat_left = [
        ("P50  (median)",    p50),
        ("P75",              _pctile(75)),
        ("P90",              _pctile(90)),
        ("P95",              p95),
        ("P99",              p99),
        ("P99.9",            _pctile(99.9)),
        ("Min",              min_lat),
        ("Max",              max_lat),
        ("Average",          round(avg_lat, 1)),
    ]
    lat_right = [
        ("Throughput (req/s)",      f"{throughput:.2f}"),
        ("Total Finalized",         completed),
        ("Test Duration",           f"{duration:.1f} s"),
        ("Harness Jobs (actual)",   summary.get("harness_jobs", 0)),
        ("Batch-Equivalent Jobs",   summary.get("batch_equivalent", 0)),
        ("Queue Reduction",         summary.get("queue_reduction", "N/A")),
        ("Peak In-Flight",          summary.get("peak_in_flight_jobs", 0)),
        ("Workers × Runners",       f"{report.get('_workers',3)} × {report.get('_runners',2)} = {report.get('_workers',3)*report.get('_runners',2)} sandboxes"),
        ("WORKER_CONCURRENCY",      24),
    ]

    nrows = max(len(lat_left), len(lat_right))
    for i in range(nrows):
        lk, lv = lat_left[i]  if i < len(lat_left)  else ("", "")
        rk, rv = lat_right[i] if i < len(lat_right) else ("", "")
        bg = _C["alt_row"] if i % 2 else _C["white"]
        # colour latency values by severity
        lat_bg = bg
        if isinstance(lv, (int, float)) and lv > 0:
            if   lv > 10000: lat_bg = _C["fail"]
            elif lv > 5000:  lat_bg = _C["tle"]
            elif lv < 1000:  lat_bg = _C["pass"]
        for ci, val in [(1, lk), (2, lv), (3, ""), (4, rk), (5, rv), (6, "")]:
            c = ws.cell(row=row, column=ci, value=val)
            _style(c, bg=lat_bg if ci == 2 else bg, bold=(ci in (1, 4)))
        row += 1
    row += 1

    # ── Section: Submission Outcomes ──────────────────────────────────────────
    _write_section(ws, row, "SUBMISSION OUTCOMES", NCOLS)
    row += 1

    _write_col_header(ws, row, 1, "Outcome",     bg=_C["hdr_light"], fg=_C["hdr_dark"])
    _write_col_header(ws, row, 2, "Count",       bg=_C["hdr_light"], fg=_C["hdr_dark"])
    _write_col_header(ws, row, 3, "% of Total",  bg=_C["hdr_light"], fg=_C["hdr_dark"])
    _write_col_header(ws, row, 4, "Description", bg=_C["hdr_light"], fg=_C["hdr_dark"], colspan=3)
    row += 1

    _OUTCOME_DESC = {
        "PASS":         "All test cases passed — full marks",
        "FAIL":         "One or more TCs wrong answer — partial/no marks",
        "TLE":          "Time limit exceeded — per-TC or global",
        "ERROR":        "Runtime error in student code",
        "BLOCKED":      "Rejected by security checker / AST pre-check (syntax error)",
        "RATE_LIMITED": "HTTP 429 — queue full, submission rejected by admission control",
        "SYSTEM_ERROR": "Infrastructure failure — not the student's fault; retried up to 3×",
        "JUDGE0_ERROR": "Judge0 / network error during submission or polling",
        "UNKNOWN":      "Status could not be determined",
    }
    STATUS_ORDER = ["PASS","FAIL","TLE","ERROR","BLOCKED","RATE_LIMITED","SYSTEM_ERROR","JUDGE0_ERROR","UNKNOWN"]
    all_statuses = sorted(status_counts.items(), key=lambda x: -x[1])
    ordered = [(s, status_counts[s]) for s in STATUS_ORDER if s in status_counts]
    ordered += [(s, c) for s, c in all_statuses if s not in STATUS_ORDER]

    for i, (status, count) in enumerate(ordered):
        bg, fg = _STATUS_STYLE.get(status, (_C["unknown"], "000000"))
        for ci, val in [
            (1, status),
            (2, count),
            (3, _pct(count, completed)),
            (4, _OUTCOME_DESC.get(status, "")),
        ]:
            c = ws.cell(row=row, column=ci, value=val)
            if ci <= 3:
                _style(c, bg=bg, fg=fg, bold=(ci == 1))
            else:
                _style(c, bg=_C["alt_row"] if i % 2 else _C["white"])
        if 4 < NCOLS:
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=NCOLS)
        row += 1
    row += 1

    # ── Section: Per-Problem Analysis ─────────────────────────────────────────
    _write_section(ws, row, "PER-PROBLEM ANALYSIS", NCOLS)
    row += 1

    for ci, h in enumerate(["Problem ID", "Submissions", "% of Total",
                              "Passed", "Pass Rate", "Avg Latency (ms)"], 1):
        _write_col_header(ws, row, ci, h, bg=_C["hdr_light"], fg=_C["hdr_dark"])
    row += 1

    # Build per-problem stats from per_problem_report
    prob_stats: dict[str, dict] = {}
    for prob_rpt in per_problem:
        pid  = prob_rpt["problem_id"]
        subs = prob_rpt.get("submissions", [])
        passed = sum(1 for s in subs if s.get("harness_status") == "PASS")
        lats_p = [s.get("latency_ms", 0) for s in subs if "latency_ms" in s]
        prob_stats[pid] = {
            "count":      len(subs),
            "passed":     passed,
            "avg_lat":    round(sum(lats_p) / len(lats_p), 0) if lats_p else 0,
        }

    total_sub_count = sum(problem_counts.values()) or 1
    for i, (pid, count) in enumerate(sorted(problem_counts.items(), key=lambda x: -x[1])):
        bg = _C["alt_row"] if i % 2 else _C["white"]
        stats = prob_stats.get(pid, {})
        passed   = stats.get("passed", 0)
        pass_pct = _pct(passed, count)
        avg_l    = stats.get("avg_lat", "–")
        for ci, val in [
            (1, pid), (2, count), (3, _pct(count, total_sub_count)),
            (4, passed), (5, pass_pct), (6, avg_l),
        ]:
            c = ws.cell(row=row, column=ci, value=val)
            _style(c, bg=bg, bold=(ci == 1))
        row += 1
    row += 1

    # ── Section: Solution Type Analysis ───────────────────────────────────────
    _write_section(ws, row, "SOLUTION TYPE ANALYSIS", NCOLS)
    row += 1

    for ci, h in enumerate(["Solution Type", "Count", "% of Total",
                              "Expected Behaviour", "", ""], 1):
        bg = _C["hdr_light"] if h else _C["white"]
        _write_col_header(ws, row, ci, h, bg=bg, fg=_C["hdr_dark"])
    row += 1

    _TYPE_DESC = {
        "accepted":            "Should PASS all test cases",
        "wrong_answer":        "Should FAIL some or all test cases",
        "time_limit_exceeded": "Should TLE on all test cases",
        "runtime_error":       "Should ERROR on all test cases",
        "syntax_error":        "Should be BLOCKED by security / AST checker",
        "partial":             "Passes a subset of test cases",
    }
    for i, (stype, count) in enumerate(sorted(type_counts.items(), key=lambda x: -x[1])):
        bg = _C["alt_row"] if i % 2 else _C["white"]
        desc = _TYPE_DESC.get(stype, "")
        for ci, val in [(1, stype), (2, count), (3, _pct(count, total_sub_count)),
                        (4, desc), (5, ""), (6, "")]:
            c = ws.cell(row=row, column=ci, value=val)
            _style(c, bg=bg, bold=(ci == 1))
        if desc:
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=NCOLS)
        row += 1
    row += 1

    # ── Section: Batch Analysis (if present) ──────────────────────────────────
    if batches:
        _write_section(ws, row, "BATCH ANALYSIS", NCOLS)
        row += 1

        for ci, h in enumerate(["Batch ID", "Users", "Passed",
                                  "Accept %", "Duration (s)", "Avg Lat (ms)"], 1):
            _write_col_header(ws, row, ci, h, bg=_C["hdr_light"], fg=_C["hdr_dark"])
        row += 1

        for i, b in enumerate(batches):
            bg = _C["alt_row"] if i % 2 else _C["white"]
            for ci, val in [
                (1, b.get("batch_id", "")),
                (2, b.get("total_users", "")),
                (3, b.get("passed", "")),
                (4, f"{b.get('acceptance_pct', 0):.1f}%"),
                (5, f"{b.get('duration_s', 0):.2f}"),
                (6, round(b.get("avg_latency_ms", 0), 0)),
            ]:
                c = ws.cell(row=row, column=ci, value=val)
                _style(c, bg=bg)
            row += 1
        row += 1

    # ── Section: Concurrency Mode Analysis ───────────────────────────────────
    _write_section(ws, row, "CONCURRENCY MODE ANALYSIS", NCOLS)
    row += 1

    _CONC_MODE_DETAIL = {
        "flat_concurrent_noramp": (
            "Flat Concurrent (no ramp)",
            "All virtual users launch simultaneously. Maximum burst — worst-case for the queue, "
            "Redis, and Judge0 admission control. Use to find saturation point.",
        ),
        "flat_concurrent_ramp": (
            "Flat Concurrent + Ramp-up",
            "Virtual users introduced evenly over the ramp window. Reduces initial spike; "
            "simulates users logging in progressively. Good for steady-state throughput testing.",
        ),
        "pipelined_batch": (
            "Pipelined Batch",
            "Users split into N-sized batches; all batches are fired simultaneously as goroutines. "
            "Within each batch users run concurrently. Useful to test Judge0 at specific concurrency levels.",
        ),
    }

    batch_size = cfg.get("batch_size", 0) or 0
    ramp       = cfg.get("ramp_up_sec", 0) or 0
    users      = cfg.get("users", 0) or 0

    if batch_size > 0:
        mode_key = "pipelined_batch"
    elif ramp > 0:
        mode_key = "flat_concurrent_ramp"
    else:
        mode_key = "flat_concurrent_noramp"

    mode_name, mode_desc = _CONC_MODE_DETAIL[mode_key]

    conc_rows = [
        ("Concurrency Mode",          mode_name),
        ("Description",               mode_desc),
        ("Virtual Users",             users),
        ("Ramp-up Duration",          f"{ramp} s" if ramp else "None (all users start instantly)"),
        ("Ramp-up User Cadence",      f"1 user every ~{_ramp_cadence_ms(cfg):.0f} ms"
                                      if ramp > 0 else "N/A"),
        ("Batch Size",                batch_size if batch_size else "N/A (flat mode)"),
        ("Number of Batches",         math.ceil(users / batch_size) if batch_size else "N/A"),
        ("Result Delivery",           _result_delivery_label(cfg)),
        ("Flask Stack Mode",          "Yes — Flask API → Redis → worker → Judge0 → SSE"
                                      if cfg.get("mode") == "flask_stack"
                                      else "No — Direct Judge0 submissions (no Flask/Redis)"),
        ("Solution Mix",              "Random (accepted + wrong_answer + TLE + error + syntax)"
                                      if not cfg.get("accept_only") else "Accepted only (100% pass target)"),
        ("Dry Run",                   cfg.get("dry_run", False)),
    ]

    for i, (k, v) in enumerate(conc_rows):
        bg = _C["alt_row"] if i % 2 else _C["white"]
        c1 = ws.cell(row=row, column=1, value=k)
        _style(c1, bg=bg, bold=True)
        c2 = ws.cell(row=row, column=2, value=v)
        _style(c2, bg=bg)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NCOLS)
        row += 1
    row += 1

    # ── Section: Score Distribution ───────────────────────────────────────────
    full_marks, partial_marks, zero_marks = _score_distribution(all_results)

    _write_section(ws, row, "SCORE DISTRIBUTION", NCOLS)
    row += 1

    for ci, h in enumerate(["Category", "Count", "% of Completed",
                              "Description", "", ""], 1):
        bg = _C["hdr_light"] if h else _C["white"]
        _write_col_header(ws, row, ci, h, bg=bg, fg=_C["hdr_dark"])
    row += 1

    dist_rows = [
        ("Full Marks  (100%)",   full_marks,    _C["pass"],    _C["pass_fg"],
         "Score == Total TCs — all test cases passed"),
        ("Partial Marks  (1–99%)", partial_marks, _C["partial"], _C["partial_fg"],
         "1 ≤ Score < Total TCs — some test cases passed"),
        ("Zero Marks  (0%)",     zero_marks,    _C["fail"],    _C["fail_fg"],
         "Score == 0 — no test cases passed (or blocked/error before any TC ran)"),
    ]

    for label, count, bg, fg, desc in dist_rows:
        c1 = ws.cell(row=row, column=1, value=label)
        _style(c1, bg=bg, fg=fg, bold=True)
        c2 = ws.cell(row=row, column=2, value=count)
        _style(c2, bg=bg, fg=fg, halign="center")
        c3 = ws.cell(row=row, column=3, value=_pct(count, completed))
        _style(c3, bg=bg, fg=fg, halign="center")
        c4 = ws.cell(row=row, column=4, value=desc)
        _style(c4, bg=_C["alt_row"] if dist_rows.index((label,count,bg,fg,desc)) % 2 else _C["white"])
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=NCOLS)
        row += 1

    # Accepted-solutions-only analysis
    if cfg.get("accept_only") or (type_counts and all(t == "accepted" for t in type_counts)):
        actual_pass  = status_counts.get("PASS", 0)
        infra_errors = (status_counts.get("JUDGE0_ERROR", 0)
                        + status_counts.get("SYSTEM_ERROR", 0))
        tle_count    = status_counts.get("TLE", 0)
        pass_pct     = _pct(actual_pass, completed) if completed else "–"
        note = (
            f"NOTE: All {completed} submissions used 'accepted' solutions "
            f"(correct code — expected 100% PASS).  "
            f"Actual PASS: {actual_pass} ({pass_pct}).  "
            f"Infrastructure errors (JUDGE0_ERROR + SYSTEM_ERROR): {infra_errors}.  "
            f"Queue TLEs: {tle_count}."
        )
        c = ws.cell(row=row, column=1, value=note)
        bg = _C["fail"] if (completed and actual_pass / completed < 0.9) else _C["tle"]
        _style(c, bg=bg, bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        row += 1

    row += 1

    # ── Section: Idempotency / Cache Analysis (only if duplicates exist) ──────
    dup_count = summary.get("duplicate_idempotent", 0)
    if dup_count > 0:
        _write_section(ws, row, "IDEMPOTENCY / WARM-CACHE ANALYSIS", NCOLS)
        row += 1

        dup_note = ws.cell(row=row, column=1, value=(
            f"This run had {dup_count} duplicate (idempotent) submissions "
            f"({_pct(dup_count, total_subs)} of total). "
            "A duplicate occurs when the same student_id + assessment_id + SHA-256(code) "
            "is re-submitted within the TTL window (2h). Flask returns HTTP 200 immediately "
            "from Redis — no re-grading occurs. Latency for duplicates reflects only "
            "Redis lookup + SSE stream time, not Judge0 execution time."
        ))
        _style(dup_note, bg=_C["alt_row"], size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        ws.row_dimensions[row].height = 30
        row += 1

        idem_rows = [
            ("Total Submissions",          total_subs),
            ("Duplicate (idempotent) hits",dup_count),
            ("Duplicate %",               _pct(dup_count, total_subs)),
            ("Fresh submissions",          total_subs - dup_count),
            ("Redis key TTL",              "2 h  (SHA-256 of student_id:assessment_id:code_hash)"),
            ("Cache scenario",             "Warm-cache: same code re-submitted within TTL window"),
        ]
        for i, (k, v) in enumerate(idem_rows):
            bg = _C["alt_row"] if i % 2 else _C["white"]
            c1 = ws.cell(row=row, column=1, value=k); _style(c1, bg=bg, bold=True)
            c2 = ws.cell(row=row, column=2, value=v); _style(c2, bg=bg)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NCOLS)
            row += 1
        row += 1

    # ── Section: Live System Metrics ──────────────────────────────────────────
    if metrics_data:
        _METRICS_COLS = [
            ("timestamp",          "Timestamp"),
            ("cpu_pct",            "CPU %"),
            ("mem_pct",            "RAM %"),
            ("mem_used_gb",        "RAM Used (GB)"),
            ("mem_avail_gb",       "RAM Avail (GB)"),
            ("disk_read_mb_s",     "Disk Read (MB/s)"),
            ("disk_write_mb_s",    "Disk Write (MB/s)"),
            ("net_sent_mb_s",      "Net TX (MB/s)"),
            ("net_recv_mb_s",      "Net RX (MB/s)"),
            ("tcp_established",    "TCP Connections"),
            ("redis_clients",      "Redis Clients"),
            ("redis_mem_used_mb",  "Redis Mem (MB)"),
            ("redis_queue_depth",  "Redis Queue Depth"),
        ]
        present = [
            (k, h) for k, h in _METRICS_COLS
            if any(k in m for m in metrics_data)
        ]
        ncols_m = len(present)

        _write_section(ws, row, "LIVE SYSTEM METRICS (captured during test)", ncols_m)
        row += 1

        # Column headers
        for ci, (k, h) in enumerate(present, 1):
            _write_col_header(ws, row, ci, h, bg=_C["hdr_light"], fg=_C["hdr_dark"])
        row += 1

        for i, m in enumerate(metrics_data):
            bg = _C["alt_row"] if i % 2 else _C["white"]
            for ci, (k, _h) in enumerate(present, 1):
                val = m.get(k, "")
                c = ws.cell(row=row, column=ci, value=val)
                # Severity colouring for CPU and RAM
                cell_bg = bg
                if k == "cpu_pct" and isinstance(val, (int, float)):
                    cell_bg = _C["cpu_crit"] if val >= 90 else (_C["cpu_warn"] if val >= 70 else _C["cpu_ok"])
                elif k == "mem_pct" and isinstance(val, (int, float)):
                    cell_bg = _C["cpu_crit"] if val >= 90 else (_C["cpu_warn"] if val >= 75 else _C["cpu_ok"])
                elif k == "redis_queue_depth" and isinstance(val, int) and val > 150:
                    cell_bg = _C["tle"]
                _style(c, bg=cell_bg)
            row += 1

        # Summary stats row at the bottom
        row += 1
        _write_section(ws, row, "Live Metrics Summary", ncols_m)
        row += 1
        for ci, (k, h) in enumerate(present, 1):
            vals = [m[k] for m in metrics_data if k in m and isinstance(m[k], (int, float))]
            if vals:
                summary_val = (
                    f"avg={sum(vals)/len(vals):.1f}  max={max(vals):.1f}  min={min(vals):.1f}"
                )
            else:
                summary_val = "–"
            c = ws.cell(row=row, column=ci, value=summary_val if ci > 1 else "Summary (avg/max/min)")
            _style(c, bg=_C["light_grey"], bold=(ci == 1), size=9)
        row += 1

    # ── Section: Container CPU & Memory ───────────────────────────────────────
    # Extract per-container series from snapshots that have docker stats
    container_series: dict = {}   # name → {cg_cpu, mem_mbs, mem_pcts, real_cpu, real_rss}
    container_order: list = []    # first-seen insertion order

    for m in metrics_data:
        for c in m.get("containers", []):
            name = c.get("name", "?")
            if name not in container_series:
                container_series[name] = {
                    "cg_cpu_pcts": [], "mem_mbs": [], "mem_pcts": [],
                    "real_cpu_pcts": [], "real_rss_mbs": [],
                }
                container_order.append(name)
            container_series[name]["cg_cpu_pcts"].append(c.get("cg_cpu_pct", c.get("cpu_pct", 0)))
            container_series[name]["mem_mbs"].append(c.get("mem_mb", 0))
            container_series[name]["mem_pcts"].append(c.get("mem_pct", 0))
            if c.get("real_cpu_pct_norm") is not None:
                container_series[name]["real_cpu_pcts"].append(c["real_cpu_pct_norm"])
            if c.get("real_rss_mb") is not None:
                container_series[name]["real_rss_mbs"].append(c["real_rss_mb"])

    if container_series:
        sorted_names = sorted(container_order)

        # ── Container peak summary ─────────────────────────────────────────
        N_CSUMM = 9
        _write_section(ws, row, "CONTAINER PERFORMANCE SUMMARY  (peaks & averages during test)", N_CSUMM)
        row += 1

        for ci, h in enumerate([
            "Container",
            "Avg cg CPU%", "Max cg CPU%",      # docker/cgroup relative
            "Avg Real CPU%", "Max Real CPU%",   # actual host normalised 0-100%
            "Avg Mem MB (cg)", "Max Mem MB (cg)",
            "Avg RSS MB (real)", "Max RSS MB (real)",
        ], 1):
            _write_col_header(ws, row, ci, h, bg=_C["hdr_light"], fg=_C["hdr_dark"])
        row += 1

        for i, name in enumerate(sorted_names):
            series  = container_series[name]
            alt_bg  = _C["alt_row"] if i % 2 else _C["white"]

            def _avg(lst): return sum(lst) / max(len(lst), 1) if lst else None
            def _max(lst): return max(lst) if lst else None

            avg_cg   = _avg(series["cg_cpu_pcts"]);  max_cg   = _max(series["cg_cpu_pcts"])
            avg_rc   = _avg(series["real_cpu_pcts"]); max_rc   = _max(series["real_cpu_pcts"])
            avg_mem  = _avg(series["mem_mbs"]);       max_mem  = _max(series["mem_mbs"])
            avg_rss  = _avg(series["real_rss_mbs"]);  max_rss  = _max(series["real_rss_mbs"])
            max_memp = _max(series["mem_pcts"]) or 0

            def _cpu_bg(v):
                if v is None: return alt_bg
                return _C["cpu_crit"] if v >= 90 else (_C["cpu_warn"] if v >= 70 else _C["cpu_ok"])
            def _mem_bg(v):
                if v is None: return alt_bg
                return _C["cpu_crit"] if max_memp >= 90 else (_C["cpu_warn"] if max_memp >= 75 else _C["cpu_ok"])

            vals = [
                (name,                                              alt_bg),
                (f"{avg_cg:.1f}%" if avg_cg is not None else "–",  alt_bg),
                (f"{max_cg:.1f}%" if max_cg is not None else "–",  _cpu_bg(max_cg)),
                (f"{avg_rc:.1f}%" if avg_rc is not None else "–",  alt_bg),
                (f"{max_rc:.1f}%" if max_rc is not None else "–",  _cpu_bg(max_rc)),
                (f"{avg_mem:.0f}" if avg_mem is not None else "–", alt_bg),
                (f"{max_mem:.0f}" if max_mem is not None else "–", _mem_bg(max_mem)),
                (f"{avg_rss:.0f}" if avg_rss is not None else "–", alt_bg),
                (f"{max_rss:.0f}" if max_rss is not None else "–", _mem_bg(max_rss)),
            ]
            for ci, (val, bg) in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=val)
                _style(c, bg=bg, bold=(ci == 1))
        row += 1
        row += 1

        # ── Container cgroup CPU % timeline ───────────────────────────────
        N_CTIMELINE = 1 + len(sorted_names)
        _write_section(ws, row, "CONTAINER CPU %  —  cgroup-relative  (docker stats, can exceed 100% per core)", N_CTIMELINE)
        row += 1

        ws.cell(row=row, column=1, value="Timestamp")
        _style(ws.cell(row=row, column=1), bg=_C["hdr_light"], fg=_C["hdr_dark"], bold=True, halign="center")
        for ci, name in enumerate(sorted_names, 2):
            c = ws.cell(row=row, column=ci, value=f"{name}\ncg CPU %")
            _style(c, bg=_C["hdr_light"], fg=_C["hdr_dark"], bold=True, halign="center")
        ws.row_dimensions[row].height = 28
        row += 1

        for i, m in enumerate(metrics_data):
            bg = _C["alt_row"] if i % 2 else _C["white"]
            ws.cell(row=row, column=1, value=m.get("timestamp", ""))
            _style(ws.cell(row=row, column=1), bg=bg)
            by_name = {cdata["name"]: cdata for cdata in m.get("containers", [])}
            for ci, name in enumerate(sorted_names, 2):
                cdata = by_name.get(name)
                if cdata:
                    cpu = cdata.get("cg_cpu_pct", cdata.get("cpu_pct", 0))
                    cpu_bg = (_C["cpu_crit"] if cpu >= 90 else _C["cpu_warn"] if cpu >= 70 else _C["cpu_ok"])
                    c = ws.cell(row=row, column=ci, value=f"{cpu:.1f}%")
                    _style(c, bg=cpu_bg, halign="center")
                else:
                    c = ws.cell(row=row, column=ci, value="–")
                    _style(c, bg=bg, halign="center")
            row += 1
        row += 1

        # ── Container real host CPU % timeline (legacy psutil field, shown when present) ────────
        _write_section(ws, row, "CONTAINER CPU %  —  actual host real_cpu_pct_norm  (legacy; populated by older collector)", N_CTIMELINE)
        row += 1

        ws.cell(row=row, column=1, value="Timestamp")
        _style(ws.cell(row=row, column=1), bg=_C["hdr_light"], fg=_C["hdr_dark"], bold=True, halign="center")
        for ci, name in enumerate(sorted_names, 2):
            c = ws.cell(row=row, column=ci, value=f"{name}\nReal CPU %")
            _style(c, bg=_C["hdr_light"], fg=_C["hdr_dark"], bold=True, halign="center")
        ws.row_dimensions[row].height = 28
        row += 1

        for i, m in enumerate(metrics_data):
            bg = _C["alt_row"] if i % 2 else _C["white"]
            ws.cell(row=row, column=1, value=m.get("timestamp", ""))
            _style(ws.cell(row=row, column=1), bg=bg)
            by_name = {cdata["name"]: cdata for cdata in m.get("containers", [])}
            for ci, name in enumerate(sorted_names, 2):
                cdata = by_name.get(name)
                val = cdata.get("real_cpu_pct_norm") if cdata else None
                if val is not None:
                    cpu_bg = (_C["cpu_crit"] if val >= 90 else _C["cpu_warn"] if val >= 70 else _C["cpu_ok"])
                    c = ws.cell(row=row, column=ci, value=f"{val:.1f}%")
                    _style(c, bg=cpu_bg, halign="center")
                else:
                    c = ws.cell(row=row, column=ci, value="–")
                    _style(c, bg=bg, halign="center")
            row += 1
        row += 1

        # ── Container cgroup memory (MB) timeline ─────────────────────────
        _write_section(ws, row, "CONTAINER MEMORY MB  —  cgroup  (docker stats)", N_CTIMELINE)
        row += 1

        ws.cell(row=row, column=1, value="Timestamp")
        _style(ws.cell(row=row, column=1), bg=_C["hdr_light"], fg=_C["hdr_dark"], bold=True, halign="center")
        for ci, name in enumerate(sorted_names, 2):
            c = ws.cell(row=row, column=ci, value=f"{name}\ncg Mem MB")
            _style(c, bg=_C["hdr_light"], fg=_C["hdr_dark"], bold=True, halign="center")
        ws.row_dimensions[row].height = 28
        row += 1

        for i, m in enumerate(metrics_data):
            bg = _C["alt_row"] if i % 2 else _C["white"]
            ws.cell(row=row, column=1, value=m.get("timestamp", ""))
            _style(ws.cell(row=row, column=1), bg=bg)
            by_name = {cdata["name"]: cdata for cdata in m.get("containers", [])}
            for ci, name in enumerate(sorted_names, 2):
                cdata = by_name.get(name)
                if cdata:
                    mem     = cdata.get("mem_mb", 0)
                    mem_pct = cdata.get("mem_pct", 0)
                    mem_bg  = (_C["cpu_crit"] if mem_pct >= 90 else _C["cpu_warn"] if mem_pct >= 75 else _C["cpu_ok"])
                    c = ws.cell(row=row, column=ci, value=int(mem))
                    _style(c, bg=mem_bg, halign="center")
                else:
                    c = ws.cell(row=row, column=ci, value="–")
                    _style(c, bg=bg, halign="center")
            row += 1
        row += 1

        # ── Container real RSS memory (MB) timeline (legacy psutil field) ────────────────────────
        _write_section(ws, row, "CONTAINER MEMORY MB  —  actual RSS  (legacy; populated by older collector)", N_CTIMELINE)
        row += 1

        ws.cell(row=row, column=1, value="Timestamp")
        _style(ws.cell(row=row, column=1), bg=_C["hdr_light"], fg=_C["hdr_dark"], bold=True, halign="center")
        for ci, name in enumerate(sorted_names, 2):
            c = ws.cell(row=row, column=ci, value=f"{name}\nReal RSS MB")
            _style(c, bg=_C["hdr_light"], fg=_C["hdr_dark"], bold=True, halign="center")
        ws.row_dimensions[row].height = 28
        row += 1

        for i, m in enumerate(metrics_data):
            bg = _C["alt_row"] if i % 2 else _C["white"]
            ws.cell(row=row, column=1, value=m.get("timestamp", ""))
            _style(ws.cell(row=row, column=1), bg=bg)
            by_name = {cdata["name"]: cdata for cdata in m.get("containers", [])}
            for ci, name in enumerate(sorted_names, 2):
                cdata = by_name.get(name)
                val = cdata.get("real_rss_mb") if cdata else None
                if val is not None:
                    rss_bg = (_C["cpu_crit"] if val >= 1500 else _C["cpu_warn"] if val >= 900 else _C["cpu_ok"])
                    c = ws.cell(row=row, column=ci, value=int(val))
                    _style(c, bg=rss_bg, halign="center")
                else:
                    c = ws.cell(row=row, column=ci, value="–")
                    _style(c, bg=bg, halign="center")
            row += 1
        row += 1

        # Set widths for container columns
        ws.column_dimensions["A"].width = 22
        for ci in range(2, N_CTIMELINE + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 18

    # ── Column widths ─────────────────────────────────────────────────────────
    _set_col_widths(ws, [30, 18, 14, 40, 18, 18])
    ws.freeze_panes = ws.cell(row=3, column=1)


# ─────────────────────────────────────────────────────────────────────────────
#  Sheet 3 — User Test Case Results
# ─────────────────────────────────────────────────────────────────────────────

def build_sheet3(wb: openpyxl.Workbook, report: dict):
    ws = wb.create_sheet("User Test Case Results")
    ws.sheet_view.showGridLines = False

    per_problem = report.get("per_problem_report", [])

    # Flatten all submissions keeping the problem_id context
    all_subs: list[dict] = []
    max_tcs = 0
    for prob_rpt in per_problem:
        pid  = prob_rpt["problem_id"]
        for sub in prob_rpt.get("submissions", []):
            entry = dict(sub)
            entry["_problem_id"] = pid
            all_subs.append(entry)
            max_tcs = max(max_tcs, len(sub.get("test_cases") or []))

    # ── Column layout ─────────────────────────────────────────────────────────
    STATIC = [
        "User ID",
        "Problem ID",
        "Solution ID",
        "Solution Type",
        "Harness Status",
        "Score",
        "Total TCs",
        "Marks %",
        "Partial?",
        "Latency (ms)",
        "Error / Notes",
        "Submitted Code",   # source_code from question bank / JSON report
    ]
    N_STATIC = len(STATIC)       # 12
    TC_COLS  = 3                 # Status | Got | Expected per TC
    TOTAL_COLS = N_STATIC + max_tcs * TC_COLS

    # ── Title ─────────────────────────────────────────────────────────────────
    _write_title(ws, 1, "USER TEST CASE RESULTS & MARKS", TOTAL_COLS, fontsize=14)
    ws.row_dimensions[1].height = 26

    # ── Sub-header: marks legend ──────────────────────────────────────────────
    legend_vals = [
        "Marks Legend:  ",
        "Full marks (100%)",   "",
        "Partial marks (1–99%)", "",
        "No marks (0%)",       "",
        "Score = TCs passed / Total TCs",
    ]
    for ci, val in enumerate(legend_vals, 1):
        c = ws.cell(row=2, column=ci, value=val)
        bg = (
            _C["pass"]    if ci == 2 else
            _C["partial"] if ci == 4 else
            _C["fail"]    if ci == 6 else
            _C["light_grey"]
        )
        _style(c, bg=bg, size=9, bold=(ci in (2, 4, 6)))
    # fill remaining legend cells
    for ci in range(len(legend_vals) + 1, TOTAL_COLS + 1):
        ws.cell(row=2, column=ci).fill = _fill(_C["light_grey"])
    ws.row_dimensions[2].height = 16

    # ── Per-user aggregated marks summary ─────────────────────────────────────
    # Aggregate: for each unique user_id, sum score and total_tcs across all submissions
    user_agg: dict = {}
    for s in all_subs:
        uid = s.get("user_id", "?")
        if uid not in user_agg:
            user_agg[uid] = {
                "score": 0, "total_tcs": 0, "submissions": 0,
                "problems": set(), "statuses": [],
            }
        user_agg[uid]["score"]       += s.get("score", 0)
        user_agg[uid]["total_tcs"]   += s.get("total_tcs", 0)
        user_agg[uid]["submissions"] += 1
        user_agg[uid]["problems"].add(s.get("_problem_id", ""))
        user_agg[uid]["statuses"].append(s.get("harness_status", ""))

    # Section header
    _write_title(ws, 3, "PER-USER MARKS SUMMARY", min(8, TOTAL_COLS), fontsize=12)

    # Column headers for summary
    SUMM_COLS = ["User ID", "Submissions", "Problems Attempted",
                 "Total Score", "Total TCs", "Marks %", "Partial?", "Dominant Status"]
    for ci, h in enumerate(SUMM_COLS, 1):
        _write_col_header(ws, 4, ci, h, bg=_C["hdr_light"], fg=_C["hdr_dark"])
    # Fill remaining columns empty
    for ci in range(len(SUMM_COLS) + 1, TOTAL_COLS + 1):
        c = ws.cell(row=4, column=ci, value="")
        c.fill = _fill(_C["hdr_light"])

    summary_start_data = 5
    for i, (uid, agg) in enumerate(sorted(user_agg.items())):
        data_row = summary_start_data + i
        score    = agg["score"]
        total    = agg["total_tcs"] or 1
        pct      = score / total * 100
        partial  = (0 < score < total)

        if pct >= 100:
            mb, mf = _C["pass"],    _C["pass_fg"]
        elif pct > 0:
            mb, mf = _C["partial"], _C["partial_fg"]
        else:
            mb, mf = _C["fail"],    _C["fail_fg"]

        # Dominant status
        from collections import Counter
        status_ctr = Counter(agg["statuses"])
        dominant   = status_ctr.most_common(1)[0][0] if status_ctr else "–"

        alt_bg = _C["alt_row"] if i % 2 else _C["white"]
        row_vals = [
            uid,
            agg["submissions"],
            len(agg["problems"]),
            score,
            agg["total_tcs"],
            f"{pct:.1f}%",
            "Yes" if partial else "No",
            dominant,
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=data_row, column=ci, value=val)
            if ci == 6:   # Marks %
                _style(c, bg=mb, fg=mf, bold=True, halign="center")
            elif ci == 7:  # Partial?
                _style(c, bg=mb if partial else alt_bg,
                       fg=mf if partial else "000000", halign="center")
            elif ci == 8:  # Status
                sbg, sfg = _STATUS_STYLE.get(dominant, (_C["unknown"], "000000"))
                _style(c, bg=sbg, fg=sfg, bold=True, halign="center")
            else:
                _style(c, bg=alt_bg, halign="center" if ci == 1 else "left")
        # Fill remaining cols with empty styled cells
        for ci in range(len(row_vals) + 1, TOTAL_COLS + 1):
            c = ws.cell(row=data_row, column=ci, value="")
            _style(c, bg=alt_bg)

    user_summary_rows = len(user_agg)
    # Totals row for the summary
    tot_row = summary_start_data + user_summary_rows
    total_score_all = sum(a["score"]     for a in user_agg.values())
    total_tc_all    = sum(a["total_tcs"] for a in user_agg.values())
    full_u   = sum(1 for a in user_agg.values() if a["score"] >= (a["total_tcs"] or 1))
    part_u   = sum(1 for a in user_agg.values()
                   if 0 < a["score"] < (a["total_tcs"] or 1))
    zero_u   = sum(1 for a in user_agg.values() if a["score"] == 0)
    for ci, val in enumerate([
        "TOTALS", len(user_agg), "", total_score_all, total_tc_all,
        _pct(total_score_all, total_tc_all or 1),
        f"{part_u} partial",
        f"Full={full_u}  Partial={part_u}  Zero={zero_u}",
    ], 1):
        c = ws.cell(row=tot_row, column=ci, value=val)
        _style(c, bg=_C["hdr_light"], bold=True, fg=_C["hdr_dark"])
    for ci in range(len(SUMM_COLS) + 1, TOTAL_COLS + 1):
        ws.cell(row=tot_row, column=ci).fill = _fill(_C["hdr_light"])

    # Blank separator row then the per-submission detail title
    sep_row   = tot_row + 1
    ws.row_dimensions[sep_row].height = 8

    detail_title_row = sep_row + 1
    _write_title(ws, detail_title_row, "PER-SUBMISSION DETAIL", TOTAL_COLS, fontsize=12)

    # ── Column headers row (for per-submission table) ─────────────────────────
    hdr_row = detail_title_row + 1

    for ci, h in enumerate(STATIC, 1):
        _write_col_header(ws, hdr_row, ci, h, bg=_C["hdr_mid"], fg="FFFFFF")

    for tc_i in range(1, max_tcs + 1):
        col = N_STATIC + (tc_i - 1) * TC_COLS + 1
        _write_col_header(ws, hdr_row, col,     f"TC{tc_i}  Status",   bg=_C["hdr_light"], fg=_C["hdr_dark"])
        _write_col_header(ws, hdr_row, col + 1, f"TC{tc_i}  Got",      bg=_C["hdr_light"], fg=_C["hdr_dark"])
        _write_col_header(ws, hdr_row, col + 2, f"TC{tc_i}  Expected", bg=_C["hdr_light"], fg=_C["hdr_dark"])

    ws.row_dimensions[hdr_row].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────────
    detail_data_start = hdr_row + 1
    for i, sub in enumerate(all_subs):
        data_row = detail_data_start + i
        alt_bg   = _C["alt_row"] if i % 2 else _C["white"]

        harness_status = sub.get("harness_status", "UNKNOWN")
        score          = sub.get("score", 0)
        total_tcs      = sub.get("total_tcs", 0)
        marks_pct      = (score / total_tcs * 100) if total_tcs > 0 else 0.0
        partial        = (0 < score < total_tcs)
        latency        = sub.get("latency_ms", "")
        error_note     = sub.get("error", "") or sub.get("judge0_status", "")
        source_code    = sub.get("source_code", "")

        static_vals = [
            sub.get("user_id", ""),
            sub.get("_problem_id", ""),
            sub.get("solution_id", ""),
            sub.get("solution_type", ""),
            harness_status,
            score,
            total_tcs,
            f"{marks_pct:.1f}%",
            "Yes" if partial else "No",
            latency,
            error_note,
            source_code,         # col 12 — Submitted Code
        ]

        for ci, val in enumerate(static_vals, 1):
            c = ws.cell(row=data_row, column=ci, value=val)
            # Default styling
            _style(c, bg=alt_bg)

            # Harness status cell (col 5)
            if ci == 5:
                bg_s, fg_s = _STATUS_STYLE.get(harness_status, (_C["unknown"], "000000"))
                _style(c, bg=bg_s, fg=fg_s, bold=True, halign="center")

            # Marks % cell (col 8)
            elif ci == 8:
                if marks_pct >= 100:
                    mb, mf = _C["pass"], _C["pass_fg"]
                elif marks_pct == 0:
                    mb, mf = _C["fail"], _C["fail_fg"]
                else:
                    mb, mf = _C["partial"], _C["partial_fg"]
                _style(c, bg=mb, fg=mf, bold=True, halign="center")

            # Partial? (col 9)
            elif ci == 9:
                if partial:
                    _style(c, bg=_C["partial"], fg=_C["partial_fg"], bold=True, halign="center")
                else:
                    _style(c, bg=alt_bg, halign="center")

            # Submitted Code (col 12) — fixed-height, monospaced font, top-aligned
            elif ci == 12:
                c.font      = Font(name="Courier New", size=8, color="1F1F1F")
                c.fill      = _fill(_C["light_grey"])
                c.border    = _thin_border()
                c.alignment = Alignment(horizontal="left", vertical="top",
                                        wrap_text=True)
                ws.row_dimensions[data_row].height = max(
                    ws.row_dimensions[data_row].height or 15,
                    min(120, 13 + source_code.count("\n") * 10)
                )

            # User ID / latency — centre
            elif ci in (1, 6, 7, 10):
                c.alignment = _center_align(wrap=False)

        # ── TC columns ────────────────────────────────────────────────────────
        test_cases = sub.get("test_cases") or []
        for tc_idx, tc in enumerate(test_cases):
            col    = N_STATIC + tc_idx * TC_COLS + 1
            status = tc.get("status", "")
            got    = tc.get("got", "") or tc.get("detail", "")
            exp    = tc.get("expected", "")

            tc_bg, tc_fg = _STATUS_STYLE.get(status, (_C["unknown"], "000000"))

            # Status cell — coloured
            cs = ws.cell(row=data_row, column=col, value=status)
            _style(cs, bg=tc_bg, fg=tc_fg, bold=True, halign="center")

            # Got cell
            cg = ws.cell(row=data_row, column=col + 1, value=got)
            _style(cg, bg=alt_bg, size=9)

            # Expected cell
            ce = ws.cell(row=data_row, column=col + 2, value=exp)
            _style(ce, bg=alt_bg, size=9)

        # Fill remaining TC columns with empty styled cells
        filled_tcs = len(test_cases)
        for tc_idx in range(filled_tcs, max_tcs):
            col = N_STATIC + tc_idx * TC_COLS + 1
            for offset in range(TC_COLS):
                c = ws.cell(row=data_row, column=col + offset, value="")
                _style(c, bg=alt_bg)

    # ── Summary row at the bottom ─────────────────────────────────────────────
    if all_subs:
        summary_row = detail_data_start + len(all_subs)
        ws.row_dimensions[summary_row].height = 16
        totals = [
            "TOTALS",
            f"{len(set(s['_problem_id'] for s in all_subs))} problems",
            "",
            "",
            "",
            sum(s.get("score", 0) for s in all_subs),
            sum(s.get("total_tcs", 0) for s in all_subs),
            _pct(
                sum(s.get("score", 0) for s in all_subs),
                sum(s.get("total_tcs", 0) for s in all_subs) or 1
            ),
            f"{sum(1 for s in all_subs if 0 < s.get('score',0) < s.get('total_tcs',1))} partial",
            "",
            f"{len(all_subs)} total submissions",
        ]
        for ci, val in enumerate(totals, 1):
            c = ws.cell(row=summary_row, column=ci, value=val)
            _style(c, bg=_C["hdr_light"], bold=True, fg=_C["hdr_dark"])

    # ── Column widths ─────────────────────────────────────────────────────────
    static_widths = [8, 18, 30, 22, 16, 8, 10, 10, 10, 12, 35, 55]
    _set_col_widths(ws, static_widths)
    for tc_i in range(max_tcs):
        base = N_STATIC + tc_i * TC_COLS + 1
        ws.column_dimensions[get_column_letter(base    )].width = 11  # Status
        ws.column_dimensions[get_column_letter(base + 1)].width = 18  # Got
        ws.column_dimensions[get_column_letter(base + 2)].width = 14  # Expected

    # Freeze row 1 (title) and columns A+B (User ID / Problem ID) so the sheet
    # scrolls freely in both directions regardless of how many users are in the
    # summary section above the detail table.
    ws.freeze_panes = "C2"


# ─────────────────────────────────────────────────────────────────────────────
#  Metrics helpers for Sheet 4
# ─────────────────────────────────────────────────────────────────────────────

def _load_run_metrics(json_path: str) -> list:
    """Auto-detect and load the sibling _metrics.jsonl for a given JSON report path.
    Returns a list of snapshot dicts (may be empty if not found)."""
    p = Path(json_path)
    candidates = [
        p.parent / (p.stem + "_metrics.jsonl"),
        p.parent / (p.stem + ".metrics.jsonl"),
    ]
    for cand in candidates:
        if cand.exists():
            snapshots = []
            with open(cand, encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if line:
                        try:
                            snapshots.append(json.loads(line))
                        except json.JSONDecodeError:
                            pass
            print(f"  Metrics loaded : {cand}  ({len(snapshots)} snapshots)")
            return snapshots
    return []


def _safe_avg(vals):
    """Return rounded average or 'N/A' for an empty list."""
    nums = [v for v in vals if isinstance(v, (int, float))]
    if not nums:
        return "N/A"
    return round(sum(nums) / len(nums), 1)


def _safe_peak(vals):
    """Return rounded max or 'N/A' for an empty list."""
    nums = [v for v in vals if isinstance(v, (int, float))]
    if not nums:
        return "N/A"
    return round(max(nums), 1)


def _extract_metrics_summary(metrics_data: list) -> dict:
    """Compute peak / avg stats from a list of JSONL snapshots.

    Host metrics (cpu_pct, mem_*) come from kernel /proc files.
    Container metrics come from docker stats (cgroup-based).

    Returns a dict with the following keys (value is 'N/A' when not available):

      HOST (kernel /proc/stat, /proc/meminfo):
        host_cpu_avg, host_cpu_peak               — cpu_pct (%)  [avg skips first snapshot]
        host_ram_avg_gb, host_ram_peak_gb         — mem_used_gb
        host_ram_avg_pct, host_ram_peak_pct       — mem_pct (%)
        redis_queue_avg, redis_queue_peak         — redis_queue_depth

      CONTAINER — cgroup (docker stats, Judge0 sandbox workers only):
        worker_cg_cpu_avg, worker_cg_cpu_peak     — sum of judge0-workers-* cg_cpu_pct
        worker_cg_cpu_norm_avg, _norm_peak        — above ÷ n_cpus  (0–100% of host)
        worker_cg_mem_avg_mb, worker_cg_mem_peak_mb

      CONTAINER — cgroup (grading worker, separate from J0 sandboxes):
        grading_cg_cpu_avg, grading_cg_cpu_peak
        grading_cg_mem_avg_mb, grading_cg_mem_peak_mb

      CONTAINER — cgroup (all other containers: API, db, redis, reconciler):
        server_cg_cpu_avg, server_cg_cpu_peak
        server_cg_mem_avg_mb, server_cg_mem_peak_mb

      snapshots — number of snapshots loaded
    """
    _ALL_KEYS = [
        "host_cpu_avg","host_cpu_peak",
        "host_ram_avg_gb","host_ram_peak_gb",
        "host_ram_avg_pct","host_ram_peak_pct",
        "redis_queue_avg","redis_queue_peak",
        "worker_cg_cpu_avg","worker_cg_cpu_peak",
        "worker_cg_cpu_norm_avg","worker_cg_cpu_norm_peak",
        "worker_cg_mem_avg_mb","worker_cg_mem_peak_mb",
        "grading_cg_cpu_avg","grading_cg_cpu_peak",
        "grading_cg_mem_avg_mb","grading_cg_mem_peak_mb",
        "server_cg_cpu_avg","server_cg_cpu_peak",
        "server_cg_mem_avg_mb","server_cg_mem_peak_mb",
        "snapshots",
    ]
    if not metrics_data:
        return {k: "N/A" for k in _ALL_KEYS}

    # Snapshot[0] CPU delta covers only the seed→first-read window (milliseconds).
    # Use _first_snapshot flag when present; otherwise skip index 0 for averages.
    snaps_for_avg = [m for m in metrics_data if not m.get("_first_snapshot", False)]
    if not snaps_for_avg:
        snaps_for_avg = metrics_data  # no flag on any snapshot — use all

    # ── Host-level series ──────────────────────────────────────────────────
    # Peak uses ALL snapshots; avg uses snaps_for_avg (excludes warmup snap)
    host_cpu_all  = [m["cpu_pct"]     for m in metrics_data   if "cpu_pct"     in m]
    host_cpu_avg  = [m["cpu_pct"]     for m in snaps_for_avg  if "cpu_pct"     in m]
    host_ram_g    = [m["mem_used_gb"] for m in snaps_for_avg  if "mem_used_gb" in m]
    host_ram_p    = [m["mem_pct"]     for m in snaps_for_avg  if "mem_pct"     in m]
    redis_q       = [m["redis_queue_depth"] for m in metrics_data
                     if "redis_queue_depth" in m
                     and isinstance(m["redis_queue_depth"], (int, float))]

    # ── Per-snapshot container aggregates ──────────────────────────────────
    # Classify containers by name:
    #   judge0-workers-1/2/3  → "workers" in name  → J0 sandbox workers
    #   judge0-grading_*      → "grading" in name  → Flask grading worker
    #   everything else       → server group (API, db, redis, reconciler…)
    _N_CPUS_DEFAULT = 4  # c5.xlarge fallback for snapshots without n_cpus field

    j0_cg_cpu_per_snap      = []
    j0_cg_cpu_norm_per_snap = []  # ÷ n_cpus → 0–100% of host capacity
    j0_cg_mem_per_snap      = []
    gr_cg_cpu_per_snap      = []
    gr_cg_mem_per_snap      = []
    sv_cg_cpu_per_snap      = []
    sv_cg_mem_per_snap      = []

    for m in metrics_data:
        n_cpus = m.get("n_cpus", _N_CPUS_DEFAULT) or _N_CPUS_DEFAULT
        j0_cg_cpu = j0_cg_mem = 0.0
        gr_cg_cpu = gr_cg_mem = 0.0
        sv_cg_cpu = sv_cg_mem = 0.0
        for c in m.get("containers", []):
            name_lower   = c.get("name", "").lower()
            is_j0_worker = "workers" in name_lower
            is_grading   = not is_j0_worker and "grading" in name_lower
            cg_cpu = c.get("cg_cpu_pct", c.get("cpu_pct", 0)) or 0
            cg_mem = c.get("mem_mb", 0) or 0
            if is_j0_worker:
                j0_cg_cpu += cg_cpu
                j0_cg_mem += cg_mem
            elif is_grading:
                gr_cg_cpu += cg_cpu
                gr_cg_mem += cg_mem
            else:
                sv_cg_cpu += cg_cpu
                sv_cg_mem += cg_mem

        if m.get("containers"):
            j0_cg_cpu_per_snap.append(j0_cg_cpu)
            j0_cg_cpu_norm_per_snap.append(round(j0_cg_cpu / n_cpus, 2))
            j0_cg_mem_per_snap.append(j0_cg_mem)
            gr_cg_cpu_per_snap.append(gr_cg_cpu)
            gr_cg_mem_per_snap.append(gr_cg_mem)
            sv_cg_cpu_per_snap.append(sv_cg_cpu)
            sv_cg_mem_per_snap.append(sv_cg_mem)

    return {
        # Host — kernel /proc (avg excludes warmup first snapshot; peak covers all)
        "host_cpu_avg":            _safe_avg(host_cpu_avg),
        "host_cpu_peak":           _safe_peak(host_cpu_all),
        "host_ram_avg_gb":         _safe_avg(host_ram_g),
        "host_ram_peak_gb":        _safe_peak(host_ram_g),
        "host_ram_avg_pct":        _safe_avg(host_ram_p),
        "host_ram_peak_pct":       _safe_peak(host_ram_p),
        "redis_queue_avg":         _safe_avg(redis_q),
        "redis_queue_peak":        _safe_peak(redis_q),
        # J0 workers — cgroup (sum across 3 workers; raw can exceed 100%×N)
        "worker_cg_cpu_avg":       _safe_avg(j0_cg_cpu_per_snap),
        "worker_cg_cpu_peak":      _safe_peak(j0_cg_cpu_per_snap),
        "worker_cg_cpu_norm_avg":  _safe_avg(j0_cg_cpu_norm_per_snap),
        "worker_cg_cpu_norm_peak": _safe_peak(j0_cg_cpu_norm_per_snap),
        "worker_cg_mem_avg_mb":    _safe_avg(j0_cg_mem_per_snap),
        "worker_cg_mem_peak_mb":   _safe_peak(j0_cg_mem_per_snap),
        # Grading worker — cgroup (Flask async worker, NOT a J0 sandbox)
        "grading_cg_cpu_avg":      _safe_avg(gr_cg_cpu_per_snap),
        "grading_cg_cpu_peak":     _safe_peak(gr_cg_cpu_per_snap),
        "grading_cg_mem_avg_mb":   _safe_avg(gr_cg_mem_per_snap),
        "grading_cg_mem_peak_mb":  _safe_peak(gr_cg_mem_per_snap),
        # Server group — cgroup (API, db, redis, reconciler…)
        "server_cg_cpu_avg":       _safe_avg(sv_cg_cpu_per_snap),
        "server_cg_cpu_peak":      _safe_peak(sv_cg_cpu_per_snap),
        "server_cg_mem_avg_mb":    _safe_avg(sv_cg_mem_per_snap),
        "server_cg_mem_peak_mb":   _safe_peak(sv_cg_mem_per_snap),
        "snapshots":               len(metrics_data),
    }


# ─────────────────────────────────────────────────────────────────────────────
#  Sheet 4 — Scenario Comparison
# ─────────────────────────────────────────────────────────────────────────────

def build_sheet4_comparison(wb: openpyxl.Workbook, reports: list):
    """
    reports: list of (filename, report_dict) tuples — one per JSON file.
    Primary report is index 0; comparison runs follow.
    """
    ws = wb.create_sheet("Scenario Comparison")
    ws.sheet_view.showGridLines = False

    n_reports = len(reports)
    # Metric column + one column per report
    NCOLS = 1 + n_reports

    # ── Auto-load per-run metrics JSONL (sibling of each JSON file) ───────────
    run_metrics: list = []  # one _extract_metrics_summary dict per report
    for fname, _ in reports:
        snaps = _load_run_metrics(fname)
        run_metrics.append(_extract_metrics_summary(snaps))

    _write_title(ws, 1, "MULTI-RUN SCENARIO COMPARISON", NCOLS, fontsize=14)

    ts = _now_utc()
    c = ws.cell(row=2, column=1, value=f"Generated: {ts} UTC  |  {n_reports} run(s) compared")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    _style(c, bg=_C["light_grey"], size=10, halign="left")

    row = 3

    # ── Report File Headers ────────────────────────────────────────────────────
    c0 = ws.cell(row=row, column=1, value="Metric")
    _style(c0, bg=_C["hdr_mid"], fg="FFFFFF", bold=True, halign="center")

    for ri, (fname, rpt) in enumerate(reports, 2):
        cfg_r = rpt.get("config", {})
        label = f"{Path(fname).name}\n({_scenario_label(cfg_r)})"
        c = ws.cell(row=row, column=ri, value=label)
        _style(c, bg=_C["hdr_mid"], fg="FFFFFF", bold=True, halign="center")
        ws.row_dimensions[row].height = 32
    row += 1

    # ── Helper to write a section divider spanning all cols ───────────────────
    def section(title: str):
        nonlocal row
        _write_section(ws, row, title, NCOLS)
        row += 1

    # ── Helper to write one metric row ────────────────────────────────────────
    def metric_row(label: str, values: list, highlight_best: bool = False,
                   lower_is_better: bool = False, fmt=None):
        nonlocal row
        alt = (row % 2 == 0)
        bg  = _C["alt_row"] if alt else _C["white"]

        c0 = ws.cell(row=row, column=1, value=label)
        _style(c0, bg=bg, bold=True)

        # Find best value for optional highlighting
        numeric_vals = [v for v in values if isinstance(v, (int, float))]
        best = None
        if highlight_best and numeric_vals:
            best = min(numeric_vals) if lower_is_better else max(numeric_vals)

        for ri, val in enumerate(values, 2):
            disp = fmt(val) if fmt and isinstance(val, (int, float)) else val
            c = ws.cell(row=row, column=ri, value=disp)
            cell_bg = bg
            if highlight_best and val == best and isinstance(val, (int, float)):
                cell_bg = _C["pass"]
            _style(c, bg=cell_bg, halign="center")
        row += 1

    # ── Section: Run Configuration ────────────────────────────────────────────
    section("RUN CONFIGURATION")

    def cfg_vals(key, default="–"):
        return [r.get("config", {}).get(key, default) for _, r in reports]

    metric_row("Target URL",           cfg_vals("target"))
    metric_row("Mode",                 cfg_vals("mode"))
    metric_row("Virtual Users",        [r.get("config",{}).get("users",0) for _,r in reports],
               highlight_best=True)
    metric_row("Ramp-up (s)",          [r.get("config",{}).get("ramp_up_sec",0) for _,r in reports])
    metric_row("Batch Size",           [r.get("config",{}).get("batch_size",0) or "flat" for _,r in reports])
    metric_row("Concurrency Mode",     [_concurrency_label(r.get("config",{})) for _,r in reports])
    metric_row("Dry Run",              [r.get("config",{}).get("dry_run",False) for _,r in reports])

    # ── Section: Infrastructure Configuration ─────────────────────────────────
    section("INFRASTRUCTURE CONFIGURATION")

    def _workers(rpt):
        return rpt.get("config", {}).get("workers", rpt.get("_workers", 3))

    def _runners(rpt):
        return rpt.get("config", {}).get("runners", rpt.get("_runners", 2))

    metric_row("Judge0 Workers (--scale workers=N)",
               [_workers(r) for _, r in reports],
               highlight_best=True)
    metric_row("MAX_RUNNERS per worker",
               [_runners(r) for _, r in reports],
               highlight_best=True)
    metric_row("Total Concurrent Sandboxes",
               [_workers(r) * _runners(r) for _, r in reports],
               highlight_best=True)
    metric_row("Puma HTTP slots (WEB_CONCURRENCY × THREADS)",
               [r.get("config", {}).get("puma_slots", "–") for _, r in reports])

    section("THROUGHPUT & DURATION")

    metric_row("Test Duration (s)",
               [round(_duration(r.get("summary",{})), 1) for _,r in reports],
               highlight_best=True, lower_is_better=True)
    metric_row("Throughput (req/s)",
               [round(r.get("summary",{}).get("throughput_per_sec",0), 2) for _,r in reports],
               highlight_best=True, lower_is_better=False)
    metric_row("Total Submissions",
               [r.get("summary",{}).get("total_submissions",0) for _,r in reports])
    metric_row("Completed",
               [r.get("summary",{}).get("completed",0) for _,r in reports])
    metric_row("Peak In-Flight Jobs",
               [r.get("summary",{}).get("peak_in_flight_jobs",0) for _,r in reports],
               highlight_best=True, lower_is_better=True)

    # ── Section: Latency ──────────────────────────────────────────────────────
    section("LATENCY  (milliseconds)")

    def lat_vals(key):
        return [r.get("summary",{}).get(key, 0) for _,r in reports]

    def lat_from_results(pct_fn):
        out = []
        for _, rpt in reports:
            lats = sorted(
                res.get("latency_ms", 0)
                for res in rpt.get("results", [])
                if "latency_ms" in res
            )
            out.append(pct_fn(lats) if lats else 0)
        return out

    def _pct_val(lats, p):
        if not lats: return 0
        # Go-compatible nearest-rank: idx = int((n-1) * p / 100)
        idx = int((len(lats) - 1) * p / 100.0)
        return lats[idx]

    # All percentiles computed consistently from raw results (same Go formula)
    metric_row("P50  (median)",  lat_from_results(lambda l: _pct_val(l, 50)),
               highlight_best=True, lower_is_better=True)
    metric_row("P75",            lat_from_results(lambda l: _pct_val(l, 75)),
               highlight_best=True, lower_is_better=True)
    metric_row("P90",            lat_from_results(lambda l: _pct_val(l, 90)),
               highlight_best=True, lower_is_better=True)
    metric_row("P95",            lat_from_results(lambda l: _pct_val(l, 95)),
               highlight_best=True, lower_is_better=True)
    metric_row("P99",            lat_from_results(lambda l: _pct_val(l, 99)),
               highlight_best=True, lower_is_better=True)
    metric_row("P99.9",          lat_from_results(lambda l: _pct_val(l, 99.9)),
               highlight_best=True, lower_is_better=True)
    metric_row("Average",
               [round(sum(res.get("latency_ms",0) for res in rpt.get("results",[]))
                      / max(len(rpt.get("results",[])), 1), 1)
                for _, rpt in reports],
               highlight_best=True, lower_is_better=True)
    metric_row("Max",
               [max((res.get("latency_ms",0) for res in rpt.get("results",[])), default=0)
                for _, rpt in reports],
               highlight_best=True, lower_is_better=True)

    # ── Section: Outcomes ─────────────────────────────────────────────────────
    section("SUBMISSION OUTCOMES")

    def status_vals(status):
        return [r.get("status_counts",{}).get(status, 0) for _,r in reports]

    def status_pct_vals(status):
        return [
            _pct(r.get("status_counts",{}).get(status,0),
                 r.get("summary",{}).get("completed",1))
            for _,r in reports
        ]

    for status in ["PASS","FAIL","TLE","ERROR","BLOCKED","RATE_LIMITED","SYSTEM_ERROR","JUDGE0_ERROR"]:
        any_present = any(r.get("status_counts",{}).get(status,0) > 0 for _,r in reports)
        if any_present:
            metric_row(f"{status}  (count)", status_vals(status))
            metric_row(f"{status}  (%)",     status_pct_vals(status))

    # ── Section: Admission Control / Idempotency ──────────────────────────────
    section("ADMISSION CONTROL & IDEMPOTENCY")

    metric_row("Rate Limited (HTTP 429)",
               [r.get("summary",{}).get("rate_limited_429",0) for _,r in reports],
               highlight_best=True, lower_is_better=True)
    metric_row("Rate Limited %",
               [_pct(r.get("summary",{}).get("rate_limited_429",0),
                     r.get("summary",{}).get("total_submissions",1))
                for _,r in reports])
    metric_row("Duplicate (idempotent)",
               [r.get("summary",{}).get("duplicate_idempotent",0) for _,r in reports])
    metric_row("Duplicate %",
               [_pct(r.get("summary",{}).get("duplicate_idempotent",0),
                     r.get("summary",{}).get("total_submissions",1))
                for _,r in reports])
    metric_row("System Errors (infra)",
               [r.get("summary",{}).get("system_errors",0) for _,r in reports],
               highlight_best=True, lower_is_better=True)

    # ── Section: Score Distribution ───────────────────────────────────────────
    section("SCORE DISTRIBUTION")

    # Compute score distributions once (normalise to handle both snake_case and PascalCase JSONs)
    score_dists = [
        _score_distribution([_normalise_result(r) for r in rpt.get("results", [])])
        for _, rpt in reports
    ]
    completed_list = [r.get("summary",{}).get("completed",1) for _,r in reports]
    metric_row("Full Marks  (100%)",    [d[0] for d in score_dists], highlight_best=True)
    metric_row("Full Marks  % of completed",
               [_pct(d[0], completed_list[i]) for i, d in enumerate(score_dists)])
    metric_row("Partial Marks  (1–99%)", [d[1] for d in score_dists])
    metric_row("Zero Marks  (0%)",      [d[2] for d in score_dists], highlight_best=True, lower_is_better=True)

    # ── Section: Infrastructure Capacity ──────────────────────────────────────
    section("INFRASTRUCTURE CAPACITY")

    metric_row("Harness Jobs (actual)",
               [r.get("summary",{}).get("harness_jobs",0) for _,r in reports])
    metric_row("Batch-Equivalent Jobs",
               [r.get("summary",{}).get("batch_equivalent",0) for _,r in reports])
    metric_row("Queue Reduction Factor",
               [r.get("summary",{}).get("queue_reduction","N/A") for _,r in reports])

    # ── Section: Host System Resources (psutil — actual host) ─────────────────
    any_host_metrics = any(ms["snapshots"] != "N/A" and ms["snapshots"] > 0
                           for ms in run_metrics)
    if any_host_metrics:
        section("HOST SYSTEM RESOURCES  (kernel /proc/stat + /proc/meminfo — actual EC2 host)")

        def hm(key): return [ms[key] for ms in run_metrics]

        metric_row("Snapshots collected",
                   [ms["snapshots"] for ms in run_metrics])
        metric_row("Peak Host CPU %",
                   hm("host_cpu_peak"),    highlight_best=True, lower_is_better=True)
        metric_row("Avg  Host CPU %",
                   hm("host_cpu_avg"),     highlight_best=True, lower_is_better=True)
        metric_row("Peak RAM Used (GB)",
                   hm("host_ram_peak_gb"), highlight_best=True, lower_is_better=True)
        metric_row("Avg  RAM Used (GB)",
                   hm("host_ram_avg_gb"),  highlight_best=True, lower_is_better=True)
        metric_row("Peak RAM Used %",
                   hm("host_ram_peak_pct"),highlight_best=True, lower_is_better=True)
        metric_row("Avg  RAM Used %",
                   hm("host_ram_avg_pct"), highlight_best=True, lower_is_better=True)
        metric_row("Peak Redis Queue Depth",
                   hm("redis_queue_peak"), highlight_best=True, lower_is_better=True)
        metric_row("Avg  Redis Queue Depth",
                   hm("redis_queue_avg"),  highlight_best=True, lower_is_better=True)

        # ── Section: Container Resources — cgroup / docker stats ──────────────
        section("CONTAINER RESOURCES — cgroup  (docker stats, NOT real host)")

        metric_row("J0 Workers: Peak cg CPU %  (raw sum, can exceed 100%×N)",
                   hm("worker_cg_cpu_peak"), highlight_best=True, lower_is_better=True)
        metric_row("J0 Workers: Avg  cg CPU %  (raw sum)",
                   hm("worker_cg_cpu_avg"),  highlight_best=True, lower_is_better=True)
        metric_row("J0 Workers: Peak cg CPU % of host  (sum÷n_cpus, 0–100%)",
                   hm("worker_cg_cpu_norm_peak"), highlight_best=True, lower_is_better=True)
        metric_row("J0 Workers: Avg  cg CPU % of host  (sum÷n_cpus, 0–100%)",
                   hm("worker_cg_cpu_norm_avg"),  highlight_best=True, lower_is_better=True)
        metric_row("J0 Workers: Peak cg Mem MB",
                   hm("worker_cg_mem_peak_mb"), highlight_best=True, lower_is_better=True)
        metric_row("J0 Workers: Avg  cg Mem MB",
                   hm("worker_cg_mem_avg_mb"),  highlight_best=True, lower_is_better=True)
        metric_row("Grading Worker: Peak cg CPU %  (Flask async, not a J0 sandbox)",
                   hm("grading_cg_cpu_peak"), highlight_best=True, lower_is_better=True)
        metric_row("Grading Worker: Avg  cg CPU %",
                   hm("grading_cg_cpu_avg"),  highlight_best=True, lower_is_better=True)
        metric_row("Grading Worker: Peak cg Mem MB",
                   hm("grading_cg_mem_peak_mb"), highlight_best=True, lower_is_better=True)
        metric_row("Grading Worker: Avg  cg Mem MB",
                   hm("grading_cg_mem_avg_mb"),  highlight_best=True, lower_is_better=True)
        metric_row("Server Group:  Peak cg CPU %  (API, db, redis, reconciler)",
                   hm("server_cg_cpu_peak"), highlight_best=True, lower_is_better=True)
        metric_row("Server Group:  Avg  cg CPU %",
                   hm("server_cg_cpu_avg"),  highlight_best=True, lower_is_better=True)
        metric_row("Server Group:  Peak cg Mem MB",
                   hm("server_cg_mem_peak_mb"), highlight_best=True, lower_is_better=True)
        metric_row("Server Group:  Avg  cg Mem MB",
                   hm("server_cg_mem_avg_mb"),  highlight_best=True, lower_is_better=True)


    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 52  # wider for new resource metric labels
    col_w = max(20, int(90 / n_reports))
    for ri in range(2, n_reports + 2):
        ws.column_dimensions[get_column_letter(ri)].width = col_w

    ws.freeze_panes = "B4"


# ─────────────────────────────────────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate EC2 load-test Excel report from load_test_report.json",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "report",
        nargs="?",
        default="load_test_report.json",
        help="Path to load_test_report.json produced by the Go load tester",
    )
    parser.add_argument(
        "--metrics",
        default=None,
        metavar="FILE",
        help="Path to JSONL metrics file from collect_ec2_metrics.py",
    )
    parser.add_argument(
        "--out",
        default=None,
        metavar="FILE",
        help="Output .xlsx path (default: same stem as report)",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=None,
        help="Judge0 worker containers (--scale workers=N). "
             "Auto-read from the JSON report if not specified.",
    )
    parser.add_argument(
        "--runners",
        type=int,
        default=None,
        help="MAX_RUNNERS per worker container. "
             "Auto-read from the JSON report if not specified.",
    )
    parser.add_argument(
        "--compare",
        nargs="*",
        metavar="FILE",
        default=None,
        help=(
            "Add Sheet 4: multi-run scenario comparison. "
            "Pass specific JSON files to compare, or omit filenames to "
            "auto-detect all load_test_report*.json in the same directory."
        ),
    )
    parser.add_argument(
        "--all",
        nargs="+",
        metavar="FILE",
        default=None,
        help=(
            "All-files mode: pass every JSON file to compare. No 'primary' "
            "designation — all files appear equally in Sheet 4. "
            "The first file is used for Sheets 1-3 as a representative sample. "
            "Overrides the positional 'report' argument and --compare."
        ),
    )
    args = parser.parse_args()

    # ── --all mode: first file becomes primary, all go into Sheet 4 ──────────
    if args.all:
        args.report = args.all[0]
        # We'll handle compare_reports directly in the --all block below

    # ── Load primary report ───────────────────────────────────────────────────
    report_path = Path(args.report)
    if not report_path.exists():
        print(f"ERROR: Report file not found: {report_path}")
        sys.exit(1)

    print(f"Reading report  : {report_path}")
    with open(report_path, encoding="utf-8") as fh:
        report = json.load(fh)

    # workers/runners: CLI flag overrides JSON config; JSON config overrides default.
    cfg_block = report.get("config", {})
    workers = args.workers if args.workers is not None else cfg_block.get("workers", 3)
    runners = args.runners if args.runners is not None else cfg_block.get("runners", 2)
    report["_workers"] = workers
    report["_runners"] = runners
    print(f"Workers         : {workers}  runners/worker: {runners}  (sandboxes: {workers*runners})")

    # ── Load optional live metrics ────────────────────────────────────────────
    metrics_data: list[dict] = []
    if args.metrics:
        mpath = Path(args.metrics)
        if mpath.exists():
            print(f"Reading metrics : {mpath}")
            with open(mpath, encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if line:
                        try:
                            metrics_data.append(json.loads(line))
                        except json.JSONDecodeError:
                            pass
            print(f"  {len(metrics_data)} metric snapshots loaded")
        else:
            print(f"WARNING: Metrics file not found: {mpath}  (continuing without)")

    # ── Resolve comparison reports ────────────────────────────────────────────
    compare_reports: list = []   # list of (filename_str, report_dict)

    if args.all:
        # --all mode: load every file; first is already loaded as primary above
        for fpath_str in args.all[1:]:
            cpath = Path(fpath_str)
            if not cpath.exists():
                print(f"WARNING: File not found: {cpath}  (skipping)")
                continue
            try:
                with open(cpath, encoding="utf-8") as fh:
                    crpt = json.load(fh)
                compare_reports.append((str(cpath), crpt))
                print(f"  File : {cpath}")
            except Exception as exc:
                print(f"WARNING: Could not read {cpath}: {exc}  (skipping)")
    elif args.compare is not None:
        # --compare with explicit files
        if args.compare:
            compare_files = [Path(f) for f in args.compare]
        else:
            # Auto-detect all load_test_report*.json in same directory
            compare_files = sorted(
                report_path.parent.glob("load_test_report*.json")
            )

        for cpath in compare_files:
            if not cpath.exists():
                print(f"WARNING: Comparison file not found: {cpath}  (skipping)")
                continue
            try:
                with open(cpath, encoding="utf-8") as fh:
                    crpt = json.load(fh)
                compare_reports.append((str(cpath), crpt))
                print(f"  Comparison file : {cpath}")
            except Exception as exc:
                print(f"WARNING: Could not read {cpath}: {exc}  (skipping)")

        if not compare_reports:
            print("WARNING: No valid comparison files found; Sheet 4 will be skipped.")

    # ── Output path ────────────────────────────────────────────────────────────
    # If the JSON lives in a "Jsons" subdirectory (e.g. Reports/Jsons/X.json),
    # place the XLSX one level up (e.g. Reports/X.xlsx) unless --out is given.
    if args.out:
        out_path = Path(args.out)
    else:
        xlsx_name = report_path.with_suffix(".xlsx").name
        if report_path.parent.name.lower() == "jsons":
            out_path = report_path.parent.parent / xlsx_name
        else:
            out_path = report_path.with_suffix(".xlsx")

    # ── Build workbook ────────────────────────────────────────────────────────
    print("Building Excel workbook…")
    wb = openpyxl.Workbook()

    build_sheet1(wb, report, workers=workers, runners=runners)
    build_sheet2(wb, report, metrics_data)
    build_sheet3(wb, report)

    if compare_reports:
        # All files go into Sheet 4 equally; primary is always listed first
        all_for_compare = [(str(report_path), report)] + [
            (f, r) for f, r in compare_reports if str(f) != str(report_path)
        ]
        build_sheet4_comparison(wb, all_for_compare)
        print(f"  Sheet 4  — Scenario Comparison  ({len(all_for_compare)} runs)")

    wb.active = wb.worksheets[0]
    wb.save(out_path)

    # ── Summary ───────────────────────────────────────────────────────────────
    n_subs = len(report.get("results", []))
    if not n_subs:
        n_subs = sum(
            len(p.get("submissions", []))
            for p in report.get("per_problem_report", [])
        )

    print(f"\n✓  Report saved : {out_path}")
    print(f"   Sheet 1  — System Info")
    print(f"   Sheet 2  — Load Test Metrics"
          f"  ({len(metrics_data)} live-metric snapshots)")
    print(f"   Sheet 3  — User Test Case Results  ({n_subs} submissions)")


if __name__ == "__main__":
    main()
