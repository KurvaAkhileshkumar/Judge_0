#!/usr/bin/env python3
"""
generate_test_report.py
───────────────────────
Generates a colour-coded Excel workbook from test_harness.py JSON output.

Usage:
    python generate_test_report.py                              # reads test_harness_results.json
    python generate_test_report.py test_harness_results.json   # explicit path
    python generate_test_report.py --out my_report.xlsx        # custom output path
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabel
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ── Colour palette ─────────────────────────────────────────────────────────
C_PASS       = "FF27AE60"   # green
C_FAIL       = "FFE74C3C"   # red
C_HEADER_BG  = "FF2C3E50"   # dark navy
C_HEADER_FG  = "FFFFFFFF"   # white
C_ALT_ROW    = "FFF0F3F4"   # very light grey for alternating rows
C_SECTION    = "FF2980B9"   # section header blue
C_SECTION_FG = "FFFFFFFF"
C_LANG = {
    "python": "FF3498DB",   # blue
    "c":      "FFE67E22",   # orange
    "cpp":    "FF9B59B6",   # purple
    "java":   "FFE74C3C",   # red
}
C_MODE = {
    "function": "FF1ABC9C",  # teal
    "stdio":    "FFF39C12",  # amber
}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="FF000000", size=11):
    return Font(bold=bold, color=color, size=size)

def _border():
    s = Side(style="thin", color="FFBDC3C7")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _set(ws, row, col, value, bold=False, bg=None, fg="FF000000",
         align="left", border=True, size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, color=fg, size=size)
    cell.alignment = _center() if align == "center" else _left()
    if bg:
        cell.fill  = _fill(bg)
    if border:
        cell.border = _border()
    return cell

def _header_row(ws, row, labels, col_start=1):
    for i, lbl in enumerate(labels):
        _set(ws, row, col_start + i, lbl,
             bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")

def _col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── Sheet 1: Summary ───────────────────────────────────────────────────────

def build_summary(wb, data):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    r = 1
    # Title
    cell = ws.cell(row=r, column=1, value="Harness Test Suite — Summary Report")
    cell.font      = Font(bold=True, size=16, color=C_HEADER_BG[2:])
    cell.alignment = _left()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1

    cell2 = ws.cell(row=r, column=1,
                    value=f"Run at: {data['run_at']}   |   API: {data['api_url']}   |   Duration: {data['duration_s']}s")
    cell2.font      = Font(size=10, italic=True, color="FF7F8C8D")
    cell2.alignment = _left()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 2

    # ── KPI boxes ────────────────────────────────────────────────────────
    kpis = [
        ("Total Tests",  data["total"],  C_HEADER_BG),
        ("Passed",       data["passed"], C_PASS),
        ("Failed",       data["failed"], C_FAIL if data["failed"] else C_PASS),
        ("Pass Rate",    f"{data['passed']/max(data['total'],1)*100:.1f}%",
                         C_PASS if data["failed"] == 0 else C_FAIL),
    ]
    for i, (label, value, bg) in enumerate(kpis):
        col = 1 + i * 2
        label_cell = ws.cell(row=r, column=col, value=label)
        label_cell.font      = Font(bold=True, size=10, color="FFFFFFFF")
        label_cell.fill      = _fill(bg)
        label_cell.alignment = _center()
        label_cell.border    = _border()
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)

        val_cell = ws.cell(row=r+1, column=col, value=value)
        val_cell.font      = Font(bold=True, size=22, color=bg[2:])
        val_cell.alignment = _center()
        val_cell.border    = _border()
        ws.merge_cells(start_row=r+1, start_column=col, end_row=r+2, end_column=col+1)

    r += 4

    # ── By Language ────────────────────────────────────────────────────────
    r += 1
    sec = ws.cell(row=r, column=1, value="Results by Language")
    sec.font = Font(bold=True, size=12, color=C_SECTION[2:])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    _header_row(ws, r, ["Language", "Total", "Passed", "Failed", "Pass Rate"], col_start=1)
    r += 1

    lang_stats = {}
    for t in data["tests"]:
        lang = t["language"].upper()
        lang_stats.setdefault(lang, {"total": 0, "passed": 0})
        lang_stats[lang]["total"]  += 1
        lang_stats[lang]["passed"] += 1 if t["passed"] else 0

    for idx, (lang, s) in enumerate(sorted(lang_stats.items())):
        pct  = s["passed"] / max(s["total"], 1) * 100
        fail = s["total"] - s["passed"]
        bg   = C_ALT_ROW if idx % 2 else "FFFFFFFF"
        _set(ws, r, 1, lang,        bg=bg)
        _set(ws, r, 2, s["total"],  bg=bg, align="center")
        _set(ws, r, 3, s["passed"], bg=C_PASS if s["passed"] == s["total"] else bg, align="center")
        _set(ws, r, 4, fail,        bg=C_FAIL if fail > 0 else bg, align="center")
        _set(ws, r, 5, f"{pct:.0f}%",
             bg=C_PASS if fail == 0 else C_FAIL, fg="FFFFFFFF" if fail == 0 or pct == 100 else "FF000000",
             align="center", bold=True)
        r += 1

    r += 1

    # ── By Mode ──────────────────────────────────────────────────────────
    sec2 = ws.cell(row=r, column=1, value="Results by Mode")
    sec2.font = Font(bold=True, size=12, color=C_SECTION[2:])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    _header_row(ws, r, ["Mode", "Total", "Passed", "Failed", "Pass Rate"], col_start=1)
    r += 1

    mode_stats = {}
    for t in data["tests"]:
        mode = t["mode"].upper()
        mode_stats.setdefault(mode, {"total": 0, "passed": 0})
        mode_stats[mode]["total"]  += 1
        mode_stats[mode]["passed"] += 1 if t["passed"] else 0

    for idx, (mode, s) in enumerate(sorted(mode_stats.items())):
        pct  = s["passed"] / max(s["total"], 1) * 100
        fail = s["total"] - s["passed"]
        bg   = C_ALT_ROW if idx % 2 else "FFFFFFFF"
        _set(ws, r, 1, mode,        bg=bg)
        _set(ws, r, 2, s["total"],  bg=bg, align="center")
        _set(ws, r, 3, s["passed"], bg=C_PASS if fail == 0 else bg, align="center")
        _set(ws, r, 4, fail,        bg=C_FAIL if fail > 0 else bg, align="center")
        _set(ws, r, 5, f"{pct:.0f}%",
             bg=C_PASS if fail == 0 else C_FAIL, fg="FFFFFFFF", align="center", bold=True)
        r += 1

    r += 1

    # ── By Scenario type (expect) ────────────────────────────────────────
    sec3 = ws.cell(row=r, column=1, value="Results by Expected Scenario")
    sec3.font = Font(bold=True, size=12, color=C_SECTION[2:])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    _header_row(ws, r, ["Scenario", "Total", "Passed", "Failed"], col_start=1)
    r += 1

    expect_stats = {}
    for t in data["tests"]:
        expect_stats.setdefault(t["expect"], {"total": 0, "passed": 0})
        expect_stats[t["expect"]]["total"]  += 1
        expect_stats[t["expect"]]["passed"] += 1 if t["passed"] else 0

    for idx, (exp, s) in enumerate(sorted(expect_stats.items())):
        fail = s["total"] - s["passed"]
        bg   = C_ALT_ROW if idx % 2 else "FFFFFFFF"
        _set(ws, r, 1, exp,         bg=bg)
        _set(ws, r, 2, s["total"],  bg=bg, align="center")
        _set(ws, r, 3, s["passed"], bg=bg, align="center")
        _set(ws, r, 4, fail,        bg=C_FAIL if fail > 0 else bg, align="center")
        r += 1

    _col_widths(ws, {
        "A": 28, "B": 10, "C": 10, "D": 10, "E": 14,
        "F": 14, "G": 14, "H": 14,
    })

    ws.freeze_panes = "A3"
    return ws


# ── Sheet 2: Detailed Results ──────────────────────────────────────────────

def build_detail(wb, data):
    ws = wb.create_sheet("Test Results")
    ws.sheet_view.showGridLines = False

    r = 1
    cell = ws.cell(row=r, column=1, value="Harness Test Suite — Detailed Results")
    cell.font      = Font(bold=True, size=14, color=C_HEADER_BG[2:])
    cell.alignment = _left()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 2

    headers = ["#", "Test Name", "Language", "Mode", "Expected Outcome",
               "Result", "Time (s)", "Status"]
    _header_row(ws, r, headers)
    r += 1

    for t in data["tests"]:
        idx   = t["index"]
        bg    = C_ALT_ROW if idx % 2 == 0 else "FFFFFFFF"
        ok    = t["passed"]
        result_text = "PASS" if ok else "FAIL"
        result_bg   = C_PASS if ok else C_FAIL

        _set(ws, r, 1, idx,              bg=bg, align="center")
        _set(ws, r, 2, t["name"],        bg=bg)
        _set(ws, r, 3, t["language"].upper(),
             bg=C_LANG.get(t["language"], bg), fg="FFFFFFFF", align="center", bold=True)
        _set(ws, r, 4, t["mode"].upper(),
             bg=C_MODE.get(t["mode"], bg), fg="FFFFFFFF", align="center", bold=True)
        _set(ws, r, 5, t["expect"],      bg=bg, align="center")
        _set(ws, r, 6, result_text,
             bg=result_bg, fg="FFFFFFFF", align="center", bold=True)
        _set(ws, r, 7, t["elapsed_s"],   bg=bg, align="center")

        # Status icon column
        icon = "✓" if ok else "✗"
        icon_cell = ws.cell(row=r, column=8, value=icon)
        icon_cell.font      = Font(bold=True, size=14,
                                   color=C_PASS[2:] if ok else C_FAIL[2:])
        icon_cell.alignment = _center()
        icon_cell.fill      = _fill(bg)
        icon_cell.border    = _border()

        r += 1

    # Totals row
    passed = data["passed"]
    failed = data["failed"]
    _set(ws, r, 1, "TOTALS", bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    _set(ws, r, 6, f"{passed} PASS / {failed} FAIL",
         bold=True, bg=C_PASS if failed == 0 else C_FAIL, fg="FFFFFFFF", align="center")
    _set(ws, r, 7, round(sum(t["elapsed_s"] for t in data["tests"]), 1),
         bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")
    _set(ws, r, 8, f"{passed/max(data['total'],1)*100:.1f}%",
         bold=True, bg=C_PASS if failed == 0 else C_FAIL, fg="FFFFFFFF", align="center")

    _col_widths(ws, {
        "A": 5, "B": 52, "C": 12, "D": 12, "E": 22,
        "F": 12, "G": 10, "H": 8,
    })
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"
    return ws


# ── Sheet 3: Language × Mode Matrix ──────────────────────────────────────

def build_matrix(wb, data):
    ws = wb.create_sheet("Language × Mode")
    ws.sheet_view.showGridLines = False

    r = 1
    cell = ws.cell(row=r, column=1, value="Pass/Fail Matrix — Language × Mode")
    cell.font      = Font(bold=True, size=13, color=C_HEADER_BG[2:])
    cell.alignment = _left()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 2

    # Build matrix
    languages = sorted({t["language"].upper() for t in data["tests"]})
    modes     = sorted({t["mode"].upper()     for t in data["tests"]})
    matrix    = {(l, m): {"passed": 0, "total": 0}
                 for l in languages for m in modes}
    for t in data["tests"]:
        key = (t["language"].upper(), t["mode"].upper())
        if key in matrix:
            matrix[key]["total"]  += 1
            matrix[key]["passed"] += 1 if t["passed"] else 0

    # Header
    _set(ws, r, 1, "Language \\ Mode", bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")
    for j, mode in enumerate(modes):
        _set(ws, r, 2 + j*3,     f"{mode}", bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")
        _set(ws, r, 2 + j*3 + 1, "Pass",   bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")
        _set(ws, r, 2 + j*3 + 2, "Fail",   bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")
    r += 1

    for i, lang in enumerate(languages):
        lang_color = C_LANG.get(lang.lower(), C_HEADER_BG)
        bg = C_ALT_ROW if i % 2 else "FFFFFFFF"
        _set(ws, r, 1, lang, bold=True, bg=lang_color, fg="FFFFFFFF", align="center")
        for j, mode in enumerate(modes):
            s    = matrix[(lang, mode)]
            p    = s["passed"]
            f    = s["total"] - p
            pct_bg = C_PASS if f == 0 and p > 0 else (C_FAIL if f > 0 else bg)
            _set(ws, r, 2 + j*3,
                 f"{p}/{s['total']}", bold=True, bg=pct_bg, fg="FFFFFFFF", align="center")
            _set(ws, r, 2 + j*3 + 1, p, bg=C_PASS if p > 0 else bg, fg="FFFFFFFF" if p > 0 else "FF000000", align="center")
            _set(ws, r, 2 + j*3 + 2, f, bg=C_FAIL if f > 0 else bg, fg="FFFFFFFF" if f > 0 else "FF000000", align="center")
        r += 1

    _col_widths(ws, {
        "A": 14, "B": 14, "C": 10, "D": 10,
        "E": 14, "F": 10, "G": 10,
    })
    ws.freeze_panes = "B3"
    return ws


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate Excel report from test_harness.py JSON output")
    parser.add_argument("json_file", nargs="?", default="test_harness_results.json")
    parser.add_argument("--out", default=None)
    args = parser.parse_args()

    json_path = Path(args.json_file)
    if not json_path.exists():
        print(f"ERROR: {json_path} not found. Run test_harness.py first.")
        sys.exit(1)

    with open(json_path) as fh:
        data = json.load(fh)

    out_path = Path(args.out) if args.out else json_path.with_suffix(".xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)          # remove default sheet

    build_summary(wb, data)
    build_detail(wb, data)
    build_matrix(wb, data)

    wb.save(out_path)
    print(f"Excel report saved → {out_path}")
    print(f"  Sheets: Summary | Test Results | Language × Mode")
    print(f"  Total: {data['passed']}/{data['total']} passed  "
          f"({'ALL PASS' if data['failed']==0 else str(data['failed'])+' FAILED'})")


if __name__ == "__main__":
    main()
