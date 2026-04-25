"""
load_test_pipelined_100pct.py
══════════════════════════════════════════════════════════════════════════════
Pipelined load test — 100% acceptance target
  • 20 batches × 10 users = 200 total submissions
  • ONLY "accepted" solutions → every test case must PASS
  • All 20 batches are fired in parallel (no waiting between batches)
  • Within each batch all 10 users are submitted concurrently
  • Full telemetry: batch-level, user-level, and per-TC level
  • Output: JSON report  +  Excel workbook with analytics

System configuration snapshot printed at startup.
══════════════════════════════════════════════════════════════════════════════
"""

from __future__ import annotations

import base64
import json
import os
import re
import sys
import time
import uuid
import threading
import traceback
import subprocess
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.series import DataPoint

# ── Constants ──────────────────────────────────────────────────────────────

JUDGE0_URL     = "http://localhost:2358"
QUESTION_BANK  = "question_bank.json"
HARNESS_FILE   = "harnesses/python_harness.py"

NUM_BATCHES    = 20
USERS_PER_BATCH = 10

# Concurrency throttle — Judge0 server has RAILS_MAX_THREADS=20
# We allow up to 20 simultaneous HTTP requests (submit + poll combined)
HTTP_SEMAPHORE  = threading.Semaphore(20)

POLL_INTERVAL_S = 0.8
MAX_POLL_TRIES  = 80    # 80 × 0.8s = 64s max per submission

REPORT_DIR = Path("batch_reports/pipelined_100pct")

JUDGE0_STATUS = {
    1: "In Queue", 2: "Processing", 3: "Accepted",
    4: "Wrong Answer", 5: "TLE", 6: "Compilation Error",
    7: "RE (SIGSEGV)", 8: "RE (SIGFPE)", 9: "RE (SIGABRT)",
    10: "RE (NZEC)", 11: "RE (Other)", 12: "Internal Error",
    13: "Exec Format Error",
}

# ── Colours for Excel ──────────────────────────────────────────────────────
CLR_HEADER_DARK  = "1F3864"   # dark navy
CLR_HEADER_MED   = "2E75B6"   # medium blue
CLR_HEADER_LIGHT = "BDD7EE"   # light blue
CLR_PASS         = "C6EFCE"   # green fill
CLR_PASS_FONT    = "276221"
CLR_FAIL         = "FFC7CE"   # red fill
CLR_FAIL_FONT    = "9C0006"
CLR_WARN         = "FFEB9C"   # yellow
CLR_WARN_FONT    = "9C5700"
CLR_WHITE        = "FFFFFF"
CLR_ROW_ALT      = "EEF3FB"   # alternating row colour

# ══════════════════════════════════════════════════════════════════════════
# Data structures
# ══════════════════════════════════════════════════════════════════════════

@dataclass
class TCRecord:
    tc_id:       str
    description: str
    stdin:       str
    expected:    str
    status:      str   # PASS | FAIL | TLE | ERROR | MISSING
    got:         str   = ""
    detail:      str   = ""

@dataclass
class UserRecord:
    batch_id:      int
    user_id:       int          # global 1-200
    user_in_batch: int          # 1-10
    problem_id:    str
    problem_title: str
    solution_id:   str
    solution_type: str          # always "accepted"
    harness_status: str         # PASS if score == total
    judge0_status:  str
    score:         int
    total_tcs:     int
    latency_ms:    int
    judge0_time_s: float
    judge0_mem_kb: int
    global_tle:    bool
    error:         str
    tc_results:    list[TCRecord] = field(default_factory=list)

    @property
    def pass_rate(self) -> float:
        return self.score / self.total_tcs if self.total_tcs else 0.0

    @property
    def all_pass(self) -> bool:
        return self.score == self.total_tcs and self.total_tcs > 0


@dataclass
class BatchRecord:
    batch_id:       int
    batch_start_ts: float   # epoch seconds
    batch_end_ts:   float
    users:          list[UserRecord] = field(default_factory=list)

    @property
    def duration_s(self) -> float:
        return self.batch_end_ts - self.batch_start_ts

    @property
    def total_users(self) -> int:
        return len(self.users)

    @property
    def passed_users(self) -> int:
        return sum(1 for u in self.users if u.all_pass)

    @property
    def acceptance_rate(self) -> float:
        return self.passed_users / self.total_users if self.total_users else 0.0

    @property
    def avg_latency_ms(self) -> float:
        lats = [u.latency_ms for u in self.users if u.latency_ms > 0]
        return sum(lats) / len(lats) if lats else 0.0

    @property
    def max_latency_ms(self) -> int:
        lats = [u.latency_ms for u in self.users if u.latency_ms > 0]
        return max(lats) if lats else 0

    @property
    def min_latency_ms(self) -> int:
        lats = [u.latency_ms for u in self.users if u.latency_ms > 0]
        return min(lats) if lats else 0

    @property
    def errors(self) -> int:
        return sum(1 for u in self.users if u.error)


# ══════════════════════════════════════════════════════════════════════════
# System info
# ══════════════════════════════════════════════════════════════════════════

def gather_system_info() -> dict:
    info: dict[str, Any] = {}

    # Docker container info
    try:
        out = subprocess.check_output(
            ["docker", "ps", "--format",
             "{{.Names}}\t{{.Status}}\t{{.Ports}}"],
            stderr=subprocess.DEVNULL, text=True
        )
        info["docker_containers"] = [
            dict(zip(["name", "status", "ports"], line.split("\t")))
            for line in out.strip().splitlines() if line
        ]
    except Exception:
        info["docker_containers"] = []

    # Judge0 system info
    try:
        r = requests.get(f"{JUDGE0_URL}/system_info", timeout=5)
        info["judge0_system_info"] = r.json()
    except Exception:
        info["judge0_system_info"] = {}

    # Judge0 config (from file)
    try:
        conf = {}
        for line in Path("judge0.conf").read_text().splitlines():
            if "=" in line and not line.startswith("#"):
                k, _, v = line.partition("=")
                # Mask passwords
                if any(s in k.lower() for s in ["password", "secret", "key"]):
                    v = "***"
                conf[k.strip()] = v.strip()
        info["judge0_conf"] = conf
    except Exception:
        info["judge0_conf"] = {}

    # docker-compose relevant settings
    try:
        dc = Path("docker-compose.yml").read_text()
        info["docker_compose_highlights"] = {
            "server_image": "judge0/judge0:1.13.0",
            "rails_max_threads": 20,
            "web_concurrency": 1,
            "max_queue_size": 5000,
            "max_runners": 50,
            "postgres_max_connections": 200,
        }
    except Exception:
        info["docker_compose_highlights"] = {}

    info["test_config"] = {
        "num_batches":        NUM_BATCHES,
        "users_per_batch":    USERS_PER_BATCH,
        "total_users":        NUM_BATCHES * USERS_PER_BATCH,
        "solution_type":      "accepted_only",
        "target_acceptance":  "100%",
        "pipeline_mode":      "all_batches_parallel",
        "poll_interval_s":    POLL_INTERVAL_S,
        "max_poll_tries":     MAX_POLL_TRIES,
        "http_concurrency":   20,
        "judge0_url":         JUDGE0_URL,
        "run_timestamp":      datetime.now().isoformat(),
    }

    return info


# ══════════════════════════════════════════════════════════════════════════
# Question bank helpers
# ══════════════════════════════════════════════════════════════════════════

def load_question_bank() -> list[dict]:
    with open(QUESTION_BANK) as f:
        data = json.load(f)
    return data["problems"]


def accepted_solutions(problems: list[dict]) -> list[tuple[dict, dict]]:
    """Return list of (problem, solution) pairs for accepted solutions only."""
    pairs = []
    for prob in problems:
        for sol in prob["solutions"]:
            if sol["type"] == "accepted":
                pairs.append((prob, sol))
                break   # one accepted solution per problem
    return pairs


# ══════════════════════════════════════════════════════════════════════════
# Harness builder (Python stdio mode)
# ══════════════════════════════════════════════════════════════════════════

def build_harness(problem: dict, solution: dict, session_id: str) -> str:
    template = Path(HARNESS_FILE).read_text()

    tc_dicts = [
        {"stdin_text": tc["stdin_text"], "expected": tc["expected"]}
        for tc in problem["test_cases"]
    ]

    student_escaped = (
        solution["source_code"]
        .replace("{", "{{")
        .replace("}", "}}")
    )
    student_b64 = base64.b64encode(
        solution["source_code"].encode()
    ).decode()
    student_repr = f"__import__('base64').b64decode('{student_b64}').decode()"

    filled = (
        template
        .replace("{SESSION_ID}",      session_id)
        .replace("{STUDENT_CODE}",    student_escaped)
        .replace("{TC_LIST_JSON}",    json.dumps(tc_dicts))
        .replace("{PER_TC_LIMIT_S}",  str(problem["per_tc_limit_s"]))
        .replace("{MEMORY_LIMIT_MB}", str(problem["memory_limit_mb"]))
    )

    # Also handle the actual template variables used in python_harness.py
    # (check what placeholders the harness uses)
    return filled


def build_harness_direct(problem: dict, solution: dict, session_id: str) -> str:
    """Build harness inline without relying on placeholder names."""
    tc_dicts = [
        {"stdin_text": tc["stdin_text"], "expected": tc["expected"]}
        for tc in problem["test_cases"]
    ]
    tc_json = json.dumps(tc_dicts)
    per_tc  = problem["per_tc_limit_s"]
    delim   = f"@@TC_RESULT__{session_id}__"
    student_b64 = base64.b64encode(solution["source_code"].encode()).decode()

    return f'''import signal, sys, io, os, select, time, traceback, json, builtins, base64

MODE  = "stdio"
DELIM = "{delim}"

_real_open   = open
_real_signal = signal.signal
_HARNESS_FILE = __file__

def _safe_open(file, mode="r", *a, **kw):
    try:
        if os.path.abspath(str(file)) == os.path.abspath(_HARNESS_FILE):
            raise PermissionError("Access denied")
    except PermissionError:
        raise
    except Exception:
        raise PermissionError("Access denied")
    return _real_open(file, mode, *a, **kw)

builtins.open = _safe_open

_sig_override = [False]
def _safe_signal(signum, handler):
    if signum == signal.SIGALRM:
        _sig_override[0] = True
        return
    return _real_signal(signum, handler)
signal.signal = _safe_signal

def _safe_exit(*a): raise SystemExit("__HARNESS_BLOCKED__")
sys.exit = builtins.exit = builtins.quit = _safe_exit

_STUDENT_SOURCE = base64.b64decode('{student_b64}').decode()

def _run_tc_child(tc, per_tc_limit_s, write_fd):
    class _TLE(Exception): pass
    def _tle(s, f): raise _TLE()
    fake_in  = io.StringIO(tc["stdin_text"])
    fake_out = io.StringIO()
    sys.stdin  = fake_in
    sys.stdout = fake_out
    _real_signal(signal.SIGALRM, _tle)
    signal.alarm(per_tc_limit_s)
    try:
        ns = {{"__name__": "__main__", "open": _safe_open, "exit": _safe_exit, "quit": _safe_exit}}
        exec(compile(_STUDENT_SOURCE, "<student>", "exec"), ns)
        signal.alarm(0)
        got      = fake_out.getvalue().strip()
        expected = str(tc["expected"]).strip()
        result   = {{"status": "PASS" if got == expected else "FAIL",
                    "got": got, "expected": expected}}
        if _sig_override[0]:
            result["warning"] = "signal_override_attempted"
    except _TLE:
        result = {{"status": "TLE", "detail": f"Exceeded {{per_tc_limit_s}}s"}}
    except SystemExit as e:
        signal.alarm(0)
        msg = "Called sys.exit()" if "__HARNESS_BLOCKED__" in str(e) else "SystemExit"
        result = {{"status": "ERROR", "detail": msg}}
    except Exception:
        signal.alarm(0)
        lines = traceback.format_exc().strip().splitlines()
        result = {{"status": "ERROR", "detail": " | ".join(lines[-2:])}}
    try:
        os.write(write_fd, json.dumps(result).encode())
        os.close(write_fd)
    except OSError:
        pass
    os._exit(0)

def _run_all(test_cases, per_tc_limit_s):
    _real_out = sys.stdout
    pids, fds = [], []
    for i, tc in enumerate(test_cases):
        try:
            r, w = os.pipe()
        except OSError:
            pids.append(None); fds.append((i, None)); continue
        try:
            pid = os.fork()
        except OSError:
            os.close(r); os.close(w)
            pids.append(None); fds.append((i, None)); continue
        if pid == 0:
            os.close(r)
            _run_tc_child(tc, per_tc_limit_s, w)
        os.close(w)
        pids.append(pid)
        fds.append((i, r))
    results  = [None] * len(test_cases)
    open_fds = [r for _, r in fds if r is not None]
    fd_idx   = {{r: i for i, r in fds if r is not None}}
    bufs     = {{r: b"" for r in open_fds}}
    t0       = time.monotonic()
    while open_fds:
        remaining = per_tc_limit_s + 1.5 - (time.monotonic() - t0)
        if remaining <= 0:
            break
        try:
            ready, _, _ = select.select(open_fds, [], [], remaining)
        except (ValueError, OSError):
            break
        if not ready:
            break
        for fd in list(ready):
            try:
                chunk = os.read(fd, 65536)
            except OSError:
                chunk = b""
            if chunk:
                bufs[fd] += chunk
            else:
                idx = fd_idx[fd]
                try:
                    results[idx] = json.loads(bufs[fd].decode())
                except Exception:
                    results[idx] = {{"status": "ERROR", "detail": "result parse error"}}
                try: os.close(fd)
                except OSError: pass
                open_fds.remove(fd)
    for pid in pids:
        if pid is not None:
            try: os.waitpid(pid, os.WNOHANG)
            except OSError: pass
    for i in range(len(results)):
        if results[i] is None:
            results[i] = {{"status": "TLE", "detail": f"Exceeded {{per_tc_limit_s}}s"}}
    for i, r in enumerate(results):
        _real_out.write(f"{{DELIM}}START_{{i+1}}\\n")
        _real_out.write(json.dumps(r) + "\\n")
        _real_out.write(f"{{DELIM}}END_{{i+1}}\\n")
        _real_out.flush()
    _real_out.write(f"{{DELIM}}DONE\\n")
    _real_out.flush()

_TEST_CASES     = {tc_json}
_PER_TC_LIMIT_S = {per_tc}

_run_all(_TEST_CASES, _PER_TC_LIMIT_S)
'''


# ══════════════════════════════════════════════════════════════════════════
# Output parser
# ══════════════════════════════════════════════════════════════════════════

def parse_harness_output(
    stdout: str,
    session_id: str,
    test_cases: list[dict],
) -> tuple[list[TCRecord], bool, int]:
    """
    Returns (tc_results, global_tle, score).
    """
    delim       = f"@@TC_RESULT__{session_id}__"
    done_marker = f"{delim}DONE"
    global_tle  = done_marker not in stdout

    n = len(test_cases)
    tc_records: list[TCRecord | None] = [None] * n

    for i in range(1, n + 1):
        start_m = f"{delim}START_{i}\n"
        end_m   = f"{delim}END_{i}\n"
        si = stdout.find(start_m)
        if si == -1:
            continue
        si += len(start_m)
        ei = stdout.find(end_m, si)
        if ei == -1:
            continue
        content = stdout[si:ei].strip()
        try:
            d = json.loads(content)
        except Exception:
            d = {"status": "ERROR", "detail": "JSON parse error"}

        tc = test_cases[i - 1]
        tc_records[i - 1] = TCRecord(
            tc_id       = tc.get("id", f"tc{i}"),
            description = tc.get("description", ""),
            stdin       = tc.get("stdin_text", ""),
            expected    = tc.get("expected", ""),
            status      = d.get("status", "UNKNOWN"),
            got         = d.get("got", ""),
            detail      = d.get("detail", ""),
        )

    # Fill nulls
    results: list[TCRecord] = []
    for i, rec in enumerate(tc_records):
        tc = test_cases[i]
        if rec is None:
            results.append(TCRecord(
                tc_id       = tc.get("id", f"tc{i+1}"),
                description = tc.get("description", ""),
                stdin       = tc.get("stdin_text", ""),
                expected    = tc.get("expected", ""),
                status      = "TLE" if global_tle else "MISSING",
                detail      = "not reached",
            ))
        else:
            results.append(rec)

    score = sum(1 for r in results if r.status == "PASS")
    return results, global_tle, score


# ══════════════════════════════════════════════════════════════════════════
# Judge0 client
# ══════════════════════════════════════════════════════════════════════════

def j0_submit(harness: str, cpu_limit_s: int, mem_limit_mb: int) -> str:
    encoded = base64.b64encode(harness.encode()).decode()
    payload = {
        "source_code":   encoded,
        "language_id":   71,
        "stdin":         "",
        "cpu_time_limit":  cpu_limit_s,
        "wall_time_limit": cpu_limit_s + 2,
        "memory_limit":    mem_limit_mb * 1024,
        "base64_encoded":  True,
        "enable_per_process_and_thread_time_limit":   True,
        "enable_per_process_and_thread_memory_limit": True,
    }
    with HTTP_SEMAPHORE:
        r = requests.post(
            f"{JUDGE0_URL}/submissions?base64_encoded=true&wait=false",
            json=payload,
            timeout=15,
        )
    r.raise_for_status()
    return r.json()["token"]


def j0_poll(token: str) -> dict:
    for attempt in range(MAX_POLL_TRIES):
        time.sleep(POLL_INTERVAL_S)
        with HTTP_SEMAPHORE:
            r = requests.get(
                f"{JUDGE0_URL}/submissions/{token}?base64_encoded=true",
                timeout=15,
            )
        r.raise_for_status()
        data = r.json()
        status_id = data["status"]["id"]
        if status_id <= 2:
            continue   # In Queue or Processing
        return data
    raise TimeoutError(f"Poll timeout after {MAX_POLL_TRIES} attempts")


def decode_b64(val: str | None) -> str:
    if not val:
        return ""
    try:
        return base64.b64decode(val).decode(errors="replace")
    except Exception:
        return val


# ══════════════════════════════════════════════════════════════════════════
# Single user execution
# ══════════════════════════════════════════════════════════════════════════

def run_user(
    batch_id:      int,
    user_in_batch: int,
    global_user_id: int,
    problem:       dict,
    solution:      dict,
) -> UserRecord:
    session_id = uuid.uuid4().hex[:12]
    start_ts   = time.monotonic()

    rec = UserRecord(
        batch_id       = batch_id,
        user_id        = global_user_id,
        user_in_batch  = user_in_batch,
        problem_id     = problem["id"],
        problem_title  = problem["title"],
        solution_id    = solution["id"],
        solution_type  = solution["type"],
        harness_status = "UNKNOWN",
        judge0_status  = "",
        score          = 0,
        total_tcs      = len(problem["test_cases"]),
        latency_ms     = 0,
        judge0_time_s  = 0.0,
        judge0_mem_kb  = 0,
        global_tle     = False,
        error          = "",
    )

    try:
        harness = build_harness_direct(problem, solution, session_id)

        cpu_limit = problem["per_tc_limit_s"] + 5   # parallel: max 1 TC + overhead
        token = j0_submit(harness, cpu_limit, problem["memory_limit_mb"])

        j0data  = j0_poll(token)
        stdout  = decode_b64(j0data.get("stdout"))
        stderr  = decode_b64(j0data.get("stderr"))

        j0_status_id   = j0data["status"]["id"]
        rec.judge0_status = j0data["status"]["description"]

        # Parse time / memory
        try:
            rec.judge0_time_s = float(j0data.get("time") or 0)
        except (TypeError, ValueError):
            rec.judge0_time_s = 0.0
        rec.judge0_mem_kb = j0data.get("memory") or 0

        tc_results, global_tle, score = parse_harness_output(
            stdout, session_id, problem["test_cases"]
        )
        rec.tc_results  = tc_results
        rec.global_tle  = global_tle
        rec.score       = score
        rec.harness_status = "PASS" if score == rec.total_tcs else "PARTIAL"

        if j0_status_id == 6:   # Compilation Error
            rec.harness_status = "COMPILE_ERROR"
            rec.error = decode_b64(j0data.get("compile_output")) or stderr

        elif j0_status_id >= 7:  # Runtime errors from Judge0 level
            rec.harness_status = "JUDGE0_ERROR"
            rec.error = f"Judge0 status {j0_status_id}: {rec.judge0_status}"

    except Exception as exc:
        rec.error          = f"{type(exc).__name__}: {exc}"
        rec.harness_status = "JUDGE0_ERROR"
        traceback.print_exc(file=sys.stderr)

    rec.latency_ms = int((time.monotonic() - start_ts) * 1000)
    return rec


# ══════════════════════════════════════════════════════════════════════════
# Batch execution
# ══════════════════════════════════════════════════════════════════════════

def run_batch(
    batch_id:     int,
    user_slots:   list[tuple[int, dict, dict]],   # (global_uid, problem, solution)
) -> BatchRecord:
    """
    Submit all users in this batch concurrently.
    Returns a BatchRecord with all results.
    """
    print(f"  [Batch {batch_id:02d}] FIRED  @ {datetime.now().strftime('%H:%M:%S.%f')[:-3]}"
          f"  ({len(user_slots)} users)", flush=True)

    batch_start = time.monotonic()
    batch       = BatchRecord(
        batch_id       = batch_id,
        batch_start_ts = time.time(),
        batch_end_ts   = 0.0,
    )

    # Submit all users in this batch in parallel via thread pool
    with ThreadPoolExecutor(max_workers=USERS_PER_BATCH) as pool:
        futures = {
            pool.submit(
                run_user,
                batch_id,
                idx + 1,      # user_in_batch (1-based)
                global_uid,
                problem,
                solution,
            ): idx
            for idx, (global_uid, problem, solution) in enumerate(user_slots)
        }
        for fut in as_completed(futures):
            try:
                rec = fut.result()
            except Exception as exc:
                # Shouldn't happen — run_user catches internally
                rec = UserRecord(
                    batch_id       = batch_id,
                    user_id        = -1,
                    user_in_batch  = futures[fut] + 1,
                    problem_id     = "unknown",
                    problem_title  = "unknown",
                    solution_id    = "unknown",
                    solution_type  = "accepted",
                    harness_status = "JUDGE0_ERROR",
                    judge0_status  = "",
                    score          = 0,
                    total_tcs      = 0,
                    latency_ms     = 0,
                    judge0_time_s  = 0.0,
                    judge0_mem_kb  = 0,
                    global_tle     = False,
                    error          = str(exc),
                )
            batch.users.append(rec)

    batch.batch_end_ts = time.time()
    duration = time.monotonic() - batch_start

    passed = batch.passed_users
    total  = batch.total_users
    pct    = batch.acceptance_rate * 100
    print(
        f"  [Batch {batch_id:02d}] DONE   @ {datetime.now().strftime('%H:%M:%S.%f')[:-3]}"
        f"  duration={duration:.2f}s  "
        f"passed={passed}/{total} ({pct:.0f}%)  "
        f"avg_latency={batch.avg_latency_ms:.0f}ms",
        flush=True
    )
    return batch


# ══════════════════════════════════════════════════════════════════════════
# Main orchestrator
# ══════════════════════════════════════════════════════════════════════════

def orchestrate() -> tuple[list[BatchRecord], dict]:
    problems = load_question_bank()
    pairs    = accepted_solutions(problems)   # (problem, solution) accepted only

    print("\n" + "═" * 72)
    print("  JUDGE0 PIPELINED LOAD TEST — 100% ACCEPTANCE TARGET")
    print("═" * 72)
    print(f"  Batches:          {NUM_BATCHES}")
    print(f"  Users/batch:      {USERS_PER_BATCH}")
    print(f"  Total users:      {NUM_BATCHES * USERS_PER_BATCH}")
    print(f"  Pipeline mode:    ALL {NUM_BATCHES} batches fired simultaneously")
    print(f"  Solution type:    accepted only (100% pass expected)")
    print(f"  Problems:         {[p['id'] for p, _ in pairs]}")
    print(f"  Judge0 URL:       {JUDGE0_URL}")
    print(f"  HTTP concurrency: 20 (RAILS_MAX_THREADS cap)")
    print("═" * 72 + "\n")

    sysinfo = gather_system_info()

    # ── Assign users to (batch, problem, solution) ────────────────────
    # Round-robin over accepted (problem, solution) pairs
    batch_slots: list[list[tuple[int, dict, dict]]] = []
    global_uid = 1
    for b in range(NUM_BATCHES):
        slots = []
        for u in range(USERS_PER_BATCH):
            pair_idx = (b * USERS_PER_BATCH + u) % len(pairs)
            prob, sol = pairs[pair_idx]
            slots.append((global_uid, prob, sol))
            global_uid += 1
        batch_slots.append(slots)

    # ── Pipeline: submit ALL batches to thread pool simultaneously ────
    print(f"  Firing all {NUM_BATCHES} batches in parallel...\n")
    overall_start = time.monotonic()

    batch_results: list[BatchRecord] = [None] * NUM_BATCHES  # type: ignore

    # Use one thread per batch — each batch itself spawns USERS_PER_BATCH threads
    # Total peak threads: NUM_BATCHES × USERS_PER_BATCH = 200
    # HTTP concurrency is still capped by HTTP_SEMAPHORE(20)
    with ThreadPoolExecutor(max_workers=NUM_BATCHES) as pool:
        future_to_bid = {
            pool.submit(run_batch, b + 1, batch_slots[b]): b
            for b in range(NUM_BATCHES)
        }
        for fut in as_completed(future_to_bid):
            b = future_to_bid[fut]
            try:
                batch_results[b] = fut.result()
            except Exception as exc:
                print(f"  [Batch {b+1:02d}] FATAL: {exc}", flush=True)
                batch_results[b] = BatchRecord(
                    batch_id       = b + 1,
                    batch_start_ts = time.time(),
                    batch_end_ts   = time.time(),
                )

    overall_duration = time.monotonic() - overall_start

    # ── Aggregate summary ─────────────────────────────────────────────
    all_users   = [u for br in batch_results for u in br.users]
    total_users = len(all_users)
    total_pass  = sum(1 for u in all_users if u.all_pass)
    all_latencies = [u.latency_ms for u in all_users if u.latency_ms > 0]
    all_latencies_sorted = sorted(all_latencies)

    def pct(lst, p):
        if not lst: return 0
        idx = int(len(lst) * p / 100)
        return lst[min(idx, len(lst) - 1)]

    tc_total = sum(u.total_tcs for u in all_users)
    tc_pass  = sum(u.score for u in all_users)

    summary = {
        "run_timestamp":         sysinfo["test_config"]["run_timestamp"],
        "overall_duration_s":    round(overall_duration, 3),
        "total_batches":         NUM_BATCHES,
        "users_per_batch":       USERS_PER_BATCH,
        "total_users":           total_users,
        "total_passed_users":    total_pass,
        "user_acceptance_rate":  f"{total_pass / total_users * 100:.1f}%" if total_users else "0%",
        "total_test_cases":      tc_total,
        "total_tc_passed":       tc_pass,
        "tc_pass_rate":          f"{tc_pass / tc_total * 100:.2f}%" if tc_total else "0%",
        "latency_ms": {
            "min":  min(all_latencies) if all_latencies else 0,
            "avg":  round(sum(all_latencies) / len(all_latencies)) if all_latencies else 0,
            "p50":  pct(all_latencies_sorted, 50),
            "p75":  pct(all_latencies_sorted, 75),
            "p90":  pct(all_latencies_sorted, 90),
            "p95":  pct(all_latencies_sorted, 95),
            "p99":  pct(all_latencies_sorted, 99),
            "max":  max(all_latencies) if all_latencies else 0,
        },
        "throughput_users_per_sec": round(total_users / overall_duration, 2) if overall_duration else 0,
        "per_problem": {},
        "system_info": sysinfo,
    }

    # Per-problem stats
    prob_map: dict[str, list[UserRecord]] = {}
    for u in all_users:
        prob_map.setdefault(u.problem_id, []).append(u)
    for pid, users in prob_map.items():
        lats = sorted([u.latency_ms for u in users if u.latency_ms > 0])
        summary["per_problem"][pid] = {
            "users":           len(users),
            "passed":          sum(1 for u in users if u.all_pass),
            "acceptance_rate": f"{sum(1 for u in users if u.all_pass) / len(users) * 100:.1f}%",
            "avg_latency_ms":  round(sum(lats) / len(lats)) if lats else 0,
            "p95_latency_ms":  pct(sorted(lats), 95),
        }

    print("\n" + "═" * 72)
    print("  SUMMARY")
    print("═" * 72)
    print(f"  Overall duration:     {overall_duration:.2f}s")
    print(f"  Total users:          {total_users}")
    print(f"  Passed (100% TCs):    {total_pass} / {total_users}  "
          f"({total_pass / total_users * 100:.1f}%)")
    print(f"  TC pass rate:         {tc_pass}/{tc_total}  "
          f"({tc_pass / tc_total * 100:.2f}%)")
    print(f"  Throughput:           {summary['throughput_users_per_sec']} users/sec")
    print(f"  Latency avg/p95/max:  "
          f"{summary['latency_ms']['avg']}ms / "
          f"{summary['latency_ms']['p95']}ms / "
          f"{summary['latency_ms']['max']}ms")
    print("═" * 72 + "\n")

    return batch_results, summary


# ══════════════════════════════════════════════════════════════════════════
# JSON report
# ══════════════════════════════════════════════════════════════════════════

def save_json_report(batch_results: list[BatchRecord], summary: dict, out_path: Path):
    report = {
        "summary": summary,
        "batches": [
            {
                "batch_id":         br.batch_id,
                "batch_start_ts":   br.batch_start_ts,
                "batch_end_ts":     br.batch_end_ts,
                "duration_s":       round(br.duration_s, 3),
                "total_users":      br.total_users,
                "passed_users":     br.passed_users,
                "acceptance_rate":  f"{br.acceptance_rate * 100:.1f}%",
                "avg_latency_ms":   round(br.avg_latency_ms),
                "max_latency_ms":   br.max_latency_ms,
                "min_latency_ms":   br.min_latency_ms,
                "errors":           br.errors,
                "users": [
                    {
                        "user_id":        u.user_id,
                        "user_in_batch":  u.user_in_batch,
                        "problem_id":     u.problem_id,
                        "solution_id":    u.solution_id,
                        "harness_status": u.harness_status,
                        "judge0_status":  u.judge0_status,
                        "score":          u.score,
                        "total_tcs":      u.total_tcs,
                        "latency_ms":     u.latency_ms,
                        "judge0_time_s":  u.judge0_time_s,
                        "judge0_mem_kb":  u.judge0_mem_kb,
                        "global_tle":     u.global_tle,
                        "error":          u.error,
                        "test_cases": [
                            {
                                "tc_id":       tc.tc_id,
                                "description": tc.description,
                                "status":      tc.status,
                                "got":         tc.got,
                                "expected":    tc.expected,
                                "detail":      tc.detail,
                            }
                            for tc in u.tc_results
                        ],
                    }
                    for u in sorted(br.users, key=lambda x: x.user_in_batch)
                ],
            }
            for br in batch_results
        ],
    }
    out_path.write_text(json.dumps(report, indent=2))
    return report


# ══════════════════════════════════════════════════════════════════════════
# Excel report
# ══════════════════════════════════════════════════════════════════════════

def _hdr_cell(ws, row, col, text,
              bg=CLR_HEADER_DARK, fg=CLR_WHITE,
              bold=True, size=11, wrap=False, align="center"):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=size)
    c.alignment = Alignment(
        horizontal=align, vertical="center",
        wrap_text=wrap
    )
    thin = Side(style="thin", color="AAAAAA")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

def _data_cell(ws, row, col, value, bold=False,
               bg=None, fg="000000", align="center",
               num_format=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    thin = Side(style="thin", color="DDDDDD")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if num_format:
        c.number_format = num_format
    return c

def _auto_width(ws, extra=4):
    for col in ws.columns:
        maxlen = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                maxlen = max(maxlen, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(maxlen + extra, 60)

def _status_colour(status: str) -> tuple[str, str]:
    """Return (bg_fill, font_colour) for a status string."""
    s = (status or "").upper()
    if s == "PASS":
        return CLR_PASS, CLR_PASS_FONT
    if s in ("FAIL", "WRONG_ANSWER", "COMPILE_ERROR", "JUDGE0_ERROR"):
        return CLR_FAIL, CLR_FAIL_FONT
    if s in ("TLE", "ERROR", "MISSING", "PARTIAL", "UNKNOWN"):
        return CLR_WARN, CLR_WARN_FONT
    return CLR_WHITE, "000000"


def build_excel(
    batch_results: list[BatchRecord],
    summary: dict,
    out_path: Path,
):
    wb = openpyxl.Workbook()

    # ── Sheet order ────────────────────────────────────────────────────
    ws_sum    = wb.active;  ws_sum.title    = "📊 Summary"
    ws_sys    = wb.create_sheet("⚙ System Config")
    ws_batch  = wb.create_sheet("🗂 Batch Report")
    ws_user   = wb.create_sheet("👤 User Report")
    ws_tc     = wb.create_sheet("🔬 TC Detail")
    ws_prob   = wb.create_sheet("📈 Problem Analytics")

    # ── 1. SUMMARY SHEET ──────────────────────────────────────────────
    ws_sum.sheet_view.showGridLines = False
    ws_sum.column_dimensions["A"].width = 38
    ws_sum.column_dimensions["B"].width = 28

    title_cell = ws_sum.cell(row=1, column=1,
        value="⚡ Judge0 Pipelined Load Test — 100% Acceptance")
    title_cell.font = Font(bold=True, size=16, color=CLR_HEADER_DARK)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.merge_cells("A1:B1")
    ws_sum.row_dimensions[1].height = 36

    ws_sum.cell(row=2, column=1, value=f"Run: {summary['run_timestamp']}")
    ws_sum.cell(row=2, column=1).font = Font(italic=True, color="888888", size=10)
    ws_sum.merge_cells("A2:B2")
    ws_sum.row_dimensions[2].height = 18

    kv_rows = [
        ("",                              ""),
        ("Test Configuration",            ""),
        ("Total Batches",                 summary["total_batches"]),
        ("Users per Batch",               summary["users_per_batch"]),
        ("Total Users",                   summary["total_users"]),
        ("Pipeline Mode",                 "All batches fired simultaneously"),
        ("Solution Type",                 "Accepted only (100% pass target)"),
        ("",                              ""),
        ("Results",                       ""),
        ("Overall Duration (s)",          summary["overall_duration_s"]),
        ("Passed Users",                  summary["total_passed_users"]),
        ("User Acceptance Rate",          summary["user_acceptance_rate"]),
        ("Total Test Cases",              summary["total_test_cases"]),
        ("TCs Passed",                    summary["total_tc_passed"]),
        ("TC Pass Rate",                  summary["tc_pass_rate"]),
        ("Throughput (users/sec)",        summary["throughput_users_per_sec"]),
        ("",                              ""),
        ("Latency",                       ""),
        ("Min (ms)",                      summary["latency_ms"]["min"]),
        ("Average (ms)",                  summary["latency_ms"]["avg"]),
        ("p50 (ms)",                      summary["latency_ms"]["p50"]),
        ("p75 (ms)",                      summary["latency_ms"]["p75"]),
        ("p90 (ms)",                      summary["latency_ms"]["p90"]),
        ("p95 (ms)",                      summary["latency_ms"]["p95"]),
        ("p99 (ms)",                      summary["latency_ms"]["p99"]),
        ("Max (ms)",                      summary["latency_ms"]["max"]),
    ]

    r = 3
    SECTION_KEYS = {"Test Configuration", "Results", "Latency"}
    for key, val in kv_rows:
        if key == "":
            r += 1
            continue
        if key in SECTION_KEYS:
            _hdr_cell(ws_sum, r, 1, key, bg=CLR_HEADER_MED, align="left", size=12)
            ws_sum.merge_cells(f"A{r}:B{r}")
            ws_sum.row_dimensions[r].height = 22
        else:
            _data_cell(ws_sum, r, 1, key, bold=True, align="left")
            is_pct = isinstance(val, str) and "%" in val
            is_pass = (
                key == "User Acceptance Rate" and val == "100.0%" or
                key == "TC Pass Rate"         and val.startswith("100")
            )
            bg_v = CLR_PASS if is_pass else None
            _data_cell(ws_sum, r, 2, val, bg=bg_v, align="right")
        r += 1

    # ── 2. SYSTEM CONFIG SHEET ────────────────────────────────────────
    ws_sys.sheet_view.showGridLines = False
    ws_sys.column_dimensions["A"].width = 36
    ws_sys.column_dimensions["B"].width = 50

    _hdr_cell(ws_sys, 1, 1, "Parameter", bg=CLR_HEADER_DARK)
    _hdr_cell(ws_sys, 1, 2, "Value",     bg=CLR_HEADER_DARK)
    ws_sys.row_dimensions[1].height = 22

    r = 2
    sysinfo = summary.get("system_info", {})

    # Docker containers
    _hdr_cell(ws_sys, r, 1, "Docker Containers", bg=CLR_HEADER_MED, align="left")
    ws_sys.merge_cells(f"A{r}:B{r}")
    r += 1
    for c in sysinfo.get("docker_containers", []):
        _data_cell(ws_sys, r, 1, c.get("name", ""), align="left")
        _data_cell(ws_sys, r, 2, f"{c.get('status','')}  |  {c.get('ports','')}",
                   align="left")
        r += 1

    r += 1
    _hdr_cell(ws_sys, r, 1, "docker-compose Highlights", bg=CLR_HEADER_MED, align="left")
    ws_sys.merge_cells(f"A{r}:B{r}")
    r += 1
    for k, v in sysinfo.get("docker_compose_highlights", {}).items():
        _data_cell(ws_sys, r, 1, k, align="left")
        _data_cell(ws_sys, r, 2, str(v), align="left")
        r += 1

    r += 1
    _hdr_cell(ws_sys, r, 1, "judge0.conf (non-secret)", bg=CLR_HEADER_MED, align="left")
    ws_sys.merge_cells(f"A{r}:B{r}")
    r += 1
    for k, v in sysinfo.get("judge0_conf", {}).items():
        _data_cell(ws_sys, r, 1, k, align="left")
        _data_cell(ws_sys, r, 2, str(v), align="left")
        r += 1

    r += 1
    _hdr_cell(ws_sys, r, 1, "Judge0 System Info (CPU/RAM)", bg=CLR_HEADER_MED, align="left")
    ws_sys.merge_cells(f"A{r}:B{r}")
    r += 1
    for k, v in sysinfo.get("judge0_system_info", {}).items():
        _data_cell(ws_sys, r, 1, k, align="left")
        _data_cell(ws_sys, r, 2, str(v), align="left")
        r += 1

    r += 1
    _hdr_cell(ws_sys, r, 1, "Test Config", bg=CLR_HEADER_MED, align="left")
    ws_sys.merge_cells(f"A{r}:B{r}")
    r += 1
    for k, v in sysinfo.get("test_config", {}).items():
        _data_cell(ws_sys, r, 1, k, align="left")
        _data_cell(ws_sys, r, 2, str(v), align="left")
        r += 1

    # ── 3. BATCH REPORT SHEET ─────────────────────────────────────────
    ws_batch.sheet_view.showGridLines = False
    batch_cols = [
        "Batch ID", "Start Time", "End Time", "Duration (s)",
        "Users", "Passed", "Failed", "Acceptance %",
        "Avg Latency (ms)", "Min Latency (ms)", "Max Latency (ms)", "Errors",
    ]
    for ci, hdr in enumerate(batch_cols, 1):
        _hdr_cell(ws_batch, 1, ci, hdr, bg=CLR_HEADER_DARK)
    ws_batch.row_dimensions[1].height = 22

    for ri, br in enumerate(batch_results, 2):
        alt_bg = CLR_ROW_ALT if ri % 2 == 0 else None
        failed = br.total_users - br.passed_users
        acceptance_pct = br.acceptance_rate * 100

        _data_cell(ws_batch, ri, 1,  br.batch_id,     bg=alt_bg, bold=True)
        _data_cell(ws_batch, ri, 2,
                   datetime.fromtimestamp(br.batch_start_ts).strftime("%H:%M:%S.%f")[:-3],
                   bg=alt_bg, align="left")
        _data_cell(ws_batch, ri, 3,
                   datetime.fromtimestamp(br.batch_end_ts).strftime("%H:%M:%S.%f")[:-3],
                   bg=alt_bg, align="left")
        _data_cell(ws_batch, ri, 4,  round(br.duration_s, 2),   bg=alt_bg)
        _data_cell(ws_batch, ri, 5,  br.total_users,             bg=alt_bg)

        passed = br.passed_users
        bg_pass = CLR_PASS if passed == br.total_users else alt_bg
        _data_cell(ws_batch, ri, 6,  passed,      bg=bg_pass,
                   fg=CLR_PASS_FONT if passed == br.total_users else "000000", bold=(passed == br.total_users))
        _data_cell(ws_batch, ri, 7,  failed,      bg=CLR_FAIL if failed > 0 else alt_bg,
                   fg=CLR_FAIL_FONT if failed > 0 else "000000")
        pct_bg = CLR_PASS if acceptance_pct == 100 else (CLR_FAIL if acceptance_pct < 50 else CLR_WARN)
        _data_cell(ws_batch, ri, 8,  f"{acceptance_pct:.1f}%",  bg=pct_bg,
                   bold=True)
        _data_cell(ws_batch, ri, 9,  round(br.avg_latency_ms),  bg=alt_bg)
        _data_cell(ws_batch, ri, 10, br.min_latency_ms,          bg=alt_bg)
        _data_cell(ws_batch, ri, 11, br.max_latency_ms,          bg=alt_bg)
        _data_cell(ws_batch, ri, 12, br.errors,
                   bg=CLR_FAIL if br.errors > 0 else alt_bg)
        ws_batch.row_dimensions[ri].height = 18

    # Totals row
    tr = len(batch_results) + 2
    all_users_flat = [u for br in batch_results for u in br.users]
    total_passed_all = sum(1 for u in all_users_flat if u.all_pass)
    total_failed_all = len(all_users_flat) - total_passed_all
    all_lats = [u.latency_ms for u in all_users_flat if u.latency_ms > 0]
    _hdr_cell(ws_batch, tr, 1,  "TOTAL",              bg=CLR_HEADER_MED)
    _hdr_cell(ws_batch, tr, 2,  "",                   bg=CLR_HEADER_MED)
    _hdr_cell(ws_batch, tr, 3,  "",                   bg=CLR_HEADER_MED)
    _hdr_cell(ws_batch, tr, 4,  "",                   bg=CLR_HEADER_MED)
    _hdr_cell(ws_batch, tr, 5,  len(all_users_flat),  bg=CLR_HEADER_MED)
    _hdr_cell(ws_batch, tr, 6,  total_passed_all,     bg=CLR_HEADER_MED)
    _hdr_cell(ws_batch, tr, 7,  total_failed_all,     bg=CLR_HEADER_MED)
    _hdr_cell(ws_batch, tr, 8,
              f"{total_passed_all / len(all_users_flat) * 100:.1f}%" if all_users_flat else "0%",
              bg=CLR_HEADER_MED)
    _hdr_cell(ws_batch, tr, 9,
              round(sum(all_lats) / len(all_lats)) if all_lats else 0,
              bg=CLR_HEADER_MED)
    _hdr_cell(ws_batch, tr, 10, min(all_lats) if all_lats else 0, bg=CLR_HEADER_MED)
    _hdr_cell(ws_batch, tr, 11, max(all_lats) if all_lats else 0, bg=CLR_HEADER_MED)
    _hdr_cell(ws_batch, tr, 12, sum(br.errors for br in batch_results), bg=CLR_HEADER_MED)
    ws_batch.row_dimensions[tr].height = 22

    _auto_width(ws_batch)

    # ── 4. USER REPORT SHEET ──────────────────────────────────────────
    ws_user.sheet_view.showGridLines = False
    user_cols = [
        "User ID", "Batch", "User #", "Problem", "Problem Title",
        "Solution ID", "Status", "Score", "Total TCs",
        "Pass Rate", "Latency (ms)", "Judge0 Time (s)",
        "Judge0 Mem (KB)", "Global TLE", "Error",
    ]
    for ci, hdr in enumerate(user_cols, 1):
        _hdr_cell(ws_user, 1, ci, hdr, bg=CLR_HEADER_DARK)
    ws_user.row_dimensions[1].height = 22

    ri = 2
    for br in batch_results:
        for u in sorted(br.users, key=lambda x: x.user_id):
            alt_bg = CLR_ROW_ALT if ri % 2 == 0 else None
            bg_status, fg_status = _status_colour(u.harness_status)
            pass_rate_pct = f"{u.pass_rate * 100:.0f}%"

            _data_cell(ws_user, ri, 1,  u.user_id,       bg=alt_bg, bold=True)
            _data_cell(ws_user, ri, 2,  u.batch_id,      bg=alt_bg)
            _data_cell(ws_user, ri, 3,  u.user_in_batch, bg=alt_bg)
            _data_cell(ws_user, ri, 4,  u.problem_id,    bg=alt_bg, align="left")
            _data_cell(ws_user, ri, 5,  u.problem_title, bg=alt_bg, align="left")
            _data_cell(ws_user, ri, 6,  u.solution_id,   bg=alt_bg, align="left")
            _data_cell(ws_user, ri, 7,  u.harness_status,
                       bg=bg_status, fg=fg_status, bold=True)
            _data_cell(ws_user, ri, 8,  u.score,         bg=alt_bg)
            _data_cell(ws_user, ri, 9,  u.total_tcs,     bg=alt_bg)
            pct_bg = CLR_PASS if u.pass_rate == 1.0 else (CLR_FAIL if u.pass_rate == 0 else CLR_WARN)
            _data_cell(ws_user, ri, 10, pass_rate_pct,   bg=pct_bg, bold=True)
            _data_cell(ws_user, ri, 11, u.latency_ms,    bg=alt_bg)
            _data_cell(ws_user, ri, 12, u.judge0_time_s, bg=alt_bg)
            _data_cell(ws_user, ri, 13, u.judge0_mem_kb, bg=alt_bg)
            _data_cell(ws_user, ri, 14, "YES" if u.global_tle else "NO",
                       bg=CLR_WARN if u.global_tle else alt_bg)
            _data_cell(ws_user, ri, 15, u.error or "",   bg=CLR_FAIL if u.error else alt_bg,
                       align="left", wrap=True)
            ws_user.row_dimensions[ri].height = 18
            ri += 1

    _auto_width(ws_user)
    ws_user.freeze_panes = "A2"

    # ── 5. TC DETAIL SHEET ────────────────────────────────────────────
    ws_tc.sheet_view.showGridLines = False
    tc_cols = [
        "User ID", "Batch", "Problem", "TC ID", "Description",
        "Status", "Expected", "Got", "Detail", "Stdin (preview)",
    ]
    for ci, hdr in enumerate(tc_cols, 1):
        _hdr_cell(ws_tc, 1, ci, hdr, bg=CLR_HEADER_DARK)
    ws_tc.row_dimensions[1].height = 22

    ri = 2
    for br in batch_results:
        for u in sorted(br.users, key=lambda x: x.user_id):
            for tc in u.tc_results:
                alt_bg = CLR_ROW_ALT if ri % 2 == 0 else None
                bg_tc, fg_tc = _status_colour(tc.status)
                stdin_preview = (tc.stdin[:40].replace("\n", "↵") + "…"
                                 if len(tc.stdin) > 40 else
                                 tc.stdin.replace("\n", "↵"))
                _data_cell(ws_tc, ri, 1,  u.user_id,       bg=alt_bg, bold=True)
                _data_cell(ws_tc, ri, 2,  u.batch_id,      bg=alt_bg)
                _data_cell(ws_tc, ri, 3,  u.problem_id,    bg=alt_bg, align="left")
                _data_cell(ws_tc, ri, 4,  tc.tc_id,        bg=alt_bg)
                _data_cell(ws_tc, ri, 5,  tc.description,  bg=alt_bg, align="left")
                _data_cell(ws_tc, ri, 6,  tc.status,       bg=bg_tc, fg=fg_tc, bold=True)
                _data_cell(ws_tc, ri, 7,  tc.expected,     bg=alt_bg, align="left")
                _data_cell(ws_tc, ri, 8,  tc.got,          bg=alt_bg, align="left")
                _data_cell(ws_tc, ri, 9,  tc.detail,       bg=alt_bg, align="left", wrap=True)
                _data_cell(ws_tc, ri, 10, stdin_preview,   bg=alt_bg, align="left")
                ws_tc.row_dimensions[ri].height = 18
                ri += 1

    _auto_width(ws_tc)
    ws_tc.freeze_panes = "A2"

    # ── 6. PROBLEM ANALYTICS SHEET ────────────────────────────────────
    ws_prob.sheet_view.showGridLines = False
    prob_cols = [
        "Problem ID", "Problem Title", "Total Users", "Passed",
        "Failed", "Acceptance %", "Avg Latency (ms)", "p95 Latency (ms)",
        "TC Pass Rate",
    ]
    for ci, hdr in enumerate(prob_cols, 1):
        _hdr_cell(ws_prob, 1, ci, hdr, bg=CLR_HEADER_DARK)
    ws_prob.row_dimensions[1].height = 22

    all_users_flat = [u for br in batch_results for u in br.users]
    prob_map: dict[str, list[UserRecord]] = {}
    for u in all_users_flat:
        prob_map.setdefault(u.problem_id, []).append(u)

    ri = 2
    for pid, users in sorted(prob_map.items()):
        lats = sorted([u.latency_ms for u in users if u.latency_ms > 0])
        passed_count = sum(1 for u in users if u.all_pass)
        failed_count = len(users) - passed_count
        tc_pass  = sum(u.score for u in users)
        tc_total = sum(u.total_tcs for u in users)
        acc_pct  = passed_count / len(users) * 100 if users else 0
        tc_pct   = tc_pass / tc_total * 100 if tc_total else 0
        avg_lat  = round(sum(lats) / len(lats)) if lats else 0
        p95_lat  = lats[int(len(lats) * 0.95)] if lats else 0
        title    = users[0].problem_title if users else ""

        alt_bg = CLR_ROW_ALT if ri % 2 == 0 else None
        _data_cell(ws_prob, ri, 1, pid,        bg=alt_bg, align="left", bold=True)
        _data_cell(ws_prob, ri, 2, title,      bg=alt_bg, align="left")
        _data_cell(ws_prob, ri, 3, len(users), bg=alt_bg)
        _data_cell(ws_prob, ri, 4, passed_count,
                   bg=CLR_PASS if passed_count == len(users) else alt_bg)
        _data_cell(ws_prob, ri, 5, failed_count,
                   bg=CLR_FAIL if failed_count > 0 else alt_bg)
        _data_cell(ws_prob, ri, 6, f"{acc_pct:.1f}%",
                   bg=CLR_PASS if acc_pct == 100 else (CLR_FAIL if acc_pct < 50 else CLR_WARN),
                   bold=True)
        _data_cell(ws_prob, ri, 7, avg_lat, bg=alt_bg)
        _data_cell(ws_prob, ri, 8, p95_lat, bg=alt_bg)
        _data_cell(ws_prob, ri, 9, f"{tc_pct:.1f}%",
                   bg=CLR_PASS if tc_pct == 100 else (CLR_FAIL if tc_pct < 50 else CLR_WARN),
                   bold=True)
        ws_prob.row_dimensions[ri].height = 18
        ri += 1

    _auto_width(ws_prob)

    # ── Add Bar Chart for Batch Acceptance % ──────────────────────────
    chart_row = len(batch_results) + 6
    ws_batch.cell(row=chart_row, column=1, value="Batch Acceptance % Chart")
    ws_batch.cell(row=chart_row, column=1).font = Font(bold=True, size=12)

    # Data for chart — batch IDs in col A, acceptance % in col 8
    chart = BarChart()
    chart.type  = "col"
    chart.style = 10
    chart.title = "Batch Acceptance Rate (%)"
    chart.y_axis.title = "Acceptance %"
    chart.x_axis.title = "Batch ID"
    chart.grouping = "clustered"
    chart.shape    = 4

    data_ref   = Reference(ws_batch, min_col=8, min_row=1,
                           max_row=len(batch_results) + 1)
    cats_ref   = Reference(ws_batch, min_col=1, min_row=2,
                           max_row=len(batch_results) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width  = 28
    chart.height = 16
    ws_batch.add_chart(chart, f"A{chart_row + 1}")

    # ── Latency Line Chart ─────────────────────────────────────────────
    lat_chart = LineChart()
    lat_chart.title    = "Avg Latency per Batch (ms)"
    lat_chart.y_axis.title = "Latency (ms)"
    lat_chart.x_axis.title = "Batch ID"
    lat_chart.style    = 10
    lat_ref  = Reference(ws_batch, min_col=9, min_row=1,
                         max_row=len(batch_results) + 1)
    lat_chart.add_data(lat_ref, titles_from_data=True)
    lat_chart.set_categories(cats_ref)
    lat_chart.width  = 28
    lat_chart.height = 16
    ws_batch.add_chart(lat_chart, f"O{chart_row + 1}")

    wb.save(out_path)
    print(f"\n  Excel report saved → {out_path}", flush=True)


# ══════════════════════════════════════════════════════════════════════════
# Entry point
# ══════════════════════════════════════════════════════════════════════════

def rebuild_excel_from_json(json_path: Path):
    """Rebuild BatchRecord / summary from a saved JSON report and re-emit Excel."""
    import shutil
    raw = json.loads(json_path.read_text())
    summary = raw["summary"]

    batch_results: list[BatchRecord] = []
    for b in raw["batches"]:
        br = BatchRecord(
            batch_id       = b["batch_id"],
            batch_start_ts = b["batch_start_ts"],
            batch_end_ts   = b["batch_end_ts"],
        )
        for u in b["users"]:
            tc_list = [
                TCRecord(
                    tc_id       = tc["tc_id"],
                    description = tc["description"],
                    stdin       = tc.get("stdin", ""),
                    expected    = tc["expected"],
                    status      = tc["status"],
                    got         = tc.get("got", ""),
                    detail      = tc.get("detail", ""),
                )
                for tc in (u.get("test_cases") or [])
            ]
            rec = UserRecord(
                batch_id       = u.get("batch_id", br.batch_id),
                user_id        = u["user_id"],
                user_in_batch  = u["user_in_batch"],
                problem_id     = u["problem_id"],
                problem_title  = u.get("problem_title", u["problem_id"]),
                solution_id    = u["solution_id"],
                solution_type  = u.get("solution_type", "accepted"),
                harness_status = u["harness_status"],
                judge0_status  = u.get("judge0_status", ""),
                score          = u["score"],
                total_tcs      = u["total_tcs"],
                latency_ms     = u["latency_ms"],
                judge0_time_s  = u.get("judge0_time_s", 0.0),
                judge0_mem_kb  = u.get("judge0_mem_kb", 0),
                global_tle     = u.get("global_tle", False),
                error          = u.get("error", ""),
                tc_results     = tc_list,
            )
            br.users.append(rec)
        batch_results.append(br)

    excel_path = json_path.with_suffix(".xlsx")
    build_excel(batch_results, summary, excel_path)
    latest_xlsx = REPORT_DIR / "latest.xlsx"
    shutil.copy2(excel_path, latest_xlsx)
    print(f"  Latest Excel: {latest_xlsx}", flush=True)


def main():
    import shutil
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    # --excel-only <json_path>  → skip load test, regenerate Excel from JSON
    if len(sys.argv) == 3 and sys.argv[1] == "--excel-only":
        rebuild_excel_from_json(Path(sys.argv[2]))
        return

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    batch_results, summary = orchestrate()

    json_path  = REPORT_DIR / f"load_test_100pct_{ts}.json"
    excel_path = REPORT_DIR / f"load_test_100pct_{ts}.xlsx"

    print("  Saving reports...", flush=True)
    save_json_report(batch_results, summary, json_path)
    print(f"  JSON report saved → {json_path}", flush=True)

    build_excel(batch_results, summary, excel_path)

    latest_json  = REPORT_DIR / "latest.json"
    latest_xlsx  = REPORT_DIR / "latest.xlsx"
    shutil.copy2(json_path, latest_json)
    shutil.copy2(excel_path, latest_xlsx)
    print(f"  Latest copies: {latest_json}, {latest_xlsx}", flush=True)

    print("\n  Done! 🏁\n", flush=True)


if __name__ == "__main__":
    main()
