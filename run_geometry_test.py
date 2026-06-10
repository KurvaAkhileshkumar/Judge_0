#!/usr/bin/env python3
"""
run_geometry_test.py
────────────────────
Load-tests the grader against geometry_question_bank.json.
Handles both stdio and function mode across all 4 languages.

Usage:
    python run_geometry_test.py
    python run_geometry_test.py --students 30 --url http://localhost:5001
    python run_geometry_test.py --bank geometry_question_bank.json --concurrency 8
"""

import argparse
import concurrent.futures
import datetime
import json
import sys
import time
import urllib.request
import urllib.error
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

API_BASE   = "http://localhost:5001"
BANK_FILE  = "geometry_question_bank.json"
SSE_TIMEOUT = 180

# ── ANSI colours ──────────────────────────────────────────────────────────────
GREEN  = "\033[92m"; RED    = "\033[91m"; YELLOW = "\033[93m"
CYAN   = "\033[96m"; BOLD   = "\033[1m";  RESET  = "\033[0m"

# ── HTTP helpers ──────────────────────────────────────────────────────────────

def post_json(url, data, timeout=30):
    body = json.dumps(data).encode()
    req  = urllib.request.Request(url, data=body,
                                  headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.status, json.loads(resp.read())


def stream_sse(url, timeout=SSE_TIMEOUT):
    req = urllib.request.Request(url, headers={"Accept": "text/event-stream"})
    deadline = time.monotonic() + timeout
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            buf = b""
            while time.monotonic() < deadline:
                chunk = resp.read(4096)
                if not chunk:
                    break
                buf += chunk
                lines = buf.split(b"\n")
                buf = lines[-1]
                for line in lines[:-1]:
                    decoded = line.decode(errors="replace").rstrip()
                    if decoded.startswith("data:"):
                        data = decoded[5:].strip()
                        if data:
                            return json.loads(data)
    except Exception as e:
        return {"stream_error": str(e)}
    return {"stream_error": "SSE closed without result"}


# ── Roster builder ────────────────────────────────────────────────────────────

def build_roster(problems, n_students):
    """
    Distribute n_students across problems round-robin.
    Each entry maps to one (problem, solution) pair.
    """
    entries = []
    for prob in problems:
        for sol in prob["solutions"]:
            entries.append({"problem": prob, "solution": sol})

    roster = []
    idx = 0
    for sid in range(1, n_students + 1):
        e = entries[idx % len(entries)]
        roster.append({
            "student_id":    f"student_{sid:03d}",
            "problem":       e["problem"],
            "solution_id":   e["solution"]["id"],
            "solution_type": e["solution"].get("type", "accepted"),
            "source_code":   e["solution"]["source_code"],
        })
        idx += 1
    return roster


# ── Submit one student ────────────────────────────────────────────────────────

def run_student(entry, api_base, run_id):
    prob   = entry["problem"]
    s_id   = entry["student_id"]
    code   = entry["source_code"]
    mode   = prob.get("mode", "stdio")

    # Build test cases — stdio uses stdin_text, function uses inputs
    api_tcs = []
    for tc in prob["test_cases"]:
        if mode == "function":
            api_tcs.append({
                "inputs":   tc["inputs"],
                "expected": tc["expected"],
            })
        else:
            api_tcs.append({
                "stdin_text": tc["stdin_text"],
                "expected":   tc["expected"],
            })

    body = {
        "student_id":      f"{s_id}_{run_id}",
        "assessment_id":   f"geo_{prob['id']}_{run_id}",
        "language":        prob["language"],
        "student_code":    code,
        "test_cases":      api_tcs,
        "mode":            mode,
        "per_tc_limit_s":  prob.get("per_tc_limit_s", 3),
        "memory_limit_mb": prob.get("memory_limit_mb", 256),
    }

    # Function-mode extras
    if mode == "function":
        body["function_name"] = prob.get("function_name", "solve")
        if prob.get("param_types"):
            body["param_types"] = prob["param_types"]
        if prob.get("return_type"):
            body["return_type"] = prob["return_type"]

    wall_start = time.monotonic()
    result_row = {
        "student_id":    s_id,
        "problem_id":    prob["id"],
        "problem_title": prob.get("title", prob["id"]),
        "language":      prob["language"],
        "mode":          mode,
        "solution_type": entry["solution_type"],
        "solution_id":   entry["solution_id"],
        "n_tcs":         len(prob["test_cases"]),
        "ticket_id":     "",
        "submit_status": "",
        "wall_time_s":   None,
        "judge0_time_s": None,
        "memory_kb":     None,
        "score":         0,
        "total":         len(prob["test_cases"]),
        "pass_rate":     0.0,
        "global_tle":    False,
        "system_error":  "",
        "security_error": "",
        "tc_results":    [],
        "error":         "",
    }

    try:
        http_status, resp = post_json(f"{api_base}/submit", body)
        ticket_id = resp.get("ticket_id", "")
        result_row["ticket_id"]     = ticket_id
        result_row["submit_status"] = resp.get("status", "")
    except Exception as e:
        result_row["error"] = f"submit_error: {e}"
        result_row["wall_time_s"] = round(time.monotonic() - wall_start, 3)
        return result_row

    result = stream_sse(f"{api_base}/results/stream/{result_row['ticket_id']}")
    result_row["wall_time_s"] = round(time.monotonic() - wall_start, 3)

    if "stream_error" in result:
        result_row["error"] = result["stream_error"]
        return result_row
    if "system_error" in result:
        result_row["system_error"] = result["system_error"]
        return result_row
    if "security_error" in result:
        result_row["security_error"] = result["security_error"]
        return result_row

    result_row["score"]         = result.get("score", 0)
    result_row["total"]         = result.get("total", len(prob["test_cases"]))
    result_row["global_tle"]    = result.get("global_tle", False)
    result_row["judge0_time_s"] = result.get("time_taken_s")
    result_row["memory_kb"]     = result.get("memory_kb")

    tc_results = result.get("tc_results", [])
    result_row["tc_results"] = tc_results
    if result_row["total"] > 0:
        result_row["pass_rate"] = round(
            100.0 * result_row["score"] / result_row["total"], 1
        )

    return result_row


# ── Console report ─────────────────────────────────────────────────────────

def print_console_report(all_results, duration_s):
    total   = len(all_results)
    perfect = sum(1 for r in all_results if r["pass_rate"] == 100.0
                  and not r["error"] and not r["system_error"] and not r["security_error"])
    errors  = sum(1 for r in all_results if r["error"] or r["system_error"])
    tles    = sum(1 for r in all_results if r["global_tle"])
    wrong   = total - perfect - errors - tles

    # Per-language breakdown
    by_lang = defaultdict(lambda: {"total": 0, "pass": 0})
    by_mode = defaultdict(lambda: {"total": 0, "pass": 0})
    by_prob = defaultdict(lambda: {"total": 0, "pass": 0})

    for r in all_results:
        lang = r["language"]; mode = r["mode"]
        # extract base question name
        pid  = r["problem_id"]
        parts = pid.rsplit("_", 2)
        q = "_".join(parts[:-2]) if len(parts) >= 3 else pid

        passed = (r["pass_rate"] == 100.0 and not r["error"] and not r["system_error"])
        by_lang[lang]["total"] += 1
        by_mode[mode]["total"] += 1
        by_prob[q]["total"]    += 1
        if passed:
            by_lang[lang]["pass"] += 1
            by_mode[mode]["pass"] += 1
            by_prob[q]["pass"]    += 1

    total_tcs  = sum(r["total"] for r in all_results)
    passed_tcs = sum(r["score"] for r in all_results)

    print(f"\n{BOLD}{'═'*68}{RESET}")
    print(f"{BOLD}  Geometry Load Test Results{RESET}")
    print(f"{'═'*68}")
    print(f"  Students:      {total}")
    print(f"  Duration:      {duration_s:.1f}s")
    print(f"  All-pass:      {GREEN}{perfect}/{total}{RESET}  ({100*perfect//total if total else 0}%)")
    print(f"  Wrong answer:  {YELLOW}{wrong}{RESET}")
    print(f"  Errors/TLE:    {RED}{errors}/{tles}{RESET}")
    print(f"  TCs passed:    {passed_tcs}/{total_tcs}")

    print(f"\n  {'Language':<10}  Pass")
    for lang in sorted(by_lang):
        d = by_lang[lang]
        pct = 100*d['pass']//d['total'] if d['total'] else 0
        print(f"  {lang:<10}  {d['pass']}/{d['total']}  ({pct}%)")

    print(f"\n  {'Mode':<12}  Pass")
    for mode in sorted(by_mode):
        d = by_mode[mode]
        pct = 100*d['pass']//d['total'] if d['total'] else 0
        print(f"  {mode:<12}  {d['pass']}/{d['total']}  ({pct}%)")

    print(f"\n  {'Question':<30}  Pass")
    for q in sorted(by_prob):
        d = by_prob[q]
        pct = 100*d['pass']//d['total'] if d['total'] else 0
        print(f"  {q:<30}  {d['pass']}/{d['total']}  ({pct}%)")

    # Failures
    failures = [r for r in all_results
                if r["pass_rate"] < 100.0 or r["error"] or r["system_error"] or r["security_error"]]
    if failures:
        print(f"\n{RED}  Failed submissions:{RESET}")
        for r in failures[:20]:
            reason = r["error"] or r["system_error"] or r["security_error"] or f"pass_rate={r['pass_rate']}%"
            print(f"    {r['student_id']} | {r['problem_id']} | {reason[:80]}")

    print(f"{'═'*68}\n")


# ── Excel report ──────────────────────────────────────────────────────────

def write_excel(all_results, run_meta, xlsx_path):
    if not HAS_OPENPYXL:
        print("  [skip] openpyxl not installed — no Excel report")
        return

    wb = openpyxl.Workbook()

    # colours
    C = {
        "hdr":  "1F4E79", "pass": "C6EFCE", "fail": "FFC7CE",
        "warn": "FFEB9C", "grey": "D9D9D9", "white": "FFFFFF",
    }
    def hdr_font(): return Font(bold=True, color="FFFFFF")
    def fill(hex_col): return PatternFill("solid", fgColor=hex_col)
    def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=False)

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Summary"
    meta_rows = [
        ("Run At",        run_meta["run_at"]),
        ("API URL",       run_meta["api_url"]),
        ("Students",      run_meta["n_students"]),
        ("Concurrency",   run_meta["concurrency"]),
        ("Bank File",     run_meta["bank_file"]),
        ("Total TCs",     run_meta["total_tcs"]),
        ("TCs Passed",    run_meta["tcs_passed"]),
        ("Duration (s)",  run_meta["duration_s"]),
    ]
    for r, (k, v) in enumerate(meta_rows, 1):
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40

    # per-language/mode summary
    start_row = len(meta_rows) + 2
    by_lang = defaultdict(lambda: {"total": 0, "pass": 0})
    by_mode = defaultdict(lambda: {"total": 0, "pass": 0})
    for r in all_results:
        ok = (r["pass_rate"] == 100.0 and not r["error"] and not r["system_error"])
        by_lang[r["language"]]["total"] += 1
        by_mode[r["mode"]]["total"] += 1
        if ok: by_lang[r["language"]]["pass"] += 1; by_mode[r["mode"]]["pass"] += 1

    hrow = start_row
    for col, h in enumerate(["Language", "Submitted", "All-Pass", "%"], 1):
        c = ws.cell(hrow, col, h); c.font = hdr_font(); c.fill = fill(C["hdr"])
    hrow += 1
    for lang in sorted(by_lang):
        d = by_lang[lang]
        pct = round(100*d["pass"]/d["total"], 1) if d["total"] else 0
        row_data = [lang, d["total"], d["pass"], f"{pct}%"]
        for col, val in enumerate(row_data, 1):
            ws.cell(hrow, col, val).fill = fill(C["pass"] if pct==100 else C["fail"])
        hrow += 1

    hrow += 1
    for col, h in enumerate(["Mode", "Submitted", "All-Pass", "%"], 1):
        c = ws.cell(hrow, col, h); c.font = hdr_font(); c.fill = fill(C["hdr"])
    hrow += 1
    for mode in sorted(by_mode):
        d = by_mode[mode]
        pct = round(100*d["pass"]/d["total"], 1) if d["total"] else 0
        for col, val in enumerate([mode, d["total"], d["pass"], f"{pct}%"], 1):
            ws.cell(hrow, col, val).fill = fill(C["pass"] if pct==100 else C["fail"])
        hrow += 1

    # ── Sheet 2: Submissions ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Submissions")
    headers = ["student_id","problem_id","language","mode","solution_type",
               "score","total","pass_rate","wall_time_s","judge0_time_s",
               "memory_kb","global_tle","system_error","security_error","error"]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(1, col, h); c.font = hdr_font(); c.fill = fill(C["hdr"])
        c.alignment = center()
    for row_i, r in enumerate(all_results, 2):
        ok = r["pass_rate"] == 100.0 and not r["error"] and not r["system_error"]
        row_fill = fill(C["pass"] if ok else C["fail"])
        for col, key in enumerate(headers, 1):
            c = ws2.cell(row_i, col, r.get(key, "")); c.fill = row_fill
    for col in range(1, len(headers)+1):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(f"  Excel report → {xlsx_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Geometry question bank load test")
    parser.add_argument("--url",         default=API_BASE)
    parser.add_argument("--students",    type=int, default=30)
    parser.add_argument("--concurrency", type=int, default=8)
    parser.add_argument("--bank",        default=BANK_FILE)
    parser.add_argument("--json-out",    default="geometry_test_results.json")
    parser.add_argument("--xlsx-out",    default="Reports/geometry_test_report.xlsx")
    args = parser.parse_args()

    bank_path = Path(args.bank)
    if not bank_path.exists():
        print(f"ERROR: {bank_path} not found."); sys.exit(1)

    with open(bank_path) as fh:
        bank = json.load(fh)
    problems = bank["problems"]

    roster = build_roster(problems, args.students)
    run_id = str(int(time.time()))

    print(f"\n{BOLD}{'═'*68}{RESET}")
    print(f"{BOLD}  Geometry Load Test{RESET}")
    print(f"{'═'*68}")
    print(f"  Students:      {args.students}")
    print(f"  Concurrency:   {args.concurrency}")
    print(f"  Bank:          {args.bank}  ({len(problems)} problem entries)")
    print(f"  API:           {args.url}")
    print(f"  Run ID:        {run_id}")

    lang_dist = Counter(e["problem"]["language"] for e in roster)
    mode_dist = Counter(e["problem"]["mode"] for e in roster)
    print(f"  Languages:     {dict(lang_dist)}")
    print(f"  Modes:         {dict(mode_dist)}")
    print(f"{'═'*68}\n")

    start_time = time.monotonic()
    all_results = []
    done_count  = [0]
    total       = len(roster)

    def submit_one(entry):
        r = run_student(entry, args.url, run_id)
        done_count[0] += 1
        ok  = r["pass_rate"] == 100.0 and not r["error"] and not r["system_error"]
        sym = f"{GREEN}✔{RESET}" if ok else f"{RED}✘{RESET}"
        print(f"  [{done_count[0]:3d}/{total}] {sym}  "
              f"{r['student_id']:<12} {r['problem_id']:<45} "
              f"{r['score']}/{r['total']}  {r['wall_time_s']:.1f}s")
        return r

    with concurrent.futures.ThreadPoolExecutor(max_workers=args.concurrency) as pool:
        futures = [pool.submit(submit_one, e) for e in roster]
        for fut in concurrent.futures.as_completed(futures):
            try:
                all_results.append(fut.result())
            except Exception as exc:
                print(f"  {RED}EXCEPTION: {exc}{RESET}")

    duration_s = round(time.monotonic() - start_time, 1)

    # Sort by student_id for deterministic output
    all_results.sort(key=lambda r: r["student_id"])

    print_console_report(all_results, duration_s)

    total_tcs  = sum(r["total"] for r in all_results)
    passed_tcs = sum(r["score"] for r in all_results)

    run_meta = {
        "run_at":      datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
        "run_id":      run_id,
        "api_url":     args.url,
        "bank_file":   args.bank,
        "n_students":  args.students,
        "concurrency": args.concurrency,
        "n_problems":  len(problems),
        "total_tcs":   total_tcs,
        "tcs_passed":  passed_tcs,
        "duration_s":  duration_s,
    }

    out = {
        "meta":     run_meta,
        "students": all_results,
    }
    with open(args.json_out, "w") as fh:
        json.dump(out, fh, indent=2)
    print(f"  JSON results → {args.json_out}")

    write_excel(all_results, run_meta, args.xlsx_out)


if __name__ == "__main__":
    main()
