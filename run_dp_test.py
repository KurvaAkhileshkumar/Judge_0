#!/usr/bin/env python3
"""
run_dp_test.py
──────────────
Runs 30 students through all 4 DP problems (fibonacci, coin_change, lcs, knapsack)
each with 30 test cases, collects full per-TC granular results, and generates
a comprehensive Excel report.

Student roster (30 students):
  • 24 correct solutions  (6 per problem)
  • 4  wrong answers      (1 per problem)
  • 2  TLE/edge solutions (1 each: naive recursion fib, incomplete DP)

Usage:
    python run_dp_test.py
    python run_dp_test.py --students 30 --url http://localhost:5001
    python run_dp_test.py --json-out dp_results.json --xlsx-out Reports/dp_report.xlsx
"""

import argparse
import concurrent.futures
import datetime
import json
import random
import sys
import time
import urllib.request
import urllib.error
from pathlib import Path

# ── Colour helpers (Excel) ────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

API_BASE    = "http://localhost:5001"
BANK_FILE   = "question_bank_mixed_30tc.json"
SSE_TIMEOUT = 180   # seconds

# ── Extra solutions (wrong / slow) one per problem ─────────────────────────
EXTRA_SOLUTIONS = {
    "fibonacci": [
        {
            "id": "fib_wrong",
            "type": "wrong_answer",
            "description": "Off-by-one: starts at 1 instead of 0",
            "source_code": "n = int(input())\na, b = 1, 1\nfor _ in range(n):\n    a, b = b, a + b\nprint(a)\n",
        },
        {
            "id": "fib_slow",
            "type": "slow_recursive",
            "description": "Naive exponential recursion (TLE for large n)",
            "source_code": (
                "import sys\nsys.setrecursionlimit(10000)\n"
                "def fib(n):\n    if n <= 1: return n\n    return fib(n-1) + fib(n-2)\n"
                "print(fib(int(input())))\n"
            ),
        },
    ],
    "coin_change": [
        {
            "id": "coin_wrong",
            "type": "wrong_answer",
            "description": "Greedy (incorrect for non-canonical coin systems)",
            "source_code": (
                "n = int(input())\ncoins = sorted(map(int, input().split()), reverse=True)\n"
                "amount = int(input())\ncount = 0\nfor c in coins:\n"
                "    count += amount // c\n    amount %= c\nprint(-1 if amount else count)\n"
            ),
        },
    ],
    "lcs": [
        {
            "id": "lcs_wrong",
            "type": "wrong_answer",
            "description": "Returns length of longer string instead of LCS",
            "source_code": "s1 = input()\ns2 = input()\nprint(max(len(s1), len(s2)))\n",
        },
    ],
    "knapsack": [
        {
            "id": "ks_wrong",
            "type": "wrong_answer",
            "description": "Sums all values regardless of capacity",
            "source_code": (
                "line1 = input().split()\nn, W = int(line1[0]), int(line1[1])\n"
                "weights = list(map(int, input().split()))\n"
                "values  = list(map(int, input().split()))\nprint(sum(values))\n"
            ),
        },
    ],
}


# ── HTTP helpers ──────────────────────────────────────────────────────────

def post_json(url, data, timeout=30):
    body = json.dumps(data).encode()
    req  = urllib.request.Request(url, data=body,
                                  headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.status, json.loads(resp.read())


def stream_sse(url, timeout=SSE_TIMEOUT):
    req = urllib.request.Request(url, headers={"Accept": "text/event-stream"})
    deadline = time.monotonic() + timeout
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            buf = b""
            while time.monotonic() < deadline:
                chunk = resp.read(4096)
                if not chunk:
                    break
                buf += chunk
                lines = buf.split(b"\n")
                buf = lines[-1]
                for line in lines[:-1]:
                    decoded = line.decode(errors="replace").rstrip()
                    if decoded.startswith("data:"):
                        data = decoded[5:].strip()
                        if data:
                            return json.loads(data)
    except Exception as e:
        return {"stream_error": str(e)}
    return {"stream_error": "SSE closed without result"}


# ── Assign students to problems ───────────────────────────────────────────

def build_student_roster(problems, n_students):
    """
    Distribute n_students across problems.
    Each problem gets correct solutions + 1 wrong solution (from EXTRA_SOLUTIONS).
    Remaining slots filled with correct solutions round-robin.
    """
    roster = []
    sid    = 1

    # First: one wrong solution per problem that has extras
    for prob in problems:
        extras = EXTRA_SOLUTIONS.get(prob["id"], [])
        for extra in extras:
            roster.append({
                "student_id":    f"student_{sid:02d}",
                "problem":       prob,
                "solution_id":   extra["id"],
                "solution_type": extra["type"],
                "source_code":   extra["source_code"],
            })
            sid += 1

    # Fill remaining with correct solutions round-robin
    correct_solutions = []
    for prob in problems:
        for sol in prob["solutions"]:
            correct_solutions.append((prob, sol))

    # repeat round-robin until we reach n_students
    idx = 0
    while sid <= n_students:
        prob, sol = correct_solutions[idx % len(correct_solutions)]
        roster.append({
            "student_id":    f"student_{sid:02d}",
            "problem":       prob,
            "solution_id":   sol["id"],
            "solution_type": sol.get("type", "accepted"),
            "source_code":   sol["source_code"],
        })
        sid += 1
        idx += 1

    return roster[:n_students]


# ── Run one student submission ────────────────────────────────────────────

def run_student(entry, api_base, run_id):
    prob   = entry["problem"]
    s_id   = entry["student_id"]
    code   = entry["source_code"]
    mode   = prob.get("mode", "stdio")

    # Build test cases for the API
    api_tcs = []
    for tc in prob["test_cases"]:
        api_tcs.append({
            "stdin_text": tc["stdin_text"],
            "expected":   tc["expected"],
        })

    body = {
        "student_id":      f"{s_id}_{run_id}",
        "assessment_id":   f"dp_test_{prob['id']}_{run_id}",
        "language":        prob["language"],
        "student_code":    code,
        "test_cases":      api_tcs,
        "mode":            mode,
        "per_tc_limit_s":  prob.get("per_tc_limit_s", 2),
        "memory_limit_mb": prob.get("memory_limit_mb", 256),
    }

    wall_start = time.monotonic()
    result_row = {
        "student_id":    s_id,
        "problem_id":    prob["id"],
        "problem_title": prob["title"],
        "difficulty":    prob.get("difficulty", ""),
        "language":      prob["language"],
        "solution_type": entry["solution_type"],
        "solution_id":   entry["solution_id"],
        "n_tcs":         len(prob["test_cases"]),
        "ticket_id":     "",
        "submit_status": "",
        "wall_time_s":   None,
        "judge0_time_s": None,
        "memory_kb":     None,
        "score":         0,
        "total":         len(prob["test_cases"]),
        "pass_rate":     0.0,
        "global_tle":    False,
        "system_error":  "",
        "security_error":"",
        "tc_results":    [],
        "tc_lookup":     {tc["id"]: tc for tc in prob["test_cases"]},
        "error":         "",
    }

    # Submit
    try:
        http_status, resp = post_json(f"{api_base}/submit", body)
        ticket_id = resp.get("ticket_id", "")
        result_row["ticket_id"]    = ticket_id
        result_row["submit_status"] = resp.get("status", "")
    except Exception as e:
        result_row["error"] = f"submit_error: {e}"
        result_row["wall_time_s"] = round(time.monotonic() - wall_start, 3)
        return result_row

    # Stream result
    result = stream_sse(f"{api_base}/results/stream/{ticket_id}")
    result_row["wall_time_s"] = round(time.monotonic() - wall_start, 3)

    if "stream_error" in result:
        result_row["error"] = result["stream_error"]
        return result_row

    if "system_error" in result:
        result_row["system_error"] = result["system_error"]
        return result_row

    if "security_error" in result:
        result_row["security_error"] = result["security_error"]
        return result_row

    result_row["score"]        = result.get("score", 0)
    result_row["total"]        = result.get("total", len(prob["test_cases"]))
    result_row["global_tle"]   = result.get("global_tle", False)
    result_row["judge0_time_s"]= result.get("time_taken_s")
    result_row["memory_kb"]    = result.get("memory_kb")
    result_row["pass_rate"]    = round(
        result_row["score"] / max(result_row["total"], 1) * 100, 1)

    # Enrich TC results with problem TC metadata
    tc_meta = {tc["id"]: tc for tc in prob["test_cases"]}
    tc_by_num = {tc["id"]: i+1 for i, tc in enumerate(prob["test_cases"])}

    raw_tcs = result.get("tc_results", [])
    for tc_r in raw_tcs:
        tc_num   = tc_r.get("tc_num", 0)
        # look up metadata by position (1-indexed)
        meta_tc = prob["test_cases"][tc_num - 1] if 0 < tc_num <= len(prob["test_cases"]) else {}
        result_row["tc_results"].append({
            "tc_num":      tc_num,
            "tc_id":       meta_tc.get("id", f"tc{tc_num}"),
            "description": meta_tc.get("description", ""),
            "stdin_text":  meta_tc.get("stdin_text", ""),
            "expected":    meta_tc.get("expected", tc_r.get("expected", "")),
            "got":         tc_r.get("got", ""),
            "status":      tc_r.get("status", ""),
            "detail":      tc_r.get("detail", ""),
            "warning":     tc_r.get("warning", ""),
        })

    return result_row


# ── Progress printer ──────────────────────────────────────────────────────

def _progress(done, total, start):
    elapsed = time.monotonic() - start
    pct     = done / max(total, 1) * 100
    print(f"\r  [{done:2d}/{total}] {pct:5.1f}%  elapsed {elapsed:.1f}s  ",
          end="", flush=True)


# ── Excel report ──────────────────────────────────────────────────────────

# Palette
C = {
    "PASS":     "FF27AE60",
    "FAIL":     "FFE74C3C",
    "TLE":      "FFF39C12",
    "MLE":      "FF9B59B6",
    "ERROR":    "FF95A5A6",
    "SEGV":     "FFEC407A",
    "MISSING":  "FFBDBDBD",
    "hdr":      "FF2C3E50",
    "hdr_fg":   "FFFFFFFF",
    "alt":      "FFF5F6FA",
    "section":  "FF2980B9",
    "python":   "FF3498DB",
    "accepted": "FF27AE60",
    "wrong_answer":   "FFE74C3C",
    "slow_recursive": "FFF39C12",
    "correct":  "FF27AE60",
    "fibonacci":  "FF3498DB",
    "coin_change":"FFE67E22",
    "lcs":        "FF9B59B6",
    "knapsack":   "FF1ABC9C",
}

STATUS_COLOR = {
    "PASS": C["PASS"], "FAIL": C["FAIL"], "TLE": C["TLE"],
    "MLE": C["MLE"], "ERROR": C["ERROR"], "SEGV": C["SEGV"],
    "MISSING": C["MISSING"], "OUTPUT": C["PASS"],
}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="FFBDC3C7")
    return Border(left=s, right=s, top=s, bottom=s)

def _set(ws, row, col, value, bold=False, bg=None, fg="FF000000",
         align="left", size=10, wrap=False, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, color=fg, size=size)
    cell.alignment = Alignment(
        horizontal="center" if align == "center" else "left",
        vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = _fill(bg)
    if border:
        cell.border = _border()
    return cell

def _hdr(ws, row, labels, col_start=1):
    for i, lbl in enumerate(labels):
        _set(ws, row, col_start + i, lbl, bold=True,
             bg=C["hdr"], fg=C["hdr_fg"], align="center")

def _widths(ws, d):
    for col, w in d.items():
        ws.column_dimensions[col].width = w


def build_summary_sheet(wb, all_results, run_meta):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    r = 1
    cell = ws.cell(row=r, column=1,
        value="DP Load Test — Granular Report")
    cell.font      = Font(bold=True, size=16, color=C["hdr"][2:])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1

    ws.cell(row=r, column=1,
        value=f"Run at: {run_meta['run_at']}   |   Students: {run_meta['n_students']}   |   "
              f"Problems: {', '.join(run_meta['problems'])}   |   "
              f"Total TCs: {run_meta['total_tcs_evaluated']}   |   "
              f"Duration: {run_meta['total_duration_s']}s"
    ).font = Font(size=10, italic=True, color="FF7F8C8D")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 2

    # KPIs
    ok_sub  = sum(1 for s in all_results if not s["error"] and not s["system_error"])
    n_pass  = sum(1 for s in all_results if s["score"] == s["total"] and s["total"] > 0)
    n_part  = sum(1 for s in all_results if 0 < s["score"] < s["total"])
    n_zero  = sum(1 for s in all_results if s["score"] == 0 and not s["error"] and not s["system_error"])
    total_tc_pass = sum(s["score"] for s in all_results)
    total_tc_all  = sum(s["total"] for s in all_results)
    avg_wall = (sum(s["wall_time_s"] for s in all_results if s["wall_time_s"])
                / max(sum(1 for s in all_results if s["wall_time_s"]), 1))

    kpis = [
        ("Students",        run_meta['n_students'], C["hdr"]),
        ("TCs Evaluated",   total_tc_all,           C["section"]),
        ("TCs Passed",      total_tc_pass,          C["PASS"]),
        ("TC Pass Rate",    f"{total_tc_pass/max(total_tc_all,1)*100:.1f}%", C["PASS"]),
        ("Full Score",      n_pass,                 C["PASS"]),
        ("Partial",         n_part,                 C["TLE"]),
        ("Zero Score",      n_zero,                 C["FAIL"]),
        ("Avg Latency",     f"{avg_wall:.2f}s",     C["hdr"]),
    ]
    for i, (label, val, bg) in enumerate(kpis):
        col = 1 + i
        lc  = ws.cell(row=r, column=col, value=label)
        lc.font = Font(bold=True, size=9, color="FFFFFFFF")
        lc.fill = _fill(bg); lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = _border()
        vc  = ws.cell(row=r+1, column=col, value=val)
        vc.font = Font(bold=True, size=18, color=bg[2:])
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = _border()
    r += 3

    # Per-problem table
    r += 1
    ws.cell(row=r, column=1, value="Per-Problem Breakdown"
            ).font = Font(bold=True, size=12, color=C["section"][2:])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1

    _hdr(ws, r, ["Problem", "Difficulty", "Lang", "Students",
                 "Correct (30/30)", "Partial", "Wrong (0)", "TC Pass Rate", "Avg Wall (s)"])
    r += 1

    prob_stats = {}
    for s in all_results:
        pid = s["problem_id"]
        prob_stats.setdefault(pid, {
            "title": s["problem_title"], "diff": s["difficulty"],
            "lang": s["language"], "n": 0,
            "full": 0, "part": 0, "zero": 0,
            "tc_pass": 0, "tc_total": 0, "wall_sum": 0, "wall_n": 0,
        })
        ps = prob_stats[pid]
        ps["n"] += 1
        ps["tc_pass"]  += s["score"]
        ps["tc_total"] += s["total"]
        if s["score"] == s["total"] and s["total"] > 0:
            ps["full"] += 1
        elif s["score"] > 0:
            ps["part"] += 1
        else:
            ps["zero"] += 1
        if s["wall_time_s"]:
            ps["wall_sum"] += s["wall_time_s"]
            ps["wall_n"]   += 1

    prob_colors = {"fibonacci": C["fibonacci"], "coin_change": C["coin_change"],
                   "lcs": C["lcs"], "knapsack": C["knapsack"]}

    for idx, (pid, ps) in enumerate(prob_stats.items()):
        bg   = C["alt"] if idx % 2 else "FFFFFFFF"
        pct  = ps["tc_pass"] / max(ps["tc_total"], 1) * 100
        wall = ps["wall_sum"] / max(ps["wall_n"], 1)
        pc   = prob_colors.get(pid, C["hdr"])
        _set(ws, r, 1, ps["title"], bold=True, bg=pc, fg="FFFFFFFF")
        _set(ws, r, 2, ps["diff"],  bg=bg, align="center")
        _set(ws, r, 3, ps["lang"].upper(), bg=C["python"], fg="FFFFFFFF", align="center")
        _set(ws, r, 4, ps["n"],     bg=bg, align="center")
        _set(ws, r, 5, ps["full"],  bg=C["PASS"] if ps["full"] == ps["n"] else bg, fg="FFFFFFFF" if ps["full"] == ps["n"] else "FF000000", align="center")
        _set(ws, r, 6, ps["part"],  bg=C["TLE"] if ps["part"] > 0 else bg, fg="FFFFFFFF" if ps["part"] > 0 else "FF000000", align="center")
        _set(ws, r, 7, ps["zero"],  bg=C["FAIL"] if ps["zero"] > 0 else bg, fg="FFFFFFFF" if ps["zero"] > 0 else "FF000000", align="center")
        _set(ws, r, 8, f"{pct:.1f}%", bg=C["PASS"] if pct == 100 else (C["TLE"] if pct > 50 else C["FAIL"]), fg="FFFFFFFF", align="center", bold=True)
        _set(ws, r, 9, f"{wall:.2f}s", bg=bg, align="center")
        r += 1

    _widths(ws, {"A": 32, "B": 12, "C": 10, "D": 10, "E": 16, "F": 10, "G": 12, "H": 14, "I": 14})
    ws.freeze_panes = "A3"


def build_submissions_sheet(wb, all_results):
    ws = wb.create_sheet("Student Submissions")
    ws.sheet_view.showGridLines = False

    r = 1
    ws.cell(row=r, column=1, value="Per-Student Submission Detail"
            ).font = Font(bold=True, size=13, color=C["hdr"][2:])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    r += 2

    _hdr(ws, r, [
        "Student", "Problem", "Difficulty", "Solution Type",
        "Score", "Total TCs", "Pass %", "Result",
        "Wall Time (s)", "Judge0 CPU (s)", "Memory (KB)", "Ticket ID"
    ])
    r += 1

    prob_colors = {"fibonacci": C["fibonacci"], "coin_change": C["coin_change"],
                   "lcs": C["lcs"], "knapsack": C["knapsack"]}
    sol_colors  = {"accepted": C["PASS"], "wrong_answer": C["FAIL"],
                   "slow_recursive": C["TLE"]}

    for idx, s in enumerate(sorted(all_results, key=lambda x: x["student_id"])):
        bg       = C["alt"] if idx % 2 else "FFFFFFFF"
        full     = s["score"] == s["total"] and s["total"] > 0
        pct      = s["score"] / max(s["total"], 1) * 100
        res_text = ("ALL PASS" if full
                    else f"PARTIAL {s['score']}/{s['total']}" if s["score"] > 0
                    else "SYSTEM ERR" if s["system_error"]
                    else "SEC BLOCK" if s["security_error"]
                    else "ERROR"     if s["error"]
                    else "ALL FAIL")
        res_bg   = (C["PASS"] if full
                    else C["TLE"] if 0 < s["score"] < s["total"]
                    else C["FAIL"])

        pc = prob_colors.get(s["problem_id"], C["hdr"])
        sc = sol_colors.get(s["solution_type"], bg)

        _set(ws, r,  1, s["student_id"],     bg=bg, bold=True)
        _set(ws, r,  2, s["problem_title"],  bg=pc, fg="FFFFFFFF")
        _set(ws, r,  3, s.get("difficulty",""), bg=bg, align="center")
        _set(ws, r,  4, s["solution_type"],  bg=sc, fg="FFFFFFFF" if sc != bg else "FF000000", align="center")
        _set(ws, r,  5, s["score"],          bg=bg, align="center", bold=True)
        _set(ws, r,  6, s["total"],          bg=bg, align="center")
        _set(ws, r,  7, f"{pct:.0f}%",       bg=C["PASS"] if full else (C["TLE"] if pct > 0 else C["FAIL"]),
             fg="FFFFFFFF", align="center", bold=True)
        _set(ws, r,  8, res_text,            bg=res_bg, fg="FFFFFFFF", align="center", bold=True)
        _set(ws, r,  9, s["wall_time_s"]  or "-", bg=bg, align="center")
        _set(ws, r, 10, s["judge0_time_s"] or "-", bg=bg, align="center")
        _set(ws, r, 11, s["memory_kb"]    or "-", bg=bg, align="center")
        _set(ws, r, 12, (s["ticket_id"] or "")[:8] + "…", bg=bg, align="center")
        r += 1

    _widths(ws, {"A": 14, "B": 32, "C": 12, "D": 18, "E": 8, "F": 10,
                 "G": 8, "H": 14, "I": 14, "J": 16, "K": 14, "L": 12})
    ws.freeze_panes = "A4"


def build_tc_detail_sheet(wb, all_results):
    ws = wb.create_sheet("TC Results (All)")
    ws.sheet_view.showGridLines = False

    r = 1
    ws.cell(row=r, column=1, value="Per-Test-Case Results — All Students"
            ).font = Font(bold=True, size=13, color=C["hdr"][2:])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    r += 2

    _hdr(ws, r, [
        "Student", "Problem", "Sol Type",
        "TC #", "TC ID", "Description",
        "Stdin (trunc)", "Expected", "Got",
        "Status", "Detail"
    ])
    r += 1

    prob_colors = {"fibonacci": C["fibonacci"], "coin_change": C["coin_change"],
                   "lcs": C["lcs"], "knapsack": C["knapsack"]}

    for s in sorted(all_results, key=lambda x: x["student_id"]):
        pc = prob_colors.get(s["problem_id"], C["hdr"])
        for tc in s["tc_results"]:
            status = tc["status"]
            sbg    = STATUS_COLOR.get(status, C["ERROR"])
            bg     = "FFFFFFFF" if tc["tc_num"] % 2 else C["alt"]

            stdin_trunc = tc["stdin_text"].replace("\n", "↵")[:30]

            _set(ws, r,  1, s["student_id"],   bg=bg, bold=False)
            _set(ws, r,  2, s["problem_id"],   bg=pc, fg="FFFFFFFF", align="center")
            _set(ws, r,  3, s["solution_type"][:8], bg=bg, align="center")
            _set(ws, r,  4, tc["tc_num"],      bg=bg, align="center", bold=True)
            _set(ws, r,  5, tc["tc_id"],       bg=bg, align="center")
            _set(ws, r,  6, tc["description"], bg=bg)
            _set(ws, r,  7, stdin_trunc,       bg=bg, wrap=True)
            _set(ws, r,  8, tc["expected"],    bg=bg, align="center")
            _set(ws, r,  9, str(tc["got"])[:40], bg=bg, align="center")
            _set(ws, r, 10, status,            bg=sbg, fg="FFFFFFFF", align="center", bold=True)
            _set(ws, r, 11, tc["detail"][:80], bg=bg, wrap=True)
            r += 1

    _widths(ws, {"A": 12, "B": 14, "C": 14, "D": 6, "E": 8, "F": 22,
                 "G": 24, "H": 12, "I": 14, "J": 10, "K": 36})
    ws.freeze_panes = "A4"
    ws.row_dimensions[3].height = 20


def build_tc_heatmap_sheet(wb, all_results, problems):
    """
    Per-problem: rows = students, cols = TC 1..30, cell = PASS/FAIL/TLE/etc.
    Gives a visual heatmap of which TCs each student passed.
    """
    ws = wb.create_sheet("TC Heatmap")
    ws.sheet_view.showGridLines = False

    r = 1
    ws.cell(row=r, column=1, value="TC Pass/Fail Heatmap — per Student per TC"
            ).font = Font(bold=True, size=13, color=C["hdr"][2:])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=35)
    r += 2

    prob_colors = {"fibonacci": C["fibonacci"], "coin_change": C["coin_change"],
                   "lcs": C["lcs"], "knapsack": C["knapsack"]}

    # Group students by problem
    by_prob = {}
    for s in all_results:
        by_prob.setdefault(s["problem_id"], []).append(s)

    for prob_id, students in sorted(by_prob.items()):
        # Section header
        pc = prob_colors.get(prob_id, C["hdr"])
        title_cell = ws.cell(row=r, column=1,
            value=f"{prob_id.upper().replace('_',' ')} — {len(students)} students")
        title_cell.font = Font(bold=True, size=11, color="FFFFFFFF")
        title_cell.fill = _fill(pc)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        title_cell.border = _border()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=33)
        r += 1

        # Column headers: Student | Sol Type | Score | TC1..TC30
        hdr_row = ["Student", "Sol Type", "Score"] + [str(i) for i in range(1, 31)]
        _hdr(ws, r, hdr_row)
        r += 1

        for s in sorted(students, key=lambda x: x["student_id"]):
            tc_map = {tc["tc_num"]: tc["status"] for tc in s["tc_results"]}
            score_str = f"{s['score']}/{s['total']}"
            sol_bg = {"accepted": C["PASS"], "wrong_answer": C["FAIL"],
                      "slow_recursive": C["TLE"]}.get(s["solution_type"], "FFFFFFFF")

            _set(ws, r, 1, s["student_id"], bold=True)
            _set(ws, r, 2, s["solution_type"][:8], bg=sol_bg, fg="FFFFFFFF", align="center")
            full = s["score"] == s["total"] and s["total"] > 0
            _set(ws, r, 3, score_str,
                 bg=C["PASS"] if full else (C["TLE"] if s["score"] > 0 else C["FAIL"]),
                 fg="FFFFFFFF", align="center", bold=True)

            for tc_num in range(1, 31):
                status = tc_map.get(tc_num, "—")
                sbg    = STATUS_COLOR.get(status, "FFFFFFFF")
                icon   = "✓" if status == "PASS" else ("✗" if status == "FAIL"
                         else "T" if status == "TLE"
                         else ("M" if status == "MLE"
                         else ("S" if status == "SEGV" else "?")))
                cell = ws.cell(row=r, column=3 + tc_num, value=icon)
                cell.font      = Font(bold=True, size=8,
                                      color="FFFFFFFF" if sbg != "FFFFFFFF" else "FF000000")
                cell.fill      = _fill(sbg) if sbg != "FFFFFFFF" else PatternFill()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = _border()

            r += 1

        r += 1  # blank row between problems

    # Set column widths
    width_map = {"A": 14, "B": 10, "C": 8}
    for i in range(1, 31):
        width_map[get_column_letter(3 + i)] = 3.5
    _widths(ws, width_map)

    for row_idx in range(1, r + 1):
        ws.row_dimensions[row_idx].height = 16

    ws.freeze_panes = "D3"


def build_tc_passrate_sheet(wb, all_results, problems):
    """Per-TC pass rate aggregated across all students for that problem."""
    ws = wb.create_sheet("TC Pass Rate Analysis")
    ws.sheet_view.showGridLines = False

    r = 1
    ws.cell(row=r, column=1, value="TC Pass Rate — Which test cases trip students up?"
            ).font = Font(bold=True, size=13, color=C["hdr"][2:])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 2

    prob_colors = {"fibonacci": C["fibonacci"], "coin_change": C["coin_change"],
                   "lcs": C["lcs"], "knapsack": C["knapsack"]}

    # Build problem metadata lookup
    prob_meta = {p["id"]: p for p in problems}

    by_prob = {}
    for s in all_results:
        by_prob.setdefault(s["problem_id"], []).append(s)

    for prob_id, students in sorted(by_prob.items()):
        pc    = prob_colors.get(prob_id, C["hdr"])
        pmeta = prob_meta.get(prob_id, {})

        title_cell = ws.cell(row=r, column=1,
            value=f"{pmeta.get('title', prob_id)}  ({len(students)} students)")
        title_cell.font = Font(bold=True, size=11, color="FFFFFFFF")
        title_cell.fill = _fill(pc)
        title_cell.alignment = Alignment(horizontal="left")
        title_cell.border = _border()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        r += 1

        _hdr(ws, r, ["TC #", "TC ID", "Description", "Stdin",
                     "Expected", "Pass", "Fail", "Other", "Pass Rate %"])
        r += 1

        for tc_idx, meta_tc in enumerate(pmeta.get("test_cases", [])):
            tc_num = tc_idx + 1
            pass_n = fail_n = other_n = 0
            for s in students:
                tc_map = {tc["tc_num"]: tc["status"] for tc in s["tc_results"]}
                st = tc_map.get(tc_num, "MISSING")
                if st == "PASS":
                    pass_n += 1
                elif st == "FAIL":
                    fail_n += 1
                else:
                    other_n += 1

            total_n  = len(students)
            pct      = pass_n / max(total_n, 1) * 100
            bg       = C["alt"] if tc_idx % 2 else "FFFFFFFF"
            pct_bg   = (C["PASS"] if pct == 100 else
                        C["TLE"]  if pct >= 50 else C["FAIL"])
            stdin_tr = meta_tc.get("stdin_text", "").replace("\n", "↵")[:30]

            _set(ws, r, 1, tc_num,                    bg=bg, align="center", bold=True)
            _set(ws, r, 2, meta_tc.get("id", ""),     bg=bg, align="center")
            _set(ws, r, 3, meta_tc.get("description",""), bg=bg)
            _set(ws, r, 4, stdin_tr,                  bg=bg)
            _set(ws, r, 5, meta_tc.get("expected",""),bg=bg, align="center")
            _set(ws, r, 6, pass_n,  bg=C["PASS"] if pass_n > 0 else bg, fg="FFFFFFFF" if pass_n > 0 else "FF000000", align="center")
            _set(ws, r, 7, fail_n,  bg=C["FAIL"] if fail_n > 0 else bg, fg="FFFFFFFF" if fail_n > 0 else "FF000000", align="center")
            _set(ws, r, 8, other_n, bg=C["TLE"]  if other_n > 0 else bg, fg="FFFFFFFF" if other_n > 0 else "FF000000", align="center")
            _set(ws, r, 9, f"{pct:.0f}%",  bg=pct_bg, fg="FFFFFFFF", align="center", bold=True)
            r += 1

        r += 1

    _widths(ws, {"A": 6, "B": 10, "C": 26, "D": 30, "E": 10,
                 "F": 8, "G": 8, "H": 8, "I": 12})
    ws.freeze_panes = "A3"


def build_latency_sheet(wb, all_results):
    ws = wb.create_sheet("Timing Analysis")
    ws.sheet_view.showGridLines = False

    r = 1
    ws.cell(row=r, column=1, value="Timing Analysis — Wall Time & Judge0 CPU"
            ).font = Font(bold=True, size=13, color=C["hdr"][2:])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 2

    # Per-problem timing stats
    _hdr(ws, r, ["Problem", "Students", "Min (s)", "Max (s)", "Avg (s)",
                 "P50 (s)", "P90 (s)", "P99 (s)"])
    r += 1

    by_prob = {}
    for s in all_results:
        if s["wall_time_s"]:
            by_prob.setdefault(s["problem_id"], []).append(s["wall_time_s"])

    prob_colors = {"fibonacci": C["fibonacci"], "coin_change": C["coin_change"],
                   "lcs": C["lcs"], "knapsack": C["knapsack"]}

    for idx, (pid, times) in enumerate(sorted(by_prob.items())):
        times.sort()
        bg = C["alt"] if idx % 2 else "FFFFFFFF"
        pc = prob_colors.get(pid, C["hdr"])
        n  = len(times)
        _set(ws, r, 1, pid, bold=True, bg=pc, fg="FFFFFFFF")
        _set(ws, r, 2, n, bg=bg, align="center")
        _set(ws, r, 3, f"{min(times):.2f}", bg=bg, align="center")
        _set(ws, r, 4, f"{max(times):.2f}", bg=bg, align="center")
        _set(ws, r, 5, f"{sum(times)/n:.2f}", bg=bg, align="center", bold=True)
        _set(ws, r, 6, f"{times[int(n*.50)]:.2f}", bg=bg, align="center")
        _set(ws, r, 7, f"{times[int(n*.90)]:.2f}", bg=bg, align="center")
        _set(ws, r, 8, f"{times[min(int(n*.99),n-1)]:.2f}", bg=bg, align="center")
        r += 1

    r += 2

    # All submissions timing table
    _hdr(ws, r, ["Student", "Problem", "Sol Type", "Score",
                 "Wall Time (s)", "Judge0 CPU (s)", "Memory (KB)", "Notes"])
    r += 1

    for s in sorted(all_results, key=lambda x: x["wall_time_s"] or 0, reverse=True):
        bg  = C["alt"] if r % 2 else "FFFFFFFF"
        pc  = prob_colors.get(s["problem_id"], C["hdr"])
        note = ""
        if s["system_error"]:  note = "system_error"
        elif s["global_tle"]:  note = "global_TLE"
        elif s["error"]:       note = s["error"][:40]

        _set(ws, r, 1, s["student_id"],  bg=bg)
        _set(ws, r, 2, s["problem_id"],  bg=pc, fg="FFFFFFFF", align="center")
        _set(ws, r, 3, s["solution_type"][:10], bg=bg, align="center")
        _set(ws, r, 4, f"{s['score']}/{s['total']}", bg=bg, align="center")
        _set(ws, r, 5, s["wall_time_s"]   or "—", bg=bg, align="center", bold=True)
        _set(ws, r, 6, s["judge0_time_s"] or "—", bg=bg, align="center")
        _set(ws, r, 7, s["memory_kb"]     or "—", bg=bg, align="center")
        _set(ws, r, 8, note, bg=C["FAIL"] if note else bg,
             fg="FFFFFFFF" if note else "FF000000")
        r += 1

    _widths(ws, {"A": 14, "B": 16, "C": 16, "D": 10,
                 "E": 14, "F": 16, "G": 14, "H": 28})
    ws.freeze_panes = "A3"


def generate_excel(all_results, problems, run_meta, xlsx_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_summary_sheet(wb, all_results, run_meta)
    build_submissions_sheet(wb, all_results)
    build_tc_detail_sheet(wb, all_results)
    build_tc_heatmap_sheet(wb, all_results, problems)
    build_tc_passrate_sheet(wb, all_results, problems)
    build_latency_sheet(wb, all_results)

    wb.save(xlsx_path)
    print(f"\n  Excel report → {xlsx_path}")
    print(f"  Sheets: Summary | Student Submissions | TC Results (All) |"
          f" TC Heatmap | TC Pass Rate Analysis | Timing Analysis")


# ── Entry point ───────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--url",      default=API_BASE)
    parser.add_argument("--students", type=int, default=30)
    parser.add_argument("--concurrency", type=int, default=10,
                        help="Max concurrent submissions (default 10)")
    parser.add_argument("--bank",     default=BANK_FILE)
    parser.add_argument("--json-out", default="dp_test_results.json")
    parser.add_argument("--xlsx-out", default="Reports/dp_test_report.xlsx")
    args = parser.parse_args()

    # Load question bank
    bank_path = Path(args.bank)
    if not bank_path.exists():
        print(f"ERROR: {bank_path} not found.")
        sys.exit(1)
    with open(bank_path) as fh:
        bank = json.load(fh)
    problems = bank["problems"]

    # Assign students to problems + solutions
    roster = build_student_roster(problems, args.students)
    run_id = str(int(time.time()))

    print(f"\n  DP Load Test — {args.students} students × 4 problems (30 TCs each)")
    print(f"  API: {args.url}   |   Concurrency: {args.concurrency}")
    print(f"  Question bank: {args.bank}")
    print(f"  Run ID: {run_id}")
    print()

    # Print roster summary
    from collections import Counter
    type_dist = Counter(e["solution_type"] for e in roster)
    prob_dist = Counter(e["problem"]["id"] for e in roster)
    print("  Solution types:", dict(type_dist))
    print("  Problem distribution:", dict(prob_dist))
    print()

    # Run all students concurrently
    start_time = time.monotonic()
    all_results = []
    done_count  = [0]
    total       = len(roster)

    print(f"  Submitting {total} students (concurrency={args.concurrency})...")

    with concurrent.futures.ThreadPoolExecutor(max_workers=args.concurrency) as pool:
        futures = {
            pool.submit(run_student, entry, args.url, run_id): entry
            for entry in roster
        }
        for future in concurrent.futures.as_completed(futures):
            result = future.result()
            all_results.append(result)
            done_count[0] += 1
            ok  = result["score"] == result["total"] and result["total"] > 0
            err = result["error"] or result["system_error"]
            icon = "✓" if ok else ("!" if err else "~")
            print(f"  [{done_count[0]:2d}/{total}] {icon} {result['student_id']}"
                  f"  {result['problem_id']:<12} {result['score']:2d}/30"
                  f"  {result['wall_time_s']:.2f}s"
                  f"  [{result['solution_type']}]")

    total_duration = round(time.monotonic() - start_time, 1)
    total_tc_pass  = sum(s["score"] for s in all_results)
    total_tc_all   = sum(s["total"] for s in all_results)

    print(f"\n  ─────────────────────────────────────────────────────")
    print(f"  Total duration : {total_duration}s")
    print(f"  TCs evaluated  : {total_tc_all}")
    print(f"  TCs passed     : {total_tc_pass}/{total_tc_all} "
          f"({total_tc_pass/max(total_tc_all,1)*100:.1f}%)")
    print(f"  Full score     : "
          f"{sum(1 for s in all_results if s['score']==s['total'] and s['total']>0)} students")

    # Save JSON
    run_meta = {
        "run_at":             datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
        "run_id":             run_id,
        "api_url":            args.url,
        "n_students":         args.students,
        "concurrency":        args.concurrency,
        "problems":           [p["id"] for p in problems],
        "total_tcs_evaluated":total_tc_all,
        "total_tcs_passed":   total_tc_pass,
        "total_duration_s":   total_duration,
    }
    report = {"meta": run_meta, "students": all_results}

    # Remove tc_lookup (not JSON serializable cleanly)
    for s in all_results:
        s.pop("tc_lookup", None)

    json_path = Path(args.json_out)
    with open(json_path, "w") as fh:
        json.dump(report, fh, indent=2, default=str)
    print(f"\n  JSON results → {json_path}")

    # Generate Excel
    if not HAS_OPENPYXL:
        print("  WARNING: openpyxl not installed. Skipping Excel report.")
        print("  Run: pip install openpyxl")
        return

    xlsx_path = Path(args.xlsx_out)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    generate_excel(all_results, problems, run_meta, xlsx_path)

    try:
        import subprocess
        subprocess.Popen(["open", str(xlsx_path)])
    except Exception:
        pass


if __name__ == "__main__":
    main()
