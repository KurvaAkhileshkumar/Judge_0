#!/usr/bin/env python3
"""
generate_cumulative_report.py
──────────────────────────────
Generates a TRUE cumulative Excel report by merging all scenario JSONs from
Reports/Jsons/ into a single aggregated dataset.

Sheets:
  1 — System Info + per-run breakdown table (aggregated config)
  2 — Load Test Metrics  (aggregated across ALL runs)
  3 — User Test Case Results  (ALL submissions from ALL runs)
  4 — Scenario Comparison  (individual per-run side-by-side, unchanged)

Usage:
    python generate_cumulative_report.py
    python generate_cumulative_report.py --out Reports/cumulative_report.xlsx
"""

import argparse
import json
import math
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.  Run: pip install -r requirements-report.txt")
    sys.exit(1)

# Import all helpers + sheet builders from the existing generator
import generate_ec2_report as G

# ── Canonical run order (matches run_load_tests.sh) ───────────────────────────
CANONICAL_ORDER = [
    "A_100u_flat",
    "A_500u_ramp10",
    "A_1000u_ramp20",
    "A_1000u_spike",
    "A_1000u_pipe50",
    "A_1000u_pipe100",
    "B_1000u_ramp20",
    "B_1000u_spike",
    "B_1000u_pipe50",
    "C_1000u_ramp20",
    "C_1000u_spike",
    "C_1000u_pipe50",
    "D_500u_ramp10",
    "D_1000u_ramp20",
    "E_1000u_ramp20",
    "E_1000u_spike",
    "E_1000u_pipe50",
]

JSONS_DIR = Path("Reports/Jsons")


# ── Load ───────────────────────────────────────────────────────────────────────

def load_reports() -> list[tuple[str, dict]]:
    """Load canonical JSONs that exist. Returns [(stem, report_dict), ...]."""
    loaded = []
    for stem in CANONICAL_ORDER:
        path = JSONS_DIR / f"{stem}.json"
        if path.exists():
            with open(path, encoding="utf-8") as fh:
                rpt = json.load(fh)
            loaded.append((stem, rpt))
            comp = rpt.get("summary", {}).get("completed", 0)
            print(f"  loaded  : {stem}.json  ({comp} submissions)")
        else:
            print(f"  MISSING : {stem}.json — skipped")
    return loaded


# ── Merge ──────────────────────────────────────────────────────────────────────

def _percentile(sorted_vals: list, p: float) -> float:
    if not sorted_vals:
        return 0.0
    idx = max(0, math.ceil(len(sorted_vals) * p / 100) - 1)
    return sorted_vals[idx]


def merge_reports(named_reports: list[tuple[str, dict]]) -> dict:
    """Merge N report dicts into a single aggregate report dict for Sheets 1–3."""
    first_cfg = named_reports[0][1].get("config", {})

    # ── Sum scalar summary fields ──────────────────────────────────────────────
    SUM_KEYS = [
        "completed", "total_submissions", "errors", "blocked_by_security",
        "rate_limited_429", "duplicate_idempotent", "system_errors",
        "harness_jobs", "batch_equivalent",
    ]
    agg_summary: dict = defaultdict(int)
    total_duration = 0.0
    peak_in_flight = 0

    for _stem, rpt in named_reports:
        s = rpt.get("summary", {})
        for k in SUM_KEYS:
            agg_summary[k] += s.get(k, 0)
        tput = s.get("throughput_per_sec", 0) or 0
        comp = s.get("completed", 0) or 0
        total_duration += (comp / tput) if tput > 0 else 0.0
        peak_in_flight = max(peak_in_flight, s.get("peak_in_flight_jobs", 0))

    # ── Pool latencies → recompute real percentiles ────────────────────────────
    all_lats: list[float] = []
    for _stem, rpt in named_reports:
        for r in rpt.get("results", []):
            lat = r.get("latency_ms") if "latency_ms" in r else r.get("LatencyMs")
            if lat is not None:
                all_lats.append(lat)
    all_lats.sort()

    completed = agg_summary["completed"]
    throughput = (completed / total_duration) if total_duration > 0 else 0.0

    merged_summary = dict(agg_summary)
    merged_summary["throughput_per_sec"]  = throughput
    merged_summary["peak_in_flight_jobs"] = peak_in_flight
    merged_summary["latency_p50_ms"]      = _percentile(all_lats, 50)
    merged_summary["latency_p95_ms"]      = _percentile(all_lats, 95)
    merged_summary["latency_p99_ms"]      = _percentile(all_lats, 99)

    # ── Aggregate count dicts ──────────────────────────────────────────────────
    merged_status:   dict = defaultdict(int)
    merged_types:    dict = defaultdict(int)
    merged_problems: dict = defaultdict(int)

    for _stem, rpt in named_reports:
        for k, v in rpt.get("status_counts",  {}).items():
            merged_status[k]   += v
        for k, v in rpt.get("type_counts",    {}).items():
            merged_types[k]    += v
        for k, v in rpt.get("problem_counts", {}).items():
            merged_problems[k] += v

    # ── Flat results list (tagged with run label for traceability) ─────────────
    merged_results: list[dict] = []
    for stem, rpt in named_reports:
        for r in rpt.get("results", []):
            entry = dict(r)
            entry["_run_label"] = stem
            merged_results.append(entry)

    # ── per_problem_report — combine submissions per problem_id ───────────────
    prob_subs: dict[str, list] = defaultdict(list)
    for stem, rpt in named_reports:
        for pp in rpt.get("per_problem_report", []):
            pid = pp["problem_id"]
            for sub in pp.get("submissions", []):
                entry = dict(sub)
                entry["_run_label"] = stem
                prob_subs[pid].append(entry)

    merged_per_problem = [
        {"problem_id": pid, "submissions": subs}
        for pid, subs in sorted(prob_subs.items())
    ]

    # ── Synthetic config (signals cumulative mode to Sheet 1 builder) ─────────
    merged_config = {
        "_cumulative":      True,
        "_n_runs":          len(named_reports),
        "_run_labels":      [s for s, _ in named_reports],
        "_total_duration_s": total_duration,
        # keep first-run values for fields that Sheet 2 reads
        "mode":             first_cfg.get("mode",       "flask_stack"),
        "target":           first_cfg.get("target",     "N/A"),
        "users":            sum(r.get("config", {}).get("users", 0) for _, r in named_reports),
        "ramp_up_sec":      0,
        "batch_size":       0,
        "dry_run":          False,
        "workers":          first_cfg.get("workers",    3),
        "runners":          first_cfg.get("runners",    2),
    }

    return {
        "config":             merged_config,
        "summary":            merged_summary,
        "status_counts":      dict(merged_status),
        "type_counts":        dict(merged_types),
        "problem_counts":     dict(merged_problems),
        "results":            merged_results,
        "per_problem_report": merged_per_problem,
        "batches":            [],
        "_workers":           first_cfg.get("workers", 3),
        "_runners":           first_cfg.get("runners", 2),
    }


# ── Sheet 1 (cumulative) ───────────────────────────────────────────────────────

def build_sheet1_cumulative(wb: openpyxl.Workbook, merged: dict,
                             named_reports: list[tuple[str, dict]]):
    """System Info sheet with aggregate header + per-run breakdown table."""
    ws = wb.active
    ws.title = "System Info"
    ws.sheet_view.showGridLines = False

    cfg     = merged["config"]
    summary = merged["summary"]
    n_runs  = cfg["_n_runs"]
    w       = cfg.get("workers", 3)
    r       = cfg.get("runners", 2)
    NCOLS   = 2

    # ── Title ─────────────────────────────────────────────────────────────────
    G._write_title(
        ws, 1,
        f"EC2 LOAD TEST — CUMULATIVE REPORT  ({n_runs} scenarios aggregated)",
        NCOLS, fontsize=14,
    )
    c2 = ws.cell(
        row=2, column=1,
        value=f"Generated: {G._now_utc()} UTC  |  {n_runs} scenario runs merged",
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    G._style(c2, bg=G._C["light_grey"], halign="center", bold=False, size=10)

    row = 3
    alt = False

    def section(title: str):
        nonlocal row, alt
        row += 1
        G._write_section(ws, row, title, NCOLS)
        row += 1
        alt = False

    def kv(key: str, value):
        nonlocal row, alt
        G._write_kv(ws, row, key, value, alt)
        row += 1
        alt = not alt

    # ── EC2 / Docker / Judge0 (static — same across all runs) ─────────────────
    section("EC2 Instance")
    kv("Instance Type",               "c5.xlarge")
    kv("vCPUs",                       4)
    kv("RAM",                         "8 GB")
    kv("Storage",                     "EBS (gp2 / gp3)")
    kv("OS",                          "Ubuntu 22.04.5 LTS")
    kv("Kernel",                      "Linux 5.15+ x86_64")
    kv("Region",                      "ap-south-1 (Mumbai)")
    kv("Network Performance",         "Up to 10 Gbps")

    section("Docker Services Layout")
    kv("Judge0 API Server (server)",  "judge0/judge0:1.13.1  —  port 2358")
    kv("Judge0 Worker containers",    f"{w} containers  (--scale workers={w})")
    kv("MAX_RUNNERS per worker",      r)
    kv("Total concurrent sandboxes",  f"{w} × {r} = {w * r}")
    kv("Flask API (api)",             "gunicorn+gevent  —  port 5001→5000")
    kv("Grading Worker",              "worker_async.py  —  WORKER_CONCURRENCY=24")
    kv("Reconciler",                  "reconciler.py  —  scan every 60s")
    kv("PostgreSQL",                  "postgres:13  —  mem_limit=400m  reservation=128m")
    kv("Redis",                       "Latest  —  mem_limit=512m  maxmemory=512mb")

    section("Judge0 Configuration (judge0.ec2.conf)")
    kv("MAX_RUNNERS (isolate sandboxes/worker)", r)
    kv("MAX_QUEUE_SIZE (Resque)",     200)
    kv("CPU_TIME_LIMIT (default)",    "5 s")
    kv("WALL_TIME_LIMIT (default)",   "10 s")
    kv("MEMORY_LIMIT (default)",      "256 000 KB  (256 MB)")
    kv("MAX_MEMORY_LIMIT",            "4 194 304 KB  (4 GB)")
    kv("MAX_PROCESSES_AND_OR_THREADS", 220)
    kv("INTERVAL (Resque idle poll)", "0.1 s")

    section("Judge0 API Server (Puma / Rails)")
    kv("WEB_CONCURRENCY (Puma processes)",    2)
    kv("RAILS_MAX_THREADS (threads/process)", 16)
    kv("HTTP slots",                          "2 × 16 = 32 concurrent HTTP requests")

    section("Flask Grading API (gunicorn + gevent)")
    kv("Gunicorn workers",            2)
    kv("Worker connections",          500)
    kv("Bind",                        "0.0.0.0:5000  (host port 5001)")
    kv("Result delivery",             "SSE  /results/stream/<ticket_id>")
    kv("MAX_QUEUE_DEPTH",             5000)
    kv("Idempotency TTL",             "2 h  (SHA-256 key stored in Redis)")

    section("Grading Worker (worker_async.py)")
    kv("WORKER_CONCURRENCY",          24)
    kv("MAX_RETRY_COUNT",             3)

    section("PostgreSQL 13")
    kv("max_connections",             200)
    kv("shared_buffers",              "128 MB")
    kv("mem_limit",                   "400 m")

    section("Redis")
    kv("maxmemory",                   "512 mb")
    kv("mem_limit",                   "512 m")

    # ── Cumulative aggregate summary ───────────────────────────────────────────
    total_dur = cfg.get("_total_duration_s", 0)
    section(f"Cumulative Aggregate  —  {n_runs} Scenarios")
    kv("Total Scenarios Run",         n_runs)
    kv("Target URL",                  cfg.get("target", "N/A"))
    kv("Mode",                        cfg.get("mode",   "N/A"))
    kv("Total Finalized (all outcomes)", summary.get("completed",           0))
    kv("PASS Count (all runs)",          merged.get("status_counts",{}).get("PASS", 0))
    kv("Total Errors",                summary.get("errors",             0))
    kv("Total Blocked",               summary.get("blocked_by_security",0))
    kv("Total Rate-Limited (429)",    summary.get("rate_limited_429",   0))
    kv("Total System Errors",         summary.get("system_errors",      0))
    kv("Sum of Run Durations",        f"{total_dur:.1f} s  (sequential runs, not wall-clock)")
    kv("Overall Throughput",
       f"{summary.get('throughput_per_sec', 0):.3f} req/s  (total_completed / sum_of_durations)")
    kv("Aggregated P50 Latency",      f"{summary.get('latency_p50_ms', 0):.0f} ms")
    kv("Aggregated P95 Latency",      f"{summary.get('latency_p95_ms', 0):.0f} ms")
    kv("Aggregated P99 Latency",      f"{summary.get('latency_p99_ms', 0):.0f} ms")

    # ── Per-run breakdown table ────────────────────────────────────────────────
    row += 1
    G._write_section(ws, row, f"Per-Run Breakdown  ({n_runs} runs)", NCOLS)
    row += 1

    # Column headers (spans both cols via wider col B)
    c1 = ws.cell(row=row, column=1, value="Run / Scenario")
    c2 = ws.cell(row=row, column=2,
                 value="Users  |  Mode  |  Ramp  |  Batch  |  Completed  |  "
                        "Throughput (req/s)  |  P50 (ms)  |  P95 (ms)  |  P99 (ms)  |  PASS  |  FAIL  |  TLE")
    G._style(c1, bg=G._C["hdr_light"], fg=G._C["hdr_dark"], bold=True)
    G._style(c2, bg=G._C["hdr_light"], fg=G._C["hdr_dark"], bold=True)
    row += 1

    for i, (stem, rpt) in enumerate(named_reports):
        rc  = rpt.get("config",  {})
        rs  = rpt.get("summary", {})
        sc  = rpt.get("status_counts", {})
        bg  = G._C["alt_row"] if i % 2 else G._C["white"]
        tput = rs.get("throughput_per_sec", 0) or 0

        detail = (
            f"users={rc.get('users')}  "
            f"mode={rc.get('mode')}  "
            f"ramp={rc.get('ramp_up_sec', 0)}s  "
            f"batch={rc.get('batch_size', 0) or 'none'}  "
            f"completed={rs.get('completed', 0)}  "
            f"tput={tput:.2f} req/s  "
            f"P50={rs.get('latency_p50_ms', 0)} ms  "
            f"P95={rs.get('latency_p95_ms', 0)} ms  "
            f"P99={rs.get('latency_p99_ms', 0)} ms  "
            f"PASS={sc.get('PASS', 0)}  "
            f"FAIL={sc.get('FAIL', 0)}  "
            f"TLE={sc.get('TLE', 0)}"
        )
        c1 = ws.cell(row=row, column=1, value=stem)
        G._style(c1, bg=bg, bold=True)
        c2 = ws.cell(row=row, column=2, value=detail)
        G._style(c2, bg=bg)
        row += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = ws.cell(row=3, column=1)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate a true cumulative load test report from all scenario JSONs"
    )
    parser.add_argument(
        "--out",
        default="Reports/cumulative_report.xlsx",
        help="Output .xlsx path (default: Reports/cumulative_report.xlsx)",
    )
    args = parser.parse_args()

    print("=" * 60)
    print("  Cumulative Load Test Report Generator")
    print("=" * 60)

    print(f"\nScanning {JSONS_DIR} ...")
    named_reports = load_reports()
    if len(named_reports) < 2:
        print("ERROR: Need at least 2 scenario JSON files in Reports/Jsons/")
        sys.exit(1)

    print(f"\nMerging {len(named_reports)} reports...")
    merged = merge_reports(named_reports)

    s = merged["summary"]
    print(f"  total completed      : {s['completed']}")
    print(f"  total submissions    : {s['total_submissions']}")
    print(f"  pooled results       : {len(merged['results'])}")
    print(f"  overall throughput   : {s['throughput_per_sec']:.3f} req/s")
    print(f"  P50 / P95 / P99 (ms) : {s['latency_p50_ms']:.0f} / "
          f"{s['latency_p95_ms']:.0f} / {s['latency_p99_ms']:.0f}")
    print(f"  status counts        : {merged['status_counts']}")

    print("\nBuilding Excel workbook...")
    wb = openpyxl.Workbook()

    # Sheet 1 — cumulative system info + per-run breakdown
    build_sheet1_cumulative(wb, merged, named_reports)
    print("  Sheet 1 done")

    # Sheet 2 — aggregated metrics: load all per-run metrics jsonl files
    all_metrics: list[dict] = []
    for stem in [s for s, _ in named_reports]:
        mpath = JSONS_DIR / f"{stem}_metrics.jsonl"
        if mpath.exists():
            count_before = len(all_metrics)
            with open(mpath, encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if line:
                        try:
                            all_metrics.append(json.loads(line))
                        except json.JSONDecodeError:
                            pass
            print(f"  Metrics loaded : {mpath}  ({len(all_metrics) - count_before} snapshots)")
    G.build_sheet2(wb, merged, metrics_data=all_metrics)
    print("  Sheet 2 done")

    # Sheet 3 — all submissions from all runs
    G.build_sheet3(wb, merged)
    print("  Sheet 3 done")

    # Sheet 4 — individual scenario comparison (unchanged)
    all_for_compare = [
        (str(JSONS_DIR / f"{stem}.json"), rpt)
        for stem, rpt in named_reports
    ]
    G.build_sheet4_comparison(wb, all_for_compare)
    print("  Sheet 4 done")

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.active = wb.worksheets[0]
    wb.save(out_path)

    n_subs = len(merged["results"])
    print(f"\n✓  Saved: {out_path}")
    print(f"   Sheet 1 — System Info + {len(named_reports)}-run breakdown  (aggregated)")
    print(f"   Sheet 2 — Load Test Metrics  ({s['completed']} total completed submissions)")
    print(f"   Sheet 3 — User Test Case Results  ({n_subs} submissions across all runs)")
    print(f"   Sheet 4 — Scenario Comparison  ({len(named_reports)} runs side-by-side)")
    print("=" * 60)


if __name__ == "__main__":
    main()
